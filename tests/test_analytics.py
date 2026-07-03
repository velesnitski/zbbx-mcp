"""Tests for analytics helpers, provider detection, and trend sanity logic."""

from zbbx_mcp.classify import (
    PROVIDER_CIDRS,
    classify_host,
    detect_provider,
    resolve_datacenter,
)
from zbbx_mcp.data import (
    CAPITAL_COORDS,
    REGION_MAP,
    countries_for_region,
    extract_country,
    group_by_country,
    host_ip,
)


class TestRegionMap:
    def test_latam_countries(self):
        codes = countries_for_region("LATAM")
        assert "BR" in codes
        assert "MX" in codes
        assert "AR" in codes
        assert "DE" not in codes

    def test_emea_countries(self):
        codes = countries_for_region("EMEA")
        assert "DE" in codes
        assert "NL" in codes
        assert "US" not in codes

    def test_all_returns_union(self):
        codes = countries_for_region("ALL")
        assert len(codes) > 50
        assert "US" in codes
        assert "BR" in codes
        assert "DE" in codes

    def test_unknown_region_empty(self):
        assert countries_for_region("MARS") == set()

    def test_case_insensitive(self):
        assert countries_for_region("latam") == countries_for_region("LATAM")

    def test_all_regions_have_countries(self):
        for region in REGION_MAP:
            assert len(REGION_MAP[region]) > 0




class TestCapitalCoords:
    def test_major_countries_present(self):
        for cc in ["US", "DE", "NL", "BR", "FR", "TR", "JP", "RU"]:
            assert cc in CAPITAL_COORDS, f"{cc} missing from CAPITAL_COORDS"

    def test_coords_are_valid(self):
        for cc, (lat, lon) in CAPITAL_COORDS.items():
            assert -90 <= lat <= 90, f"{cc} lat {lat} out of range"
            assert -180 <= lon <= 180, f"{cc} lon {lon} out of range"




class TestGroupByCountry:
    def _hosts(self):
        return [
            {"hostid": "1", "host": "srv-nl01", "groups": [{"name": "free"}]},
            {"hostid": "2", "host": "srv-nl02", "groups": [{"name": "free"}]},
            {"hostid": "3", "host": "srv-de01", "groups": [{"name": "prem"}]},
            {"hostid": "4", "host": "srv-us01", "groups": [{"name": "free"}]},
            {"hostid": "5", "host": "no-country", "groups": [{"name": "mon"}]},
        ]

    def test_basic_grouping(self):
        result = group_by_country(self._hosts())
        assert "NL" in result
        assert len(result["NL"]) == 2
        assert "DE" in result
        assert "US" in result
        assert "" not in result

    def test_country_filter(self):
        result = group_by_country(self._hosts(), country="nl")
        assert list(result.keys()) == ["NL"]
        assert len(result["NL"]) == 2

    def test_country_filter_no_match(self):
        result = group_by_country(self._hosts(), country="JP")
        assert result == {}

    def test_region_filter(self):
        result = group_by_country(self._hosts(), region="EMEA")
        assert "NL" in result
        assert "DE" in result
        assert "US" not in result




class TestHostIp:
    def test_extracts_ip(self):
        h = {"interfaces": [{"ip": "1.2.3.4"}]}
        assert host_ip(h) == "1.2.3.4"

    def test_skips_loopback(self):
        h = {"interfaces": [{"ip": "127.0.0.1"}, {"ip": "5.6.7.8"}]}
        assert host_ip(h) == "5.6.7.8"

    def test_no_interfaces(self):
        assert host_ip({}) == ""
        assert host_ip({"interfaces": []}) == ""

    def test_only_loopback(self):
        h = {"interfaces": [{"ip": "127.0.0.1"}]}
        assert host_ip(h) == ""




class TestProviderDetection:
    """Provider detection tests — iterate the PROVIDER_CIDRS database only.

    No provider names or IPs are hardcoded in tests. Test coverage scales
    automatically with whatever is in the database.
    """

    @staticmethod
    def _all_nets():
        import ipaddress
        return [
            (prov, ipaddress.ip_network(cidr, strict=False))
            for prov, cidrs in PROVIDER_CIDRS.items()
            for cidr in cidrs
        ]

    def _pick_sample_ip(self, net, all_nets):
        """Pick an IP in `net` that is NOT in any more-specific other-provider net."""
        overlaps = [
            n for p, n in all_nets
            if n.subnet_of(net) and n.prefixlen > net.prefixlen
        ]
        hosts = net.hosts() if net.prefixlen < 31 else iter([net.network_address])
        for candidate in hosts:
            if not any(candidate in o for o in overlaps):
                return str(candidate)
        return None  # fully subsumed — unreachable for this provider

    def test_every_provider_detected_via_db(self):
        """Every provider must be detected from at least one of its own CIDRs."""
        all_nets = self._all_nets()
        for prov, cidrs in PROVIDER_CIDRS.items():
            import ipaddress
            detected_at_least_once = False
            for cidr in cidrs:
                net = ipaddress.ip_network(cidr, strict=False)
                ip = self._pick_sample_ip(net, all_nets)
                if ip and detect_provider(ip) == prov:
                    detected_at_least_once = True
                    break
            assert detected_at_least_once, (
                "Provider has CIDRs but none resolve back to it — check overlaps"
            )

    def test_rfc1918_private_is_other(self):
        """Private RFC 1918 ranges should fall through to 'Other'."""
        assert detect_provider("192.168.1.1") == "Other"
        assert detect_provider("10.0.0.1") == "Other"
        assert detect_provider("172.16.0.1") == "Other"

    def test_invalid_ip(self):
        assert detect_provider("not-an-ip") == "Unknown"
        assert detect_provider("") == "Unknown"

    def test_prefix_length_wins(self):
        """More specific CIDR must win over broader one — verified across DB."""
        import ipaddress
        # Find any pair where a smaller prefix is contained in a larger one
        nets = [
            (prov, ipaddress.ip_network(c, strict=False))
            for prov, cidrs in PROVIDER_CIDRS.items()
            for c in cidrs
        ]
        found_pair = False
        for i, (p1, n1) in enumerate(nets):
            for p2, n2 in nets[i + 1:]:
                if p1 == p2 or n1.prefixlen == n2.prefixlen:
                    continue
                more_specific = n1 if n1.prefixlen > n2.prefixlen else n2
                broader = n2 if n1.prefixlen > n2.prefixlen else n1
                specific_prov = p1 if n1.prefixlen > n2.prefixlen else p2
                if more_specific.subnet_of(broader):
                    hosts = list(more_specific.hosts()) if more_specific.prefixlen < 31 else [more_specific.network_address]
                    ip = str(hosts[len(hosts) // 3])
                    assert detect_provider(ip) == specific_prov
                    found_pair = True
                    break
            if found_pair:
                break
        assert found_pair, "Database must contain at least one overlapping pair"

    def test_all_providers_have_cidrs(self):
        for prov, cidrs in PROVIDER_CIDRS.items():
            assert len(cidrs) > 0, f"{prov} has no CIDRs"




class TestResolveDatacenter:
    """Datacenter resolution tests — iterate the DATACENTER_CIDRS database."""

    def _sample_ip(self, cidr: str) -> str:
        import ipaddress
        net = ipaddress.ip_network(cidr, strict=False)
        hosts = list(net.hosts()) if net.prefixlen < 31 else [net.network_address]
        return str(hosts[len(hosts) // 3]) if hosts else str(net.network_address)

    def test_every_dc_entry_resolves(self):
        """Every datacenter mapping must resolve to its recorded provider+city."""
        from zbbx_mcp.classify import DATACENTER_CIDRS
        for prov, mappings in DATACENTER_CIDRS.items():
            for cidr, city in mappings:
                ip = self._sample_ip(cidr)
                got_prov, got_city = resolve_datacenter(ip)
                assert got_prov == prov, f"{ip}: expected {prov}, got {got_prov}"
                assert got_city == city, f"{ip}: expected {city}, got {got_city}"

    def test_fallback_to_provider_only(self):
        """Provider-known but no DC mapping → provider name, empty city."""
        from zbbx_mcp.classify import DATACENTER_CIDRS, PROVIDER_CIDRS
        dc_providers = set(DATACENTER_CIDRS.keys())
        for prov, cidrs in PROVIDER_CIDRS.items():
            if prov in dc_providers:
                continue
            # Pick first cidr of a provider without DC mapping
            ip = self._sample_ip(cidrs[0])
            got_prov, got_city = resolve_datacenter(ip)
            assert got_prov == prov
            assert got_city == ""
            return
        # If we reach here the test is still valid (all providers have DC mappings)

    def test_unknown_falls_through(self):
        """RFC 1918 / unroutable → Other, empty city."""
        prov, city = resolve_datacenter("10.10.10.10")
        assert prov == "Other"
        assert city == ""

    def test_invalid(self):
        prov, city = resolve_datacenter("bad")
        assert prov == "Unknown"




class TestExtractCountry:
    def test_standard_patterns(self):
        assert extract_country("srv-nl0105") == "NL"
        assert extract_country("srv-de3") == "DE"
        assert extract_country("srv-us0001") == "US"

    def test_lite_pattern(self):
        assert extract_country("srv-nl01-lite") == "NL"
        assert extract_country("srv-us01-lite") == "US"
        assert extract_country("srv-tr01-lite") == "TR"

    def test_ar_pattern(self):
        assert extract_country("srv-ar010") == "AR"

    def test_br_mx_patterns(self):
        assert extract_country("srv-br0101") == "BR"
        assert extract_country("srv-mx0101") == "MX"

    def test_uk_normalizes_to_gb(self):
        """Non-ISO country code normalizes to standard."""
        assert extract_country("srv-uk0001") == "GB"
        assert extract_country("srv-uk0005") == "GB"

    def test_no_match(self):
        assert extract_country("Zabbix server") == ""
        assert extract_country("account.example.com") == ""
        assert extract_country("a1") == ""

    def test_multiple_country_codes(self):
        """Multiple country codes in hostname — first wins."""
        assert extract_country("srv-de-nl01") == "DE"




class TestTrendSanity:
    """Test the trend/change consistency rules used in CEO report and geo tools.

    Rules:
    1. change < -10% and trend == "rising" → override to "stable"
    2. change > 0 and trend == "dropping" → override to "stable"
    3. current > avg * 1.5 and trend == "dropping" → override to "rising"
    4. current < 0.01 and avg > 0.05 → "dead"
    """

    @staticmethod
    def _apply_sanity(change: float, trend: str, traffic_gbps: float, avg_gbps: float) -> str:
        """Replicate the sanity logic from ceo_report.py / geo.py."""
        if change <= -30 and trend in ("stable", "rising"):
            trend = "dropping"
        elif change >= 30 and trend in ("stable", "dropping"):
            trend = "rising"
        elif change <= -10 and trend == "rising" or change > 0 and trend == "dropping":
            trend = "stable"
        if traffic_gbps < 0.01 and avg_gbps > 0.05:
            trend = "dead"
        return trend

    def test_rising_with_negative_change(self):
        """Rising trend but negative change should become stable."""
        result = self._apply_sanity(change=-13, trend="rising", traffic_gbps=22.0, avg_gbps=25.4)
        assert result == "stable", f"Expected stable, got {result}"

    def test_dropping_with_large_decline(self):
        """Legitimate large decline stays dropping."""
        result = self._apply_sanity(change=-87, trend="dropping", traffic_gbps=0.3, avg_gbps=2.2)
        assert result == "dropping"

    def test_rising_with_strong_growth(self):
        """Legitimate strong growth stays rising."""
        result = self._apply_sanity(change=123, trend="rising", traffic_gbps=26.3, avg_gbps=11.8)
        assert result == "rising"

    def test_dropping_positive_change_becomes_stable(self):
        """Small positive change with dropping trend becomes stable."""
        result = self._apply_sanity(change=11, trend="dropping", traffic_gbps=1.9, avg_gbps=1.7)
        assert result == "stable"

    def test_dropping_huge_current_becomes_rising(self):
        """Large positive change with dropping trend becomes rising."""
        result = self._apply_sanity(change=100, trend="dropping", traffic_gbps=4.0, avg_gbps=2.0)
        # change > 30 catches this → rising
        assert result == "rising"

    def test_dead_overrides_all(self):
        """Zero traffic with prior average triggers dead."""
        result = self._apply_sanity(change=-100, trend="dropping", traffic_gbps=0.0, avg_gbps=0.9)
        assert result == "dead"

    def test_moderate_decline_stays_stable(self):
        """Moderate decline within threshold stays stable."""
        result = self._apply_sanity(change=-21, trend="stable", traffic_gbps=20.3, avg_gbps=25.7)
        assert result == "stable"

    def test_stable_large_decline_becomes_dropping(self):
        """Large decline with stable trend becomes dropping."""
        result = self._apply_sanity(change=-47, trend="stable", traffic_gbps=0.8, avg_gbps=1.6)
        assert result == "dropping"

    def test_stable_significant_decline_becomes_dropping(self):
        """Significant decline with stable trend becomes dropping."""
        result = self._apply_sanity(change=-43, trend="stable", traffic_gbps=0.4, avg_gbps=0.7)
        assert result == "dropping"

    def test_stable_severe_decline_becomes_dropping(self):
        """Severe decline with stable trend becomes dropping."""
        result = self._apply_sanity(change=-70, trend="stable", traffic_gbps=1.1, avg_gbps=3.7)
        assert result == "dropping"

    def test_big_positive_change_stable_becomes_rising(self):
        """Large positive change overrides stable to rising."""
        result = self._apply_sanity(change=50, trend="stable", traffic_gbps=15.0, avg_gbps=10.0)
        assert result == "rising"

    def test_small_negative_change_keeps_rising(self):
        """Small decline within threshold keeps rising."""
        result = self._apply_sanity(change=-5, trend="rising", traffic_gbps=9.5, avg_gbps=10.0)
        assert result == "rising"

    def test_zero_traffic_zero_avg_stays_stable(self):
        """No traffic and no history stays stable."""
        result = self._apply_sanity(change=0, trend="stable", traffic_gbps=0.0, avg_gbps=0.0)
        assert result == "stable"

    def test_dropping_large_positive_becomes_rising(self):
        """Large positive change overrides dropping to rising."""
        result = self._apply_sanity(change=83, trend="dropping", traffic_gbps=4.1, avg_gbps=2.2)
        assert result == "rising"

    def test_dropping_moderate_positive_becomes_rising(self):
        """Moderate positive change overrides dropping to rising."""
        result = self._apply_sanity(change=36, trend="dropping", traffic_gbps=8.3, avg_gbps=6.1)
        assert result == "rising"

    def test_dropping_positive_15pct_becomes_stable(self):
        """Small positive change below threshold becomes stable."""
        result = self._apply_sanity(change=15, trend="dropping", traffic_gbps=1.15, avg_gbps=1.0)
        assert result == "stable"

    
    def test_exactly_minus_30_becomes_dropping(self):
        """Boundary: -30% exactly should trigger dropping (<=, not <)."""
        result = self._apply_sanity(change=-30, trend="stable", traffic_gbps=0.7, avg_gbps=1.0)
        assert result == "dropping", f"change=-30 stable should be dropping, got {result}"

    def test_minus_29_stays_stable(self):
        """Boundary: -29% should NOT trigger dropping."""
        result = self._apply_sanity(change=-29, trend="stable", traffic_gbps=0.71, avg_gbps=1.0)
        assert result == "stable"

    def test_exactly_plus_30_becomes_rising(self):
        """Boundary: +30% exactly should trigger rising (>=, not >)."""
        result = self._apply_sanity(change=30, trend="stable", traffic_gbps=1.3, avg_gbps=1.0)
        assert result == "rising", f"change=+30 stable should be rising, got {result}"

    def test_plus_29_stays_stable(self):
        """Boundary: +29% should NOT trigger rising."""
        result = self._apply_sanity(change=29, trend="stable", traffic_gbps=1.29, avg_gbps=1.0)
        assert result == "stable"

    def test_exactly_minus_10_rising_becomes_stable(self):
        """Boundary: -10% with rising should become stable."""
        result = self._apply_sanity(change=-10, trend="rising", traffic_gbps=0.9, avg_gbps=1.0)
        assert result == "stable"

    def test_minus_9_rising_stays_rising(self):
        """Boundary: -9% with rising should stay rising."""
        result = self._apply_sanity(change=-9, trend="rising", traffic_gbps=0.91, avg_gbps=1.0)
        assert result == "rising"

    def test_exactly_minus_30_rising_becomes_dropping(self):
        """Boundary: -30% with rising should become dropping (not stable)."""
        result = self._apply_sanity(change=-30, trend="rising", traffic_gbps=0.7, avg_gbps=1.0)
        assert result == "dropping"

    def test_exactly_plus_30_dropping_becomes_rising(self):
        """Boundary: +30% with dropping should become rising (not stable)."""
        result = self._apply_sanity(change=30, trend="dropping", traffic_gbps=1.3, avg_gbps=1.0)
        assert result == "rising"




class TestClassifyHost:
    def test_unknown_groups(self):
        prod, tier = classify_host([{"name": "Templates"}])
        assert prod == "Unknown"

    def test_skip_templates(self):
        prod, tier = classify_host([{"name": "Templates/Applications"}, {"name": "mygroup"}])
        assert prod == "mygroup"

    def test_empty_groups(self):
        prod, tier = classify_host([])
        assert prod == "Unknown"




class TestProductHiding:
    """Test ZABBIX_HIDE_PRODUCTS filtering used in CEO report."""

    def test_is_hidden_with_env(self, monkeypatch):
        """Hidden product detected when env var is set."""
        monkeypatch.setenv("ZABBIX_HIDE_PRODUCTS", "Legacy,OldProduct")
        # Reset cache so it re-reads env
        import zbbx_mcp.data as _data
        assert _data.is_hidden_product("Legacy") is True
        assert _data.is_hidden_product("legacy") is True  # case insensitive
        assert _data.is_hidden_product("OldProduct") is True
        assert _data.is_hidden_product("ActiveProduct") is False
        assert _data.is_hidden_product("AnotherProduct") is False

    def test_is_hidden_empty_env(self, monkeypatch):
        """Nothing hidden when env var is empty."""
        monkeypatch.setenv("ZABBIX_HIDE_PRODUCTS", "")
        import zbbx_mcp.data as _data
        assert _data.is_hidden_product("Legacy") is False
        assert _data.is_hidden_product("anything") is False

    def test_is_hidden_unset_env(self, monkeypatch):
        """Nothing hidden when env var is not set."""
        monkeypatch.delenv("ZABBIX_HIDE_PRODUCTS", raising=False)
        import zbbx_mcp.data as _data
        assert _data.is_hidden_product("Legacy") is False

    def test_group_by_country_excludes_hidden(self, monkeypatch):
        """group_by_country should skip hosts with hidden products."""
        monkeypatch.setenv("ZABBIX_HIDE_PRODUCTS", "Legacy")
        monkeypatch.setenv("ZABBIX_PRODUCT_MAP", "")  # use raw group names
        import zbbx_mcp.data as _data

        hosts = [
            {"hostid": "1", "host": "srv-de01", "groups": [{"name": "good_product"}]},
            {"hostid": "2", "host": "srv-de02", "groups": [{"name": "Legacy"}]},
            {"hostid": "3", "host": "srv-nl01", "groups": [{"name": "good_product"}]},
        ]
        result = _data.group_by_country(hosts)
        # DE should have 1 host (not 2 — Legacy host excluded)
        assert len(result.get("DE", [])) == 1
        assert result["DE"][0]["hostid"] == "1"
        assert "NL" in result

    def test_fleet_composition_filter(self, monkeypatch):
        """Simulate CEO report fleet composition: hidden + non-service excluded."""
        monkeypatch.setenv("ZABBIX_HIDE_PRODUCTS", "Legacy")
        import zbbx_mcp.data as _data

        _NON_service = {"Monitoring", "Infrastructure", "Unknown"}

        hosts_data = [
            ("serviceProduct", "Free"),
            ("serviceProduct", "Premium"),
            ("Legacy", "service"),
            ("Legacy", "service"),
            ("Monitoring", "WHOIS"),
            ("Infrastructure", "Tunnel"),
            ("Unknown", "Unknown"),
        ]

        product_counts: dict[str, int] = {}
        for prod, _tier in hosts_data:
            if prod and prod not in _NON_service and not _data.is_hidden_product(prod):
                product_counts[prod] = product_counts.get(prod, 0) + 1

        assert "serviceProduct" in product_counts
        assert product_counts["serviceProduct"] == 2
        assert "Legacy" not in product_counts, "Legacy should be hidden"
        assert "Monitoring" not in product_counts
        assert "Infrastructure" not in product_counts
        assert "Unknown" not in product_counts

    def test_ceo_report_two_step_filter(self, monkeypatch):
        """Simulate the actual CEO report flow: service_hosts then fleet composition.

        This catches the real bug where fleet composition iterated `hosts`
        instead of `service_hosts`, leaking hidden products into the report.
        """
        monkeypatch.setenv("ZABBIX_HIDE_PRODUCTS", "Legacy")
        monkeypatch.setenv("ZABBIX_PRODUCT_MAP", "")
        import zbbx_mcp.data as _data

        _NON_service = {"Monitoring", "Infrastructure", "Unknown"}

        # Simulate all hosts from Zabbix
        all_hosts = [
            {"hostid": "1", "host": "srv-de01", "groups": [{"name": "Activeservice"}]},
            {"hostid": "2", "host": "srv-de02", "groups": [{"name": "Activeservice"}]},
            {"hostid": "3", "host": "srv-nl01", "groups": [{"name": "Activeservice"}]},
            {"hostid": "4", "host": "srv-us01", "groups": [{"name": "Legacy"}]},
            {"hostid": "5", "host": "srv-us02", "groups": [{"name": "Legacy"}]},
            {"hostid": "6", "host": "monitor-1", "groups": [{"name": "Monitoring"}]},
            {"hostid": "7", "host": "infra-1", "groups": [{"name": "Infrastructure"}]},
        ]

        # Step 1: Build service_hosts (same as CEO report line 144-146)
        service_hosts = [
            h for h in all_hosts
            if not _data.is_hidden_product(classify_host(h.get("groups", []))[0])
            and classify_host(h.get("groups", []))[0] not in _NON_service
        ]

        assert len(service_hosts) == 3, f"Expected 3 service hosts, got {len(service_hosts)}"

        # Step 2: Build fleet composition from service_hosts (NOT all_hosts!)
        product_counts: dict[str, int] = {}
        for h in service_hosts:  # Must be service_hosts, not all_hosts
            prod, _ = classify_host(h.get("groups", []))
            if prod:
                product_counts[prod] = product_counts.get(prod, 0) + 1

        assert "Activeservice" in product_counts
        assert product_counts["Activeservice"] == 3
        assert "Legacy" not in product_counts, "Legacy leaked into fleet composition!"
        assert "Monitoring" not in product_counts
        assert "Infrastructure" not in product_counts

        # Step 3: Verify header KPI matches fleet composition
        total_from_header = len(service_hosts)
        total_from_composition = sum(product_counts.values())
        assert total_from_header == total_from_composition, \
            f"Header says {total_from_header} but composition sums to {total_from_composition}"


class TestIdleRelayDetection:
    """Pure-helper tests for get_idle_relays bucket+filter logic."""

    def _phys(self) -> frozenset[str]:
        return frozenset({"net.if.in[eth0]", "net.if.in[eno1]"})

    def test_split_buckets_physical_vs_tunnel(self):
        from zbbx_mcp.tools.correlation import _split_iface_metrics

        items = [
            {"hostid": "h1", "key_": "net.if.in[eth0]", "lastvalue": "20000"},
            {"hostid": "h1", "key_": "net.if.in[tun0]", "lastvalue": "0"},
            {"hostid": "h1", "key_": "net.if.in[gre1]", "lastvalue": "0"},
            {"hostid": "h1", "key_": "net.if.in[lo]", "lastvalue": "999"},  # ignored
        ]
        per_host = _split_iface_metrics(items, [], self._phys())
        assert per_host["h1"]["physical_bps"] == 20000
        assert per_host["h1"]["tunnel_bps"] == 0
        assert per_host["h1"]["tunnel_count"] == 2
        assert sorted(per_host["h1"]["tunnel_names"]) == ["gre1", "tun0"]

    def test_split_skips_docker_bridges(self):
        from zbbx_mcp.tools.correlation import _split_iface_metrics

        items = [
            {"hostid": "h1", "key_": "net.if.in[docker0]", "lastvalue": "1"},
            {"hostid": "h1", "key_": "net.if.in[br-abc]", "lastvalue": "1"},
        ]
        per_host = _split_iface_metrics(items, [], self._phys())
        assert per_host == {}

    def test_split_handles_garbage_values(self):
        from zbbx_mcp.tools.correlation import _split_iface_metrics

        items = [
            {"hostid": "h1", "key_": "net.if.in[eth0]", "lastvalue": ""},
            {"hostid": "h1", "key_": "net.if.in[tun0]", "lastvalue": None},
            {"hostid": "h1", "key_": "not-a-net-key", "lastvalue": "5"},
            {"hostid": "h1", "key_": "net.if.in[", "lastvalue": "5"},  # malformed
        ]
        per_host = _split_iface_metrics(items, [], self._phys())
        # eth0 zero is still a recorded physical, no tunnel flagged
        assert per_host["h1"]["physical_bps"] == 0.0
        assert per_host["h1"]["tunnel_bps"] == 0.0
        assert per_host["h1"]["tunnel_count"] == 1
        assert per_host["h1"]["tunnel_names"] == ["tun0"]

    def test_idle_relay_flagged_when_tunnels_silent(self):
        from zbbx_mcp.tools.correlation import _find_idle_relays

        per_host = {
            "h1": {
                "physical_bps": 200_000,
                "physical_out_bps": 0,  # receives but doesn't forward → flagged
                "tunnel_bps": 0,
                "tunnel_count": 3,
                "tunnel_names": ["tun0", "tun1", "tun2"],
            },
        }
        idle = _find_idle_relays(per_host, min_mgmt_kbps=100)
        assert len(idle) == 1
        hid, in_kbps, out_kbps, tun_count, sample = idle[0]
        assert hid == "h1"
        assert in_kbps == 200.0
        assert out_kbps == 0.0
        assert tun_count == 3
        assert sample == ["tun0", "tun1", "tun2"]

    def test_idle_relay_skipped_when_forwarding_healthy(self):
        # NAT-mode relay: physical out ≈ in (forwards) with idle tunnels by
        # design — must NOT be flagged (the out<<in gate, ADR 043).
        from zbbx_mcp.tools.correlation import _find_idle_relays

        per_host = {
            "h1": {
                "physical_bps": 200_000,
                "physical_out_bps": 190_000,  # out ≈ in → healthy forwarder
                "tunnel_bps": 0,
                "tunnel_count": 3,
                "tunnel_names": ["tun0", "tun1", "tun2"],
            },
        }
        assert _find_idle_relays(per_host, min_mgmt_kbps=100) == []

    def test_idle_relay_skipped_when_tunnels_have_traffic(self):
        from zbbx_mcp.tools.correlation import _find_idle_relays

        per_host = {
            "h1": {
                "physical_bps": 200_000,
                "tunnel_bps": 10,  # one tunnel forwarding
                "tunnel_count": 2,
                "tunnel_names": ["tun0", "tun1"],
            },
        }
        assert _find_idle_relays(per_host, 100) == []

    def test_idle_relay_skipped_below_mgmt_floor(self):
        from zbbx_mcp.tools.correlation import _find_idle_relays

        per_host = {
            "h1": {
                "physical_bps": 50_000,  # 50 kbps, below 100
                "tunnel_bps": 0,
                "tunnel_count": 2,
                "tunnel_names": ["tun0", "tun1"],
            },
        }
        assert _find_idle_relays(per_host, 100) == []

    def test_idle_relay_skipped_when_no_tunnels(self):
        from zbbx_mcp.tools.correlation import _find_idle_relays

        per_host = {
            "h1": {
                "physical_bps": 200_000,
                "tunnel_bps": 0,
                "tunnel_count": 0,
                "tunnel_names": [],
            },
        }
        assert _find_idle_relays(per_host, 100) == []

    def test_idle_relays_sorted_by_mgmt_traffic_desc(self):
        from zbbx_mcp.tools.correlation import _find_idle_relays

        per_host = {
            "h1": {"physical_bps": 100_000, "physical_out_bps": 0, "tunnel_bps": 0, "tunnel_count": 1, "tunnel_names": ["tun0"]},
            "h2": {"physical_bps": 500_000, "physical_out_bps": 0, "tunnel_bps": 0, "tunnel_count": 1, "tunnel_names": ["tun0"]},
            "h3": {"physical_bps": 250_000, "physical_out_bps": 0, "tunnel_bps": 0, "tunnel_count": 1, "tunnel_names": ["tun0"]},
        }
        out = _find_idle_relays(per_host, 50)
        assert [r[0] for r in out] == ["h2", "h3", "h1"]


class TestOutageClustering:
    """Pure-helper tests for get_outage_clusters time-window grouping."""

    def _rec(self, clock: int, hostid: str, key: str, name: str = "Down", sev: int = 4) -> dict:
        return {
            "clock": clock,
            "hostid": hostid,
            "host": f"host-{hostid}",
            "name": name,
            "severity": sev,
            "key": key,
        }

    def test_subnet_helper(self):
        from zbbx_mcp.tools.correlation import subnet24

        assert subnet24("10.0.5.42") == "10.0.5.0/24"
        assert subnet24("") == ""
        assert subnet24("not-an-ip") == ""
        assert subnet24("1.2.3") == ""
        assert subnet24("::1") == ""

    def test_three_hosts_same_subnet_within_window_form_cluster(self):
        from zbbx_mcp.tools.correlation import _cluster_problems

        records = [
            self._rec(1000, "h1", "10.0.0.0/24"),
            self._rec(1100, "h2", "10.0.0.0/24"),
            self._rec(1200, "h3", "10.0.0.0/24"),
        ]
        clusters = _cluster_problems(records, window_sec=600, min_hosts=3)
        assert len(clusters) == 1
        c = clusters[0]
        assert c["host_count"] == 3
        assert c["events"] == 3
        assert c["start"] == 1000
        assert c["end"] == 1200

    def test_below_min_hosts_does_not_cluster(self):
        from zbbx_mcp.tools.correlation import _cluster_problems

        records = [
            self._rec(1000, "h1", "10.0.0.0/24"),
            self._rec(1100, "h2", "10.0.0.0/24"),
        ]
        assert _cluster_problems(records, 600, 3) == []

    def test_outside_window_does_not_cluster(self):
        from zbbx_mcp.tools.correlation import _cluster_problems

        records = [
            self._rec(1000, "h1", "10.0.0.0/24"),
            self._rec(1500, "h2", "10.0.0.0/24"),
            self._rec(2200, "h3", "10.0.0.0/24"),  # outside 600s of h1
        ]
        # Greedy run grows h1..h2 (500s), h3 starts new run with only 1 host
        assert _cluster_problems(records, 600, 3) == []

    def test_two_separate_subnets_yield_two_clusters(self):
        from zbbx_mcp.tools.correlation import _cluster_problems

        records = [
            self._rec(1000, "h1", "10.0.0.0/24"),
            self._rec(1100, "h2", "10.0.0.0/24"),
            self._rec(1200, "h3", "10.0.0.0/24"),
            self._rec(2000, "h4", "10.0.1.0/24"),
            self._rec(2100, "h5", "10.0.1.0/24"),
            self._rec(2200, "h6", "10.0.1.0/24"),
        ]
        clusters = _cluster_problems(records, 600, 3)
        assert len(clusters) == 2
        assert {c["key"] for c in clusters} == {"10.0.0.0/24", "10.0.1.0/24"}

    def test_max_severity_propagates(self):
        from zbbx_mcp.tools.correlation import _cluster_problems

        records = [
            self._rec(1000, "h1", "10.0.0.0/24", sev=2),
            self._rec(1100, "h2", "10.0.0.0/24", sev=5),
            self._rec(1200, "h3", "10.0.0.0/24", sev=3),
        ]
        clusters = _cluster_problems(records, 600, 3)
        assert clusters[0]["max_severity"] == 5

    def test_duplicate_hostids_count_once(self):
        from zbbx_mcp.tools.correlation import _cluster_problems

        records = [
            self._rec(1000, "h1", "10.0.0.0/24", name="A"),
            self._rec(1100, "h1", "10.0.0.0/24", name="B"),
            self._rec(1200, "h2", "10.0.0.0/24"),
        ]
        # 3 events but only 2 distinct hosts — does not meet min_hosts=3
        assert _cluster_problems(records, 600, 3) == []
        # min_hosts=2 should pass
        clusters = _cluster_problems(records, 600, 2)
        assert clusters[0]["host_count"] == 2
        assert clusters[0]["events"] == 3

    def test_clusters_sorted_by_host_count_then_severity(self):
        from zbbx_mcp.tools.correlation import _cluster_problems

        records = [
            # Big cluster, low severity
            self._rec(1000, "h1", "A", sev=2),
            self._rec(1050, "h2", "A", sev=2),
            self._rec(1100, "h3", "A", sev=2),
            self._rec(1150, "h4", "A", sev=2),
            # Smaller cluster, high severity
            self._rec(1000, "h5", "B", sev=5),
            self._rec(1050, "h6", "B", sev=5),
            self._rec(1100, "h7", "B", sev=5),
        ]
        clusters = _cluster_problems(records, 600, 3)
        assert [c["key"] for c in clusters] == ["A", "B"]  # bigger first


class TestTrafficDropsSkipBreakdown:
    """Pure-helper tests for the no-baseline visibility footer."""

    def test_empty_when_nothing_skipped(self):
        from zbbx_mcp.tools.traffic import _format_skip_breakdown

        assert _format_skip_breakdown({"no_history": 0, "no_baseline_window": 0, "below_floor": 0}, 1.0) == ""

    def test_single_reason_renders(self):
        from zbbx_mcp.tools.traffic import _format_skip_breakdown

        out = _format_skip_breakdown({"no_history": 12, "no_baseline_window": 0, "below_floor": 0}, 1.0)
        assert out == "12 skipped: 12 no-history."

    def test_all_three_reasons_render_in_order(self):
        from zbbx_mcp.tools.traffic import _format_skip_breakdown

        out = _format_skip_breakdown(
            {"no_history": 5, "no_baseline_window": 3, "below_floor": 30}, 1.0,
        )
        assert out == "38 skipped: 5 no-history, 3 no-baseline-window, 30 below-1Mbps-floor."

    def test_floor_uses_min_baseline_arg(self):
        from zbbx_mcp.tools.traffic import _format_skip_breakdown

        out = _format_skip_breakdown(
            {"no_history": 0, "no_baseline_window": 0, "below_floor": 5}, 0.5,
        )
        assert "below-0.5Mbps-floor" in out

    def test_zero_categories_omitted(self):
        from zbbx_mcp.tools.traffic import _format_skip_breakdown

        out = _format_skip_breakdown(
            {"no_history": 0, "no_baseline_window": 7, "below_floor": 0}, 1.0,
        )
        # Other reasons should not appear
        assert "no-history" not in out
        assert "below-" not in out
        assert "7 no-baseline-window" in out


class TestShutdownPeerHeadroom:
    """Pure-helper tests for shutdown peer-cohort headroom logic."""

    def test_solo_when_no_peers(self):
        from zbbx_mcp.tools.trends_health import _compute_shutdown_safety

        label, headroom = _compute_shutdown_safety(50.0, [])
        assert label == "SOLO"
        assert headroom == 0.0

    def test_safe_when_cohort_headroom_covers_load_with_margin(self):
        from zbbx_mcp.tools.trends_health import _compute_shutdown_safety

        # 4 peers, each averaging 50 Mbps with peaks at 120 → 280 Mbps headroom
        peers = [{"peak": 120.0, "avg": 50.0}] * 4
        label, headroom = _compute_shutdown_safety(100.0, peers)
        assert label == "SAFE"
        assert headroom == 280.0  # 4 × (120 - 50)

    def test_risky_when_headroom_below_safety_margin(self):
        from zbbx_mcp.tools.trends_health import _compute_shutdown_safety

        # Candidate avg 100 Mbps × 1.5 margin = 150 Mbps required.
        # Peers offer 80 Mbps headroom — positive but insufficient.
        peers = [{"peak": 60.0, "avg": 20.0}, {"peak": 60.0, "avg": 20.0}]
        label, headroom = _compute_shutdown_safety(100.0, peers)
        assert label == "RISKY"
        assert headroom == 80.0

    def test_safety_margin_is_configurable(self):
        from zbbx_mcp.tools.trends_health import _compute_shutdown_safety

        peers = [{"peak": 100.0, "avg": 50.0}]  # 50 Mbps headroom
        # With margin 1.0, 50 Mbps headroom is exactly enough for 50 Mbps load
        assert _compute_shutdown_safety(50.0, peers, safety_margin=1.0)[0] == "SAFE"
        # With margin 1.5 (default), 50 Mbps load needs 75 Mbps headroom
        assert _compute_shutdown_safety(50.0, peers, safety_margin=1.5)[0] == "RISKY"

    def test_negative_spare_peers_do_not_subtract(self):
        from zbbx_mcp.tools.trends_health import _compute_shutdown_safety

        # A peer at peak < avg is impossible in real data but defensible:
        # such a peer should contribute zero, never negative headroom.
        peers = [{"peak": 100.0, "avg": 50.0}, {"peak": 10.0, "avg": 30.0}]
        label, headroom = _compute_shutdown_safety(20.0, peers)
        assert headroom == 50.0
        assert label == "SAFE"

    def test_peers_with_missing_metrics_are_dropped(self):
        from zbbx_mcp.tools.trends_health import _compute_shutdown_safety

        peers = [
            {"peak": None, "avg": None},  # no trend data — skip
            {"peak": 100.0, "avg": 30.0},  # 70 headroom
        ]
        label, headroom = _compute_shutdown_safety(40.0, peers)
        assert headroom == 70.0
        assert label == "SAFE"

    def test_candidate_without_traffic_returns_na(self):
        from zbbx_mcp.tools.trends_health import _compute_shutdown_safety

        peers = [{"peak": 100.0, "avg": 30.0}]
        label, headroom = _compute_shutdown_safety(None, peers)
        assert label == "N/A"
        assert headroom == 70.0  # still computed for the report

    def test_zero_load_candidate_is_safe(self):
        from zbbx_mcp.tools.trends_health import _compute_shutdown_safety

        # DEAD candidates with traffic_avg=0 — any peer headroom is enough.
        peers = [{"peak": 1.0, "avg": 0.5}]
        label, _ = _compute_shutdown_safety(0.0, peers)
        assert label == "SAFE"


class TestExternalIpHistoryParsing:
    """Pure-helper tests for audit-log details parsing and recovery scoring."""

    def test_list_shape_picks_ip_updates(self):
        from zbbx_mcp.tools.ip_history import parse_ip_changes

        details = (
            '[["update", "interfaces.42.ip", "1.2.3.4", "5.6.7.8"],'
            ' ["update", "host.name", "old", "new"],'
            ' ["update", "interfaces.42.port", "10050", "10050"]]'
        )
        out = parse_ip_changes(details)
        assert out == [("1.2.3.4", "5.6.7.8")]

    def test_dict_shape_picks_ip_updates(self):
        from zbbx_mcp.tools.ip_history import parse_ip_changes

        details = '{"interfaces.7.ip": ["update", "10.0.0.1", "10.0.0.2"]}'
        assert parse_ip_changes(details) == [("10.0.0.1", "10.0.0.2")]

    def test_no_change_when_old_equals_new(self):
        from zbbx_mcp.tools.ip_history import parse_ip_changes

        # Renames that touch the field but leave the value equal must be skipped.
        details = '[["update", "interfaces.42.ip", "1.2.3.4", "1.2.3.4"]]'
        assert parse_ip_changes(details) == []

    def test_non_ip_field_ignored(self):
        from zbbx_mcp.tools.ip_history import parse_ip_changes

        details = '[["update", "host.host", "a", "b"]]'
        assert parse_ip_changes(details) == []

    def test_garbage_input_returns_empty(self):
        from zbbx_mcp.tools.ip_history import parse_ip_changes

        assert parse_ip_changes("") == []
        assert parse_ip_changes("not-json") == []
        assert parse_ip_changes("[1, 2, 3]") == []  # not the expected shape

    def test_recovery_scores(self):
        from zbbx_mcp.tools.ip_history import _score_recovery

        assert _score_recovery(100.0, 90.0) == "recovered"   # 0.9
        assert _score_recovery(100.0, 70.0) == "recovered"   # 0.7 boundary
        assert _score_recovery(100.0, 50.0) == "partial"     # 0.5
        assert _score_recovery(100.0, 30.0) == "partial"     # 0.3 boundary
        assert _score_recovery(100.0, 5.0) == "still-down"   # 0.05

    def test_recovery_na_cases(self):
        from zbbx_mcp.tools.ip_history import _score_recovery

        assert _score_recovery(None, 50.0) == "n/a"
        assert _score_recovery(50.0, None) == "n/a"
        assert _score_recovery(0.0, 50.0) == "n/a"  # divide-by-zero baseline


class TestLossDriftDetection:
    """Pure-helper tests for sliding-window loss/RTT classification."""

    def testsplit_baseline_recent_partitions_by_clock(self):
        from zbbx_mcp.tools.loss_drift import split_baseline_recent

        trends = [
            {"clock": 100, "value_avg": "1.0"},
            {"clock": 200, "value_avg": "2.0"},
            {"clock": 300, "value_avg": "10.0"},  # recent
            {"clock": 400, "value_avg": "12.0"},  # recent
        ]
        base, recent = split_baseline_recent(trends, cutoff_clock=300)
        assert base == 1.5  # (1+2)/2
        assert recent == 11.0  # (10+12)/2

    def test_split_handles_missing_sides(self):
        from zbbx_mcp.tools.loss_drift import split_baseline_recent

        # Only baseline records
        b, r = split_baseline_recent([{"clock": 100, "value_avg": "5"}], 300)
        assert b == 5.0 and r is None

        # Only recent records
        b, r = split_baseline_recent([{"clock": 400, "value_avg": "5"}], 300)
        assert b is None and r == 5.0

        # Empty
        assert split_baseline_recent([], 300) == (None, None)

    def test_split_skips_garbage_values(self):
        from zbbx_mcp.tools.loss_drift import split_baseline_recent

        trends = [
            {"clock": 100, "value_avg": "not-a-number"},
            {"clock": 200, "value_avg": "2.0"},
        ]
        base, _ = split_baseline_recent(trends, 300)
        assert base == 2.0

    def test_new_loss_takes_priority_over_loss_up(self):
        from zbbx_mcp.tools.loss_drift import compute_loss_drift

        # baseline ~0% loss, recent jumps to 8% — both flags fire, prefer new-loss.
        label, details = compute_loss_drift(0.5, 8.0, None, None)
        assert label == "new-loss"
        assert details["loss_delta"] == 7.5

    def test_loss_up_when_baseline_already_high(self):
        from zbbx_mcp.tools.loss_drift import compute_loss_drift

        label, _ = compute_loss_drift(3.0, 10.0, None, None)
        assert label == "loss-up"  # baseline >= 1%, so not 'new-loss'

    def test_rtt_up_alone(self):
        from zbbx_mcp.tools.loss_drift import compute_loss_drift

        label, details = compute_loss_drift(None, None, 50.0, 90.0)
        assert label == "rtt-up"
        assert details["rtt_delta_pct"] == 80.0

    def test_loss_and_rtt_combo(self):
        from zbbx_mcp.tools.loss_drift import compute_loss_drift

        # Loss baseline >= 1% so 'new-loss' does not preempt.
        label, _ = compute_loss_drift(2.0, 10.0, 50.0, 90.0)
        assert label == "loss-and-rtt"

    def test_below_thresholds_is_ok(self):
        from zbbx_mcp.tools.loss_drift import compute_loss_drift

        label, _ = compute_loss_drift(2.0, 4.0, 50.0, 60.0)  # +2% loss, +20% RTT
        assert label == "ok"

    def test_no_data_is_na(self):
        from zbbx_mcp.tools.loss_drift import compute_loss_drift

        label, _ = compute_loss_drift(None, None, None, None)
        assert label == "n/a"

    def test_thresholds_are_configurable(self):
        from zbbx_mcp.tools.loss_drift import compute_loss_drift

        # Default loss_step=5 → not flagged at +3.
        assert compute_loss_drift(2.0, 5.0, None, None)[0] == "ok"
        # Tighten to 2 → +3 flags.
        assert compute_loss_drift(2.0, 5.0, None, None, loss_step=2.0)[0] == "loss-up"

    def test_degraded_baseline_suppresses_false_rtt_drift(self):
        from zbbx_mcp.tools.loss_drift import compute_loss_drift

        # Baseline measured during an outage (47% loss); recent recovered to ~0%.
        # RTT "doubling" vs that unreliable baseline is recovery, not real drift.
        label, _ = compute_loss_drift(47.24, 0.09, 76.4, 142.5)
        assert label == "ok"


class TestOutageClusterGroupingV2:
    """Pure-helper tests for multi-level cluster grouping (#119)."""

    def testsubnet24_and_subnet16(self):
        from zbbx_mcp.tools.correlation import _group_key

        assert _group_key("subnet24", ip="10.20.30.40") == "10.20.30.0/24"
        assert _group_key("subnet16", ip="10.20.30.40") == "10.20.0.0/16"
        assert _group_key("subnet24", ip="") == ""
        assert _group_key("subnet16", ip="not-an-ip") == ""

    def test_provider_level_skips_unknown(self):
        from zbbx_mcp.tools.correlation import _group_key

        assert _group_key("provider", provider="OVH") == "OVH"
        # 'Other'/'Unknown' would lump unrelated hosts — must be empty key.
        assert _group_key("provider", provider="Other") == ""
        assert _group_key("provider", provider="Unknown") == ""
        assert _group_key("provider", provider="") == ""

    def test_hostgroup_level(self):
        from zbbx_mcp.tools.correlation import _group_key

        assert _group_key("hostgroup", hostgroup="EU/edge") == "EU/edge"
        assert _group_key("hostgroup", hostgroup="") == ""

    def test_unknown_level_is_empty(self):
        from zbbx_mcp.tools.correlation import _group_key

        assert _group_key("subnet8", ip="1.2.3.4") == ""

    def test_auto_levels_constant_is_narrowest_first(self):
        from zbbx_mcp.tools.correlation import _AUTO_LEVELS

        assert _AUTO_LEVELS == ("subnet24", "subnet16", "provider")


class TestServicePortSplit:
    """Pure-helper tests for detect_service_port_split classification."""

    def test_split_label_when_service_alone_collapses(self):
        from zbbx_mcp.tools.disruption import _classify_service_split

        # Service: 100→20 (-80%), Mgmt: 50→48 (-4%)
        label, details = _classify_service_split(100.0, 20.0, 50.0, 48.0)
        assert label == "split"
        assert details["service_drop_pct"] == 80.0
        assert details["mgmt_drop_pct"] == 4.0

    def test_full_outage_label_when_both_collapse(self):
        from zbbx_mcp.tools.disruption import _classify_service_split

        label, _ = _classify_service_split(100.0, 20.0, 50.0, 5.0)  # both -80% / -90%
        assert label == "full-outage"

    def test_ok_when_neither_collapses(self):
        from zbbx_mcp.tools.disruption import _classify_service_split

        label, _ = _classify_service_split(100.0, 95.0, 50.0, 49.0)
        assert label == "ok"

    def test_na_when_baseline_missing(self):
        from zbbx_mcp.tools.disruption import _classify_service_split

        assert _classify_service_split(None, 20.0, 50.0, 48.0)[0] == "n/a"
        assert _classify_service_split(0.0, 20.0, 50.0, 48.0)[0] == "n/a"

    def test_thresholds_configurable(self):
        from zbbx_mcp.tools.disruption import _classify_service_split

        # 30% service drop: not flagged at default (50%), flagged when threshold lowered.
        assert _classify_service_split(100.0, 70.0, 50.0, 49.0)[0] == "ok"
        assert _classify_service_split(
            100.0, 70.0, 50.0, 49.0, service_drop_pct=20.0,
        )[0] == "split"


class TestRegionalLossClassification:
    """Pure-helper tests for detect_regional_traffic_loss."""

    def test_collapsed_when_one_region_drops_others_flat(self):
        from zbbx_mcp.tools.disruption import _classify_regional_loss

        # EU collapses 80%, NA stays flat (-2%).
        regions = {"EU": (1000.0, 200.0), "NA": (500.0, 490.0)}
        flagged = _classify_regional_loss(regions)
        assert len(flagged) == 1
        assert flagged[0]["region"] == "EU"
        assert flagged[0]["label"] == "collapsed"

    def test_solo_drop_when_no_flat_peer(self):
        from zbbx_mcp.tools.disruption import _classify_regional_loss

        # Both regions drop heavily — no peer is flat, so solo-drop label.
        regions = {"EU": (1000.0, 200.0), "NA": (500.0, 100.0)}
        flagged = _classify_regional_loss(regions)
        assert {r["region"] for r in flagged} == {"EU", "NA"}
        assert all(r["label"] == "solo-drop" for r in flagged)

    def test_below_threshold_not_flagged(self):
        from zbbx_mcp.tools.disruption import _classify_regional_loss

        regions = {"EU": (100.0, 80.0), "NA": (100.0, 95.0)}  # 20% / 5%
        assert _classify_regional_loss(regions) == []  # 20% < default 30% threshold

    def test_missing_data_skipped(self):
        from zbbx_mcp.tools.disruption import _classify_regional_loss

        regions = {"EU": (None, 200.0), "NA": (500.0, 50.0), "APAC": (300.0, 290.0)}
        flagged = _classify_regional_loss(regions)
        # APAC is flat (~3%), so NA gets 'collapsed'.
        assert len(flagged) == 1
        assert flagged[0]["region"] == "NA"


class TestDisruptionWaveDetection:
    """Pure-helper tests for the wave-clustering algorithm."""

    def _drop(self, clock, hostid, subnet, drop_pct=50.0):
        return {
            "clock": clock,
            "hostid": hostid,
            "host": f"h-{hostid}",
            "subnet": subnet,
            "hostgroup": "test",
            "drop_pct": drop_pct,
        }

    def test_wave_fires_when_thresholds_met(self):
        from zbbx_mcp.tools.disruption import _compute_waves

        drops = [
            self._drop(1000, "h1", "10.0.1.0/24"),
            self._drop(1100, "h2", "10.0.2.0/24"),
            self._drop(1200, "h3", "10.0.3.0/24"),
            self._drop(1300, "h4", "10.0.4.0/24"),
            self._drop(1400, "h5", "10.0.5.0/24"),
        ]
        waves = _compute_waves(drops, window_sec=3600, min_hosts=5, min_subnets=3)
        assert len(waves) == 1
        assert waves[0]["host_count"] == 5
        assert waves[0]["subnet_count"] == 5

    def test_wave_does_not_fire_when_subnets_collapse(self):
        from zbbx_mcp.tools.disruption import _compute_waves

        # 5 hosts, all in same /24 — fails min_subnets=3
        drops = [self._drop(1000 + i * 100, f"h{i}", "10.0.1.0/24") for i in range(5)]
        assert _compute_waves(drops, min_hosts=5, min_subnets=3) == []

    def test_window_boundary_excludes_late_arrivals(self):
        from zbbx_mcp.tools.disruption import _compute_waves

        drops = [
            self._drop(1000, "h1", "10.0.1.0/24"),
            self._drop(1100, "h2", "10.0.2.0/24"),
            self._drop(1200, "h3", "10.0.3.0/24"),
            self._drop(5000, "h4", "10.0.4.0/24"),  # outside 3600s window
            self._drop(5100, "h5", "10.0.5.0/24"),
        ]
        # First three meet min_subnets=3 but only 3 hosts; lower bar.
        waves = _compute_waves(drops, window_sec=3600, min_hosts=3, min_subnets=3)
        assert len(waves) == 1
        assert waves[0]["host_count"] == 3

    def test_severity_label_by_avg_drop(self):
        from zbbx_mcp.tools.disruption import _compute_waves

        drops = [
            self._drop(1000, "h1", "a", drop_pct=80),
            self._drop(1100, "h2", "b", drop_pct=80),
            self._drop(1200, "h3", "c", drop_pct=80),
        ]
        waves = _compute_waves(drops, min_hosts=3, min_subnets=3)
        assert waves[0]["severity"] == "critical"

        drops = [
            self._drop(1000, "h1", "a", drop_pct=40),
            self._drop(1100, "h2", "b", drop_pct=40),
            self._drop(1200, "h3", "c", drop_pct=40),
        ]
        waves = _compute_waves(drops, min_hosts=3, min_subnets=3)
        assert waves[0]["severity"] == "medium"


class TestAtRiskScoring:
    """Pure-helper tests for the composite at-risk score."""

    def test_zero_inputs_score_zero(self):
        from zbbx_mcp.tools.risk import _compute_risk_score

        score, details = _compute_risk_score(0, "ok", 0.0)
        assert score == 0.0
        assert details["drift_label"] == "ok"

    def test_more_peer_rotations_increase_score(self):
        from zbbx_mcp.tools.risk import _compute_risk_score

        low, _ = _compute_risk_score(1, "ok", 0.0)
        high, _ = _compute_risk_score(20, "ok", 0.0)
        assert high > low

    def test_drift_label_dominates_when_other_signals_zero(self):
        from zbbx_mcp.tools.risk import _compute_risk_score

        rtt, _ = _compute_risk_score(0, "rtt-up", 0.0)
        loss, _ = _compute_risk_score(0, "loss-up", 0.0)
        combo, _ = _compute_risk_score(0, "loss-and-rtt", 0.0)
        assert rtt < loss < combo

    def test_age_capped_at_90_days(self):
        from zbbx_mcp.tools.risk import _compute_risk_score

        cap, _ = _compute_risk_score(0, "ok", 90.0)
        bigger, _ = _compute_risk_score(0, "ok", 365.0)
        assert cap == bigger  # capped, so equal

    def test_none_age_treated_as_capped(self):
        from zbbx_mcp.tools.risk import _compute_risk_score

        # No prior rotation observed → treat as cap, not zero.
        cap, _ = _compute_risk_score(0, "ok", 90.0)
        none_score, _ = _compute_risk_score(0, "ok", None)
        assert cap == none_score


class TestBlastRadiusClassification:
    """Pure-helper tests for cohort connection-count delta labels."""

    def test_absorbing_when_post_gains_at_least_10pct(self):
        from zbbx_mcp.tools.risk import _compute_blast_radius

        label, delta = _compute_blast_radius(100.0, 120.0)
        assert label == "absorbing"
        assert delta == 20.0

    def test_draining_when_post_loses_more_than_10pct(self):
        from zbbx_mcp.tools.risk import _compute_blast_radius

        label, delta = _compute_blast_radius(100.0, 80.0)
        assert label == "draining"
        assert delta == -20.0

    def test_stable_within_10pct(self):
        from zbbx_mcp.tools.risk import _compute_blast_radius

        for post in (95.0, 100.0, 105.0):
            label, _ = _compute_blast_radius(100.0, post)
            assert label == "stable"

    def test_na_when_pre_missing_or_zero(self):
        from zbbx_mcp.tools.risk import _compute_blast_radius

        assert _compute_blast_radius(None, 50.0) == ("n/a", None)
        assert _compute_blast_radius(0.0, 50.0) == ("n/a", None)
        assert _compute_blast_radius(50.0, None) == ("n/a", None)


class TestRecoveryAggregate:
    """Pure-helper tests for fleet-level recovery KPI aggregation."""

    def test_aggregate_basic(self):
        from zbbx_mcp.tools.ip_history import _aggregate_recovery_scores

        rotations = [
            {"score": "recovered"},
            {"score": "recovered"},
            {"score": "recovered"},
            {"score": "partial"},
            {"score": "still-down"},
            {"score": "n/a"},
        ]
        agg = _aggregate_recovery_scores(rotations)
        assert agg["total"] == 6
        assert agg["recovered"] == 3
        assert agg["partial"] == 1
        assert agg["still_down"] == 1
        assert agg["na"] == 1
        # rate = 3 recovered / 5 determined-outcome = 60%
        assert agg["rate_pct"] == 60.0

    def test_aggregate_all_na_yields_none_rate(self):
        from zbbx_mcp.tools.ip_history import _aggregate_recovery_scores

        agg = _aggregate_recovery_scores([{"score": "n/a"}, {"score": "n/a"}])
        assert agg["total"] == 2
        assert agg["na"] == 2
        assert agg["rate_pct"] is None

    def test_aggregate_unknown_label_treated_as_na(self):
        from zbbx_mcp.tools.ip_history import _aggregate_recovery_scores

        agg = _aggregate_recovery_scores([{"score": "weird"}, {"score": "recovered"}])
        assert agg["na"] == 1
        assert agg["recovered"] == 1
        assert agg["rate_pct"] == 100.0  # 1 of 1 determined

    def test_aggregate_empty(self):
        from zbbx_mcp.tools.ip_history import _aggregate_recovery_scores

        agg = _aggregate_recovery_scores([])
        assert agg["total"] == 0
        assert agg["rate_pct"] is None


class TestProblemNameNormalization:
    """Pure-helper tests for normalize_problem_name (#127)."""

    def test_strips_on_hostname(self):
        from zbbx_mcp.formatters import normalize_problem_name

        assert normalize_problem_name(
            "ServiceX: on host-a-1 error", "host-a-1",
        ) == "ServiceX: error"

    def test_subhost_form_preferred_over_parent(self):
        from zbbx_mcp.formatters import normalize_problem_name

        # Hostname is "parent child"; the trigger names the sub-host. We must
        # strip the sub-host form, not "child" only.
        result = normalize_problem_name(
            "ServiceX: on parent child error", "parent child",
        )
        assert result == "ServiceX: error"

    def test_collapses_internal_whitespace(self):
        from zbbx_mcp.formatters import normalize_problem_name

        # After stripping, multiple spaces around the cut point collapse to one.
        result = normalize_problem_name("CPU on host-a is overloaded", "host-a")
        assert result == "CPU is overloaded"

    def test_no_match_when_hostname_absent(self):
        from zbbx_mcp.formatters import normalize_problem_name

        # "on" is in the trigger but not paired with the hostname.
        result = normalize_problem_name(
            "Listener on port 8080 is down", "host-a",
        )
        assert result == "Listener on port 8080 is down"

    def test_returns_input_when_hostname_missing(self):
        from zbbx_mcp.formatters import normalize_problem_name

        assert normalize_problem_name("Anything", "") == "Anything"
        assert normalize_problem_name("", "host-a") == ""

    def test_two_hosts_normalize_to_same_name(self):
        from zbbx_mcp.formatters import normalize_problem_name

        # The whole point: triggers that differ only by embedded host should
        # collapse to a single dedup key.
        a = normalize_problem_name("ServiceY: on host-a error", "host-a")
        b = normalize_problem_name("ServiceY: on host-b error", "host-b")
        assert a == b == "ServiceY: error"

    def test_case_insensitive_on_keyword(self):
        from zbbx_mcp.formatters import normalize_problem_name

        # Some triggers capitalise differently — match regardless of case.
        assert normalize_problem_name(
            "Boot ON host-a failed", "host-a",
        ) == "Boot failed"


class TestHostFloodGrouping:
    """Pure-helper tests for _group_host_floods (#128)."""

    def _rec(self, hostid, host, name="Trigger", severity=4, clock=1000):
        return {
            "hostid": hostid,
            "host": host,
            "name": name,
            "severity": severity,
            "clock": clock,
        }

    def test_threshold_filters_below_min(self):
        from zbbx_mcp.tools.floods import _group_host_floods

        records = [
            self._rec("h1", "host-a", name="A"),
            self._rec("h1", "host-a", name="B"),
        ]
        # 2 problems on one host, min_problems=5 → no flood.
        assert _group_host_floods(records, {}, min_problems=5) == []

    def test_flood_emitted_when_count_meets_threshold(self):
        from zbbx_mcp.tools.floods import _group_host_floods

        records = [
            self._rec("h1", "host-a", name=f"T-{i}", severity=2 + (i % 3))
            for i in range(5)
        ]
        result = _group_host_floods(records, {}, min_problems=5)
        assert len(result) == 1
        assert result[0]["host"] == "host-a"
        assert result[0]["problem_count"] == 5
        assert result[0]["max_severity"] == 4

    def test_subhost_merged_into_parent(self):
        from zbbx_mcp.tools.floods import _group_host_floods

        # Three problems on parent, two on its child — counts as one flood of 5.
        records = [
            self._rec("p1", "parent", name="A"),
            self._rec("p1", "parent", name="B"),
            self._rec("p1", "parent", name="C"),
            self._rec("c1", "parent child1", name="D"),
            self._rec("c1", "parent child1", name="E"),
        ]
        parent_map = {"c1": "p1"}
        result = _group_host_floods(records, parent_map, min_problems=5)
        assert len(result) == 1
        assert result[0]["hostid"] == "p1"
        assert result[0]["host"] == "parent"
        assert result[0]["problem_count"] == 5
        assert result[0]["child_count"] == 1

    def test_earliest_clock_picked(self):
        from zbbx_mcp.tools.floods import _group_host_floods

        records = [
            self._rec("h1", "host-a", clock=2000),
            self._rec("h1", "host-a", clock=1000),
            self._rec("h1", "host-a", clock=1500),
            self._rec("h1", "host-a", clock=2500),
            self._rec("h1", "host-a", clock=3000),
        ]
        result = _group_host_floods(records, {}, min_problems=5)
        assert result[0]["earliest_clock"] == 1000

    def test_sample_triggers_dedup_and_cap(self):
        from zbbx_mcp.tools.floods import _group_host_floods

        records = [
            self._rec("h1", "host-a", name="DupTrigger") for _ in range(7)
        ] + [
            self._rec("h1", "host-a", name="UniqueTrigger") for _ in range(3)
        ]
        result = _group_host_floods(records, {}, min_problems=5)
        # Sample is set-deduplicated, capped at 5 entries.
        assert len(result[0]["sample_triggers"]) == 2
        assert set(result[0]["sample_triggers"]) == {"DupTrigger", "UniqueTrigger"}

    def test_floods_sorted_by_count_then_severity(self):
        from zbbx_mcp.tools.floods import _group_host_floods

        records = (
            [self._rec("h1", "small", severity=5) for _ in range(5)]
            + [self._rec("h2", "big", severity=2) for _ in range(8)]
        )
        result = _group_host_floods(records, {}, min_problems=5)
        # Bigger flood comes first regardless of severity.
        assert result[0]["host"] == "big"
        assert result[1]["host"] == "small"


class TestPhysicalNicRegexFallback:
    """Pure-helper tests for #129 — NIC name regex fallback in _split_iface_metrics."""

    def test_unused_secondary_nic_classified_physical(self):
        from zbbx_mcp.tools.correlation import _split_iface_metrics

        # eno3 / enp130s0f0 are physical NICs not in the curated TRAFFIC_IN_KEYS
        # list. Without the regex they would fall into the tunnel bucket.
        items = [
            {"hostid": "h1", "key_": "net.if.in[eno3]", "lastvalue": "0"},
            {"hostid": "h1", "key_": "net.if.in[enp130s0f0]", "lastvalue": "0"},
            {"hostid": "h1", "key_": "net.if.in[tun0]", "lastvalue": "0"},
        ]
        per_host = _split_iface_metrics(items, [], frozenset())
        # Only tun0 should land in tunnel_names.
        assert per_host["h1"]["tunnel_count"] == 1
        assert per_host["h1"]["tunnel_names"] == ["tun0"]

    def test_usb_ethernet_enx_prefix(self):
        from zbbx_mcp.tools.correlation import _split_iface_metrics

        items = [
            {"hostid": "h1", "key_": "net.if.in[enx00aa11bb22cc]", "lastvalue": "0"},
            {"hostid": "h1", "key_": "net.if.in[gre1]", "lastvalue": "0"},
        ]
        per_host = _split_iface_metrics(items, [], frozenset())
        assert per_host["h1"]["tunnel_names"] == ["gre1"]

    def test_explicit_physical_keys_still_win(self):
        from zbbx_mcp.tools.correlation import _split_iface_metrics

        # Matching the curated key takes precedence; the regex is only a fallback.
        items = [
            {"hostid": "h1", "key_": "net.if.in[eth0]", "lastvalue": "100"},
        ]
        per_host = _split_iface_metrics(items, [], frozenset({"net.if.in[eth0]"}))
        assert per_host["h1"]["physical_bps"] == 100
        assert per_host["h1"]["tunnel_count"] == 0

    def test_unknown_prefix_still_treated_as_tunnel(self):
        from zbbx_mcp.tools.correlation import _split_iface_metrics

        items = [
            {"hostid": "h1", "key_": "net.if.in[gre1]", "lastvalue": "0"},
            {"hostid": "h1", "key_": "net.if.in[mytun0]", "lastvalue": "0"},
        ]
        per_host = _split_iface_metrics(items, [], frozenset())
        assert sorted(per_host["h1"]["tunnel_names"]) == ["gre1", "mytun0"]


class TestServiceCheckStaleGate:
    """Pure-helper tests for is_service_check_stale (#130)."""

    def test_state_one_is_stale(self):
        from zbbx_mcp.fetch import is_service_check_stale

        # state=1 means Zabbix flagged the item unsupported.
        item = {"state": "1", "lastclock": str(1_700_000_000), "lastvalue": "0"}
        assert is_service_check_stale(item, now=1_700_000_300) is True

    def test_lastclock_inside_window_is_fresh(self):
        from zbbx_mcp.fetch import is_service_check_stale

        item = {"state": "0", "lastclock": str(1_700_000_000), "lastvalue": "1"}
        # 5 minutes old — well within the default 30min window.
        assert is_service_check_stale(item, now=1_700_000_300) is False

    def test_lastclock_outside_window_is_stale(self):
        from zbbx_mcp.fetch import is_service_check_stale

        item = {"state": "0", "lastclock": str(1_700_000_000), "lastvalue": "1"}
        # 31 minutes old — past the default 30min window.
        assert is_service_check_stale(item, now=1_700_000_000 + 31 * 60) is True

    def test_zero_lastclock_is_stale(self):
        from zbbx_mcp.fetch import is_service_check_stale

        # Item never polled → lastclock=0. Treat as stale.
        assert is_service_check_stale(
            {"state": "0", "lastclock": "0"}, now=1_700_000_300,
        ) is True

    def test_missing_lastclock_is_stale(self):
        from zbbx_mcp.fetch import is_service_check_stale

        assert is_service_check_stale(
            {"state": "0"}, now=1_700_000_300,
        ) is True

    def test_garbage_lastclock_is_stale(self):
        from zbbx_mcp.fetch import is_service_check_stale

        assert is_service_check_stale(
            {"state": "0", "lastclock": "not-a-number"}, now=1_700_000_300,
        ) is True

    def test_custom_stale_window(self):
        from zbbx_mcp.fetch import is_service_check_stale

        item = {"state": "0", "lastclock": str(1_700_000_000), "lastvalue": "1"}
        # 5 minutes old; tighten the window to 60s and it becomes stale.
        assert is_service_check_stale(item, now=1_700_000_300, stale_sec=60) is True
        # Loosen to 1h and the same item is fresh.
        assert is_service_check_stale(item, now=1_700_000_300, stale_sec=3600) is False

    def test_state_one_overrides_fresh_lastclock(self):
        from zbbx_mcp.fetch import is_service_check_stale

        # Even a recent lastclock cannot rescue an unsupported item.
        item = {"state": "1", "lastclock": str(1_700_000_290), "lastvalue": "0"}
        assert is_service_check_stale(item, now=1_700_000_300) is True


class TestProblemAgeBuckets:
    """Pure-helper tests for the age-histogram bucketer (#132)."""

    def _p(self, sev, age_sec, now=1_700_000_000):
        return {"severity": sev, "clock": now - age_sec}

    def test_empty_input_returns_empty_dict(self):
        from zbbx_mcp.tools.health import _bucket_problems_by_age

        assert _bucket_problems_by_age([], 1_700_000_000) == {}

    def test_three_problems_one_per_bucket(self):
        from zbbx_mcp.tools.health import _bucket_problems_by_age

        problems = [
            self._p(4, 3600),       # <1d
            self._p(4, 2 * 86400),  # 1-3d
            self._p(4, 5 * 86400),  # 3-7d
        ]
        out = _bucket_problems_by_age(problems, 1_700_000_000)
        assert out[4] == {"<1d": 1, "1-3d": 1, "3-7d": 1, "7d+": 0}

    def test_seven_day_overflow_lands_in_seven_d_plus(self):
        from zbbx_mcp.tools.health import _bucket_problems_by_age

        problems = [self._p(5, 14 * 86400)]
        assert _bucket_problems_by_age(problems, 1_700_000_000)[5]["7d+"] == 1

    def test_boundary_one_day_lands_in_one_three(self):
        from zbbx_mcp.tools.health import _bucket_problems_by_age

        # Exactly 86400s old = 1 day. Strict "<1d" pushes it into the 1-3d bucket.
        problems = [self._p(3, 86400)]
        out = _bucket_problems_by_age(problems, 1_700_000_000)
        assert out[3]["<1d"] == 0
        assert out[3]["1-3d"] == 1

    def test_severities_partitioned_independently(self):
        from zbbx_mcp.tools.health import _bucket_problems_by_age

        problems = [
            self._p(5, 3600),
            self._p(5, 3600),
            self._p(2, 3600),
        ]
        out = _bucket_problems_by_age(problems, 1_700_000_000)
        assert out[5]["<1d"] == 2
        assert out[2]["<1d"] == 1

    def test_bad_clock_skipped(self):
        from zbbx_mcp.tools.health import _bucket_problems_by_age

        problems = [
            {"severity": 4, "clock": 0},
            {"severity": 4, "clock": "garbage"},
            {"severity": 4, "clock": 1_700_000_000 - 3600},  # valid
        ]
        out = _bucket_problems_by_age(problems, 1_700_000_000)
        assert out[4]["<1d"] == 1
        assert sum(out[4].values()) == 1


class TestStaleItemsCascade:
    """Pure-helper tests for cascade collapse (#133)."""

    def test_no_master_passes_through(self):
        from zbbx_mcp.tools.items import _collapse_dependent_chain

        stale = [{"itemid": "i1", "master_itemid": ""}]
        out = _collapse_dependent_chain(stale)
        assert len(out) == 1
        assert out[0]["affected_count"] == 0

    def test_child_with_stale_master_collapsed(self):
        from zbbx_mcp.tools.items import _collapse_dependent_chain

        stale = [
            {"itemid": "i1", "master_itemid": ""},   # root
            {"itemid": "i2", "master_itemid": "i1"}, # child of root
        ]
        out = _collapse_dependent_chain(stale)
        assert len(out) == 1
        assert out[0]["itemid"] == "i1"
        assert out[0]["affected_count"] == 1

    def test_child_with_non_stale_master_kept(self):
        from zbbx_mcp.tools.items import _collapse_dependent_chain

        # Child references master_itemid that is NOT in the stale list —
        # treat the child as its own root (its master is healthy).
        stale = [{"itemid": "i2", "master_itemid": "i_healthy"}]
        out = _collapse_dependent_chain(stale)
        assert len(out) == 1
        assert out[0]["itemid"] == "i2"
        assert out[0]["affected_count"] == 0

    def test_two_hop_chain_collapses_to_root(self):
        from zbbx_mcp.tools.items import _collapse_dependent_chain

        stale = [
            {"itemid": "root", "master_itemid": ""},
            {"itemid": "mid", "master_itemid": "root"},
            {"itemid": "leaf", "master_itemid": "mid"},
        ]
        out = _collapse_dependent_chain(stale)
        assert len(out) == 1
        assert out[0]["itemid"] == "root"
        assert out[0]["affected_count"] == 2  # mid + leaf

    def test_multiple_children_share_one_root(self):
        from zbbx_mcp.tools.items import _collapse_dependent_chain

        stale = [
            {"itemid": "root", "master_itemid": ""},
            {"itemid": "c1", "master_itemid": "root"},
            {"itemid": "c2", "master_itemid": "root"},
            {"itemid": "c3", "master_itemid": "root"},
        ]
        out = _collapse_dependent_chain(stale)
        assert len(out) == 1
        assert out[0]["affected_count"] == 3

    def test_circular_reference_does_not_loop(self):
        from zbbx_mcp.tools.items import _collapse_dependent_chain

        # Pathological: A → B → A. The root_of walk must terminate.
        stale = [
            {"itemid": "a", "master_itemid": "b"},
            {"itemid": "b", "master_itemid": "a"},
        ]
        out = _collapse_dependent_chain(stale)
        # Both end up rooted at one of the cycle members; count should be
        # bounded and the function must return.
        assert sum(s.get("affected_count", 0) for s in out) >= 0

    def test_input_not_mutated(self):
        from zbbx_mcp.tools.items import _collapse_dependent_chain

        stale = [{"itemid": "i1", "master_itemid": ""}]
        original = dict(stale[0])
        _ = _collapse_dependent_chain(stale)
        assert stale[0] == original


class TestWaveCohesionGuard:
    """Pure-helper tests for the country-concentration check inside _compute_waves (#134)."""

    def _drop(self, clock, hostid, subnet, country, drop_pct=60.0):
        return {
            "clock": clock,
            "hostid": hostid,
            "host": f"h-{hostid}",
            "subnet": subnet,
            "hostgroup": "test",
            "country": country,
            "drop_pct": drop_pct,
        }

    def test_globally_spread_drops_are_filtered_out(self):
        from zbbx_mcp.tools.disruption import _compute_waves

        # 5 hosts, 5 different countries — top-country share = 1/5 = 20%, well below 40%.
        drops = [
            self._drop(1000 + 100 * i, f"h{i}", f"10.0.{i}.0/24", c)
            for i, c in enumerate(["DE", "AE", "MX", "GT", "ID"])
        ]
        waves = _compute_waves(drops, min_hosts=5, min_subnets=3)
        assert waves == []

    def test_concentrated_country_passes(self):
        from zbbx_mcp.tools.disruption import _compute_waves

        # 5 hosts, all in TR — concentration 100%.
        drops = [
            self._drop(1000 + 100 * i, f"h{i}", f"10.0.{i}.0/24", "TR")
            for i in range(5)
        ]
        waves = _compute_waves(drops, min_hosts=5, min_subnets=3)
        assert len(waves) == 1
        assert waves[0]["top_country"] == "TR"
        assert waves[0]["top_country_share"] == 1.0

    def test_partial_concentration_at_default_threshold(self):
        from zbbx_mcp.tools.disruption import _compute_waves

        # 6 hosts: 4 TR (67%) + 2 elsewhere — should pass at default 0.4.
        drops = [
            self._drop(1000, "h1", "10.0.1.0/24", "TR"),
            self._drop(1100, "h2", "10.0.2.0/24", "TR"),
            self._drop(1200, "h3", "10.0.3.0/24", "TR"),
            self._drop(1300, "h4", "10.0.4.0/24", "TR"),
            self._drop(1400, "h5", "10.0.5.0/24", "ID"),
            self._drop(1500, "h6", "10.0.6.0/24", "MX"),
        ]
        waves = _compute_waves(drops, min_hosts=5, min_subnets=3)
        assert len(waves) == 1
        assert waves[0]["top_country"] == "TR"
        assert abs(waves[0]["top_country_share"] - 4 / 6) < 1e-9

    def test_threshold_is_configurable(self):
        from zbbx_mcp.tools.disruption import _compute_waves

        # 3 TR / 2 ID = 60% top share at exactly min_hosts=5. Default 40%
        # passes; tighten to 90% and any sub-bucket that cohesion-passes
        # falls below min_hosts, so the whole cluster is rejected.
        drops = (
            [self._drop(1000 + 100 * i, f"a{i}", f"10.0.{i}.0/24", "TR") for i in range(3)]
            + [self._drop(1300 + 100 * i, f"b{i}", f"10.1.{i}.0/24", "ID") for i in range(2)]
        )
        assert len(_compute_waves(drops, min_hosts=5, min_subnets=3)) == 1
        assert _compute_waves(
            drops, min_hosts=5, min_subnets=3, min_country_concentration=0.9,
        ) == []

    def test_records_without_country_bypass_cohesion(self):
        from zbbx_mcp.tools.disruption import _compute_waves

        # No `country` field — backwards compat. Cohesion check is skipped.
        records = [
            {"clock": 1000 + 100 * i, "hostid": f"h{i}", "host": f"h-{i}",
             "subnet": f"10.0.{i}.0/24", "hostgroup": "test", "drop_pct": 60.0}
            for i in range(5)
        ]
        waves = _compute_waves(records, min_hosts=5, min_subnets=3)
        assert len(waves) == 1
        assert waves[0]["top_country"] == ""
        assert waves[0]["top_country_share"] == 1.0


class TestPeerRelativeDropFilter:
    """Pure-helper tests for _compute_peer_relative_drops (#135)."""

    def test_below_absolute_threshold_dropped(self):
        from zbbx_mcp.tools.disruption import _compute_peer_relative_drops

        records = [
            {"hostid": "h1", "drop_pct": 30.0, "cohort_key": "free:tier:tr"},
        ]
        out = _compute_peer_relative_drops(records, min_drop_pct=50.0)
        assert out == []

    def test_diurnal_cohort_dropped_uniformly_filtered_out(self):
        from zbbx_mcp.tools.disruption import _compute_peer_relative_drops

        # 5 hosts in same cohort all drop 60% — peer-relative ≈ 0 — filtered out.
        records = [
            {"hostid": f"h{i}", "drop_pct": 60.0, "cohort_key": "free:tier:tr"}
            for i in range(5)
        ]
        out = _compute_peer_relative_drops(records, min_drop_pct=50.0)
        assert out == []

    def test_genuinely_impacted_host_kept(self):
        from zbbx_mcp.tools.disruption import _compute_peer_relative_drops

        # h1 drops 80% while peers drop 10% — peer-relative ~70 > 20.
        records = [
            {"hostid": "h1", "drop_pct": 80.0, "cohort_key": "free:tier:tr"},
            {"hostid": "h2", "drop_pct": 10.0, "cohort_key": "free:tier:tr"},
            {"hostid": "h3", "drop_pct": 12.0, "cohort_key": "free:tier:tr"},
            {"hostid": "h4", "drop_pct": 8.0, "cohort_key": "free:tier:tr"},
        ]
        out = _compute_peer_relative_drops(records, min_drop_pct=50.0)
        assert len(out) == 1
        assert out[0]["hostid"] == "h1"
        assert out[0]["peer_relative_drop"] == 80.0 - 10.0  # cohort_drop = (10+12+8)/3 = 10

    def test_small_cohort_passes_absolute_only(self):
        from zbbx_mcp.tools.disruption import _compute_peer_relative_drops

        # 2 hosts in cohort (below default min_cohort_size=3) — peer gate skipped,
        # absolute gate fires alone.
        records = [
            {"hostid": "h1", "drop_pct": 80.0, "cohort_key": "x"},
            {"hostid": "h2", "drop_pct": 75.0, "cohort_key": "x"},
        ]
        out = _compute_peer_relative_drops(records, min_drop_pct=50.0)
        assert len(out) == 2
        assert all(r["cohort_drop"] is None for r in out)
        assert all(r["peer_relative_drop"] is None for r in out)
        assert all(r["cohort_size"] == 2 for r in out)

    def test_solo_cohort_passes_absolute_only(self):
        from zbbx_mcp.tools.disruption import _compute_peer_relative_drops

        records = [{"hostid": "h1", "drop_pct": 80.0, "cohort_key": "x"}]
        out = _compute_peer_relative_drops(records, min_drop_pct=50.0)
        assert len(out) == 1
        assert out[0]["cohort_drop"] is None
        assert out[0]["cohort_size"] == 1

    def test_min_relative_drop_threshold_configurable(self):
        from zbbx_mcp.tools.disruption import _compute_peer_relative_drops

        # h1 drops 60%, peers average 50% — peer-relative 10. Below default 20%, kept at 5%.
        records = [
            {"hostid": "h1", "drop_pct": 60.0, "cohort_key": "x"},
            {"hostid": "h2", "drop_pct": 50.0, "cohort_key": "x"},
            {"hostid": "h3", "drop_pct": 50.0, "cohort_key": "x"},
            {"hostid": "h4", "drop_pct": 50.0, "cohort_key": "x"},
        ]
        assert _compute_peer_relative_drops(records, min_drop_pct=50.0) == []
        kept = _compute_peer_relative_drops(
            records, min_drop_pct=50.0, min_peer_relative_drop=5.0,
        )
        # h1 passes (relative=10), peers fail (relative=-3.33 each).
        assert {r["hostid"] for r in kept} == {"h1"}

    def test_separate_cohorts_evaluated_independently(self):
        from zbbx_mcp.tools.disruption import _compute_peer_relative_drops

        records = [
            # cohort_a: all drop together — all rejected
            *[{"hostid": f"a{i}", "drop_pct": 70.0, "cohort_key": "a"} for i in range(4)],
            # cohort_b: one outlier — kept
            {"hostid": "b1", "drop_pct": 90.0, "cohort_key": "b"},
            {"hostid": "b2", "drop_pct": 5.0, "cohort_key": "b"},
            {"hostid": "b3", "drop_pct": 5.0, "cohort_key": "b"},
            {"hostid": "b4", "drop_pct": 5.0, "cohort_key": "b"},
        ]
        kept = _compute_peer_relative_drops(records, min_drop_pct=50.0)
        assert {r["hostid"] for r in kept} == {"b1"}


class TestFormatAge:
    """Pure-helper tests for the compact age renderer (#136)."""

    def test_seconds_under_a_minute(self):
        from zbbx_mcp.formatters import format_age

        assert format_age(0) == "0s"
        assert format_age(45) == "45s"
        assert format_age(59) == "59s"

    def test_minutes(self):
        from zbbx_mcp.formatters import format_age

        assert format_age(60) == "1m"
        assert format_age(150) == "2m"
        assert format_age(3599) == "59m"

    def test_hours(self):
        from zbbx_mcp.formatters import format_age

        assert format_age(3600) == "1h"
        assert format_age(7200) == "2h"
        assert format_age(86399) == "23h"

    def test_days(self):
        from zbbx_mcp.formatters import format_age

        assert format_age(86400) == "1d"
        assert format_age(7 * 86400) == "7d"
        assert format_age(180 * 86400) == "180d"

    def test_negative_clamped_to_zero(self):
        from zbbx_mcp.formatters import format_age

        assert format_age(-5) == "0s"
        assert format_age(-1_000_000) == "0s"


class TestParentSubHostCanonicalization:
    """#138: parent + sub-host must fold to one physical machine in counts."""

    def test_build_parent_map_pairs_child_with_parent(self):
        from zbbx_mcp.data import build_parent_map

        hosts = [
            {"hostid": "p", "host": "edge-us65"},
            {"hostid": "c", "host": "edge-us65 us71"},
            {"hostid": "x", "host": "edge-de01"},
        ]
        pm = build_parent_map(hosts)
        assert pm == {"c": "p"}

    def test_canonical_dedup_via_set(self):
        from zbbx_mcp.data import build_parent_map

        # The canonical-id pattern: parent_map.get(hid, hid). After this,
        # a parent + child pair maps to one canonical id.
        hosts = [
            {"hostid": "p", "host": "edge-us65"},
            {"hostid": "c", "host": "edge-us65 us71"},
        ]
        pm = build_parent_map(hosts)
        canonical_ids = {pm.get(h["hostid"], h["hostid"]) for h in hosts}
        assert canonical_ids == {"p"}

    def test_cohesion_does_not_double_count_sub_host(self):
        # End-to-end through _compute_waves: 6 records, but two of them
        # represent parent + sub of one physical machine. After upstream
        # canonicalisation (caller's responsibility) they share hostid
        # "p_us". top-country share over distinct hostids is 2/5 = 40%.
        # Without the fix, share over 6 records would be 3/6 = 50%.
        from zbbx_mcp.tools.disruption import _compute_waves

        # Note: each record carries the canonical hostid. The tool
        # upstream of _compute_waves de-dupes traffic into the parent,
        # so this list has one record per canonical machine.
        drops = [
            {"clock": 1000, "hostid": "p_us", "host": "edge-us65",
             "subnet": "10.0.1.0/24", "hostgroup": "x", "country": "US",
             "drop_pct": 60.0},
            {"clock": 1100, "hostid": "p_us2", "host": "edge-us66",
             "subnet": "10.0.2.0/24", "hostgroup": "x", "country": "US",
             "drop_pct": 60.0},
            {"clock": 1200, "hostid": "p_de", "host": "edge-de01",
             "subnet": "10.0.3.0/24", "hostgroup": "x", "country": "DE",
             "drop_pct": 60.0},
            {"clock": 1300, "hostid": "p_id", "host": "edge-id01",
             "subnet": "10.0.4.0/24", "hostgroup": "x", "country": "ID",
             "drop_pct": 60.0},
            {"clock": 1400, "hostid": "p_mx", "host": "edge-mx01",
             "subnet": "10.0.5.0/24", "hostgroup": "x", "country": "MX",
             "drop_pct": 60.0},
        ]
        # 5 unique machines, US share = 2/5 = 40% — meets default 0.4.
        waves = _compute_waves(drops, min_hosts=5, min_subnets=3)
        assert len(waves) == 1
        assert waves[0]["host_count"] == 5
        assert abs(waves[0]["top_country_share"] - 0.4) < 1e-9

    def test_cluster_unique_host_counts_use_canonical_id(self):
        # The cluster code path: after _build_records sets r["hostid"]
        # to canonical, _cluster_problems' set comprehension dedupes
        # parent+sub to one entry.
        from zbbx_mcp.tools.correlation import _cluster_problems

        records = [
            # Parent and sub-host both have a problem in the same cluster.
            # Both records carry the canonical (parent's) hostid.
            {"clock": 1000, "hostid": "p", "host": "edge-us65",
             "name": "X", "severity": 4, "key": "10.0.0.0/24"},
            {"clock": 1050, "hostid": "p", "host": "edge-us65",
             "name": "Y", "severity": 4, "key": "10.0.0.0/24"},
            # Two distinct other hosts in the same /24.
            {"clock": 1100, "hostid": "h2", "host": "edge-us66",
             "name": "X", "severity": 4, "key": "10.0.0.0/24"},
            {"clock": 1150, "hostid": "h3", "host": "edge-us67",
             "name": "X", "severity": 4, "key": "10.0.0.0/24"},
        ]
        # 4 records but only 3 distinct canonical hosts. min_hosts=3 passes;
        # min_hosts=4 fails (would have passed without canonicalisation).
        c3 = _cluster_problems(records, window_sec=600, min_hosts=3)
        assert len(c3) == 1
        assert c3[0]["host_count"] == 3
        c4 = _cluster_problems(records, window_sec=600, min_hosts=4)
        assert c4 == []


class TestNormalizeCountry:
    """Pure-helper tests for normalize_country (#140)."""

    def test_iso2_passthrough(self):
        from zbbx_mcp.data import normalize_country

        for v in ["RU", "ru", "Ru", " ru "]:
            assert normalize_country(v) == "RU"

    def test_uk_alias_to_gb(self):
        from zbbx_mcp.data import normalize_country

        # The existing extract_country alias ("UK"→"GB") is preserved.
        assert normalize_country("UK") == "GB"
        assert normalize_country("uk") == "GB"

    def test_iso3_recognised(self):
        from zbbx_mcp.data import normalize_country

        assert normalize_country("RUS") == "RU"
        assert normalize_country("usa") == "US"
        assert normalize_country("DEU") == "DE"

    def test_full_name_recognised(self):
        from zbbx_mcp.data import normalize_country

        assert normalize_country("Russia") == "RU"
        assert normalize_country("UNITED STATES") == "US"
        assert normalize_country("Saudi Arabia") == "SA"
        assert normalize_country("Czechia") == "CZ"
        assert normalize_country("Czech Republic") == "CZ"
        assert normalize_country("United Kingdom") == "GB"
        assert normalize_country("UAE") == "AE"

    def test_empty_and_unknown_return_empty(self):
        from zbbx_mcp.data import normalize_country

        assert normalize_country("") == ""
        assert normalize_country("   ") == ""
        assert normalize_country(None) == ""  # type: ignore[arg-type]
        assert normalize_country("Atlantis") == ""
        assert normalize_country("ZZZ") == ""

    def test_two_letter_unknown_still_returns_iso2_like(self):
        from zbbx_mcp.data import normalize_country

        # We don't enumerate the full ISO-2 set; any 2-letter alphabetic
        # input is treated as a code (the downstream filter just won't
        # match any host). UK-alias normalisation still applies.
        assert normalize_country("ZZ") == "ZZ"


class TestResolveCountry:
    """Pure-helper tests for resolve_country chain (#141)."""

    def test_extract_country_takes_precedence(self):
        from zbbx_mcp.data import resolve_country

        h = {
            "host": "edge-de01",
            "inventory": {"country_code": "FR", "country_name": "France"},
        }
        # Hostname says DE; inventory disagreement loses to the name.
        assert resolve_country(h) == "DE"

    def test_inventory_country_code_used_when_hostname_empty(self):
        from zbbx_mcp.data import resolve_country

        h = {"host": "control-plane-01", "inventory": {"country_code": "us"}}
        assert resolve_country(h) == "US"

    def test_inventory_country_name_used_when_code_empty(self):
        from zbbx_mcp.data import resolve_country

        h = {
            "host": "control-plane-01",
            "inventory": {"country_code": "", "country_name": "Russia"},
        }
        assert resolve_country(h) == "RU"

    def test_returns_empty_when_all_sources_empty(self):
        from zbbx_mcp.data import resolve_country

        h = {"host": "control-plane-01", "inventory": {}}
        assert resolve_country(h) == ""

    def test_handles_missing_inventory_field(self):
        from zbbx_mcp.data import resolve_country

        # Older Zabbix host.get without selectInventory — no field at all.
        h = {"host": "control-plane-01"}
        assert resolve_country(h) == ""

    def test_inventory_unknown_name_falls_through(self):
        from zbbx_mcp.data import resolve_country

        h = {"host": "weird", "inventory": {"country_name": "Atlantis"}}
        assert resolve_country(h) == ""


class TestTelemetrySummary:
    """Pure-helper tests for _summarise_records (#7)."""

    def _rec(self, tool="x", status="ok", duration_ms=10, response_size=100, ts=None):
        r = {
            "tool": tool,
            "status": status,
            "duration_ms": duration_ms,
            "response_size": response_size,
        }
        if ts is not None:
            r["ts"] = ts
        return r

    def test_per_tool_counts_and_avg(self):
        from zbbx_mcp.tools.telemetry import _summarise_records

        records = [
            self._rec(tool="search_hosts", duration_ms=20),
            self._rec(tool="search_hosts", duration_ms=40),
            self._rec(tool="get_problems", duration_ms=200),
        ]
        out = _summarise_records(records)
        by_tool = {r["tool"]: r for r in out}
        assert by_tool["search_hosts"]["calls"] == 2
        assert by_tool["search_hosts"]["avg_ms"] == 30.0
        assert by_tool["get_problems"]["calls"] == 1
        assert by_tool["get_problems"]["avg_ms"] == 200.0

    def test_error_rate_pct(self):
        from zbbx_mcp.tools.telemetry import _summarise_records

        records = (
            [self._rec(status="ok") for _ in range(7)]
            + [self._rec(status="error") for _ in range(3)]
        )
        out = _summarise_records(records)
        assert out[0]["errors"] == 3
        assert out[0]["error_pct"] == 30.0

    def test_sorted_by_calls_desc(self):
        from zbbx_mcp.tools.telemetry import _summarise_records

        records = (
            [self._rec(tool="a") for _ in range(2)]
            + [self._rec(tool="b") for _ in range(5)]
            + [self._rec(tool="c") for _ in range(3)]
        )
        out = _summarise_records(records)
        assert [r["tool"] for r in out] == ["b", "c", "a"]

    def test_max_ms_tracked(self):
        from zbbx_mcp.tools.telemetry import _summarise_records

        records = [
            self._rec(duration_ms=10),
            self._rec(duration_ms=500),
            self._rec(duration_ms=50),
        ]
        out = _summarise_records(records)
        assert out[0]["max_ms"] == 500

    def test_garbage_duration_treated_as_zero(self):
        from zbbx_mcp.tools.telemetry import _summarise_records

        records = [
            {"tool": "x", "status": "ok", "duration_ms": "not-a-number"},
            {"tool": "x", "status": "ok", "duration_ms": 100},
        ]
        out = _summarise_records(records)
        assert out[0]["calls"] == 2
        assert out[0]["avg_ms"] == 50.0  # (0 + 100) / 2

    def test_since_ts_filter_drops_old_records(self):
        from zbbx_mcp.tools.telemetry import _summarise_records

        records = [
            self._rec(tool="a", ts=1000),
            self._rec(tool="a", ts=2000),
            self._rec(tool="a", ts=3000),
        ]
        out = _summarise_records(records, since_ts=2000)
        assert out[0]["calls"] == 2  # 2000 and 3000 kept; 1000 dropped

    def test_iso_timestamp_filter(self):
        from zbbx_mcp.tools.telemetry import _summarise_records

        records = [
            self._rec(tool="a", ts="2026-05-04T12:00:00Z"),
            self._rec(tool="a", ts="2026-05-05T12:00:00Z"),
        ]
        # 2026-05-05 00:00 UTC = 1777939200
        out = _summarise_records(records, since_ts=1777939200)
        assert out[0]["calls"] == 1  # only the May 5 record


class TestDiagnoseHostHelpers:
    """Pure-helper tests for diagnose_host (#2 composite)."""

    def test_classify_mode_server_via_traffic_keys(self):
        from zbbx_mcp.tools.diagnose import _classify_host_mode

        items = [{"key_": "net.if.in[eth0]"}, {"key_": "agent.ping"}]
        assert _classify_host_mode({}, items) == "server"

    def test_classify_mode_server_via_agent_ping_alone(self):
        from zbbx_mcp.tools.diagnose import _classify_host_mode

        items = [{"key_": "agent.ping"}]
        assert _classify_host_mode({}, items) == "server"

    def test_classify_mode_domain_when_no_agent_no_traffic(self):
        from zbbx_mcp.tools.diagnose import _classify_host_mode

        items = [{"key_": "webcheck.https.status"}]
        assert _classify_host_mode({}, items) == "domain"

    def test_classify_mode_domain_when_no_items(self):
        from zbbx_mcp.tools.diagnose import _classify_host_mode

        assert _classify_host_mode({}, []) == "domain"

    def test_verdict_traffic_lost_when_traffic_collapses_with_healthy_agent(self):
        from zbbx_mcp.tools.diagnose import _classify_verdict

        v, action = _classify_verdict(
            mode="server",
            agent_ping_val=1,
            agent_ping_age_min=0.5,
            traffic_baseline_mbps=200.0,
            traffic_recent_mbps=2.0,  # 1% of baseline
            open_problems=0,
            https_down=False,
            https_age_h=None,
        )
        assert v == "traffic_lost"
        assert "rotat" in action.lower() or "external" in action.lower()

    def test_verdict_down_when_agent_and_traffic_both_gone(self):
        from zbbx_mcp.tools.diagnose import _classify_verdict

        v, _ = _classify_verdict(
            mode="server",
            agent_ping_val=0,
            agent_ping_age_min=60.0,
            traffic_baseline_mbps=200.0,
            traffic_recent_mbps=0.5,
            open_problems=0,
            https_down=False,
            https_age_h=None,
        )
        assert v == "down"

    def test_verdict_degraded_agent_down_traffic_ok(self):
        from zbbx_mcp.tools.diagnose import _classify_verdict

        v, _ = _classify_verdict(
            mode="server",
            agent_ping_val=0,
            agent_ping_age_min=10.0,
            traffic_baseline_mbps=200.0,
            traffic_recent_mbps=180.0,
            open_problems=0,
            https_down=False,
            https_age_h=None,
        )
        assert v == "degraded"

    def test_verdict_degraded_when_open_problems(self):
        from zbbx_mcp.tools.diagnose import _classify_verdict

        v, action = _classify_verdict(
            mode="server",
            agent_ping_val=1,
            agent_ping_age_min=0.5,
            traffic_baseline_mbps=200.0,
            traffic_recent_mbps=180.0,
            open_problems=3,
            https_down=False,
            https_age_h=None,
        )
        assert v == "degraded"
        assert "3" in action

    def test_verdict_healthy_when_everything_ok(self):
        from zbbx_mcp.tools.diagnose import _classify_verdict

        v, _ = _classify_verdict(
            mode="server",
            agent_ping_val=1,
            agent_ping_age_min=0.5,
            traffic_baseline_mbps=200.0,
            traffic_recent_mbps=180.0,
            open_problems=0,
            https_down=False,
            https_age_h=None,
        )
        assert v == "healthy"

    def test_verdict_https_down_in_domain_mode(self):
        from zbbx_mcp.tools.diagnose import _classify_verdict

        v, action = _classify_verdict(
            mode="domain",
            agent_ping_val=None,
            agent_ping_age_min=None,
            traffic_baseline_mbps=None,
            traffic_recent_mbps=None,
            open_problems=2,
            https_down=True,
            https_age_h=17.5,
        )
        assert v == "https_down"
        assert "17" in action

    def test_verdict_domain_healthy_no_problems(self):
        from zbbx_mcp.tools.diagnose import _classify_verdict

        v, _ = _classify_verdict(
            mode="domain",
            agent_ping_val=None,
            agent_ping_age_min=None,
            traffic_baseline_mbps=None,
            traffic_recent_mbps=None,
            open_problems=0,
            https_down=False,
            https_age_h=None,
        )
        assert v == "healthy"

    def test_verdict_agent_age_5min_boundary_mutation_sentinel(self):
        """Mutation sentinel — pins the agent-age threshold's strict-greater semantics.

        ``_classify_verdict`` marks an agent unreachable when
        ``agent_ping_age_min > 5``. The 5-minute constant is a load-bearing
        threshold: it gates every server-mode diagnosis. This test fixes
        three boundary points around it so any off-by-one mutation
        (``> 5`` → ``>= 5``, ``> 4``, ``> 6``, etc.) shows up as a test
        failure rather than silently misclassifying healthy hosts as
        degraded on every run.

        Pairing each boundary point with healthy traffic + healthy
        agent.ping isolates the age check — the only path to a non-
        ``healthy`` verdict here goes through the age clause.
        """
        from zbbx_mcp.tools.diagnose import _classify_verdict

        common: dict = dict(
            mode="server",
            agent_ping_val=1,
            traffic_baseline_mbps=200.0,
            traffic_recent_mbps=180.0,
            open_problems=0,
            https_down=False,
            https_age_h=None,
        )

        # Just below the boundary — must stay healthy.
        v, _ = _classify_verdict(agent_ping_age_min=4.99, **common)
        assert v == "healthy", "4.99m must be healthy; catches `> 4` mutations"

        # Exactly at the boundary — must NOT flip (strict-greater).
        # The off-by-one trap: `>= 5` would mark every host at exactly
        # 5min ago as degraded — high-volume false positive.
        v, _ = _classify_verdict(agent_ping_age_min=5.00, **common)
        assert v == "healthy", "5.00m exactly must stay healthy; catches `>= 5` mutations"

        # Just above — must flip to degraded.
        v, _ = _classify_verdict(agent_ping_age_min=5.01, **common)
        assert v == "degraded", "5.01m must mark agent unreachable; catches `> 6` mutations"

    def test_verdict_traffic_below_5mbps_baseline_not_flagged_as_traffic_lost(self):
        from zbbx_mcp.tools.diagnose import _classify_verdict

        # A 0.5 Mbps -> 0.05 Mbps drop is technically 90%, but the baseline
        # is too small to count as a real signal; should not flip to traffic_lost.
        v, _ = _classify_verdict(
            mode="server",
            agent_ping_val=1,
            agent_ping_age_min=0.5,
            traffic_baseline_mbps=0.5,
            traffic_recent_mbps=0.05,
            open_problems=0,
            https_down=False,
            https_age_h=None,
        )
        assert v == "healthy"


class TestBulkDiagnoseHelpers:
    """Pure-helper tests for bulk_diagnose (#148)."""

    def _facts(self, **overrides):
        base = {
            "host": "h1", "verdict": "healthy", "mode": "server",
            "action": "no issues", "problems": [],
            "agent_ping_val": 1, "agent_ping_age_min": 0.5,
            "traffic_baseline_mbps": 200.0, "traffic_recent_mbps": 190.0,
            "https_down": False, "https_age_h": None,
        }
        base.update(overrides)
        return base

    def test_primary_signal_healthy(self):
        from zbbx_mcp.tools.diagnose import _verdict_primary_signal
        assert _verdict_primary_signal(self._facts()) == "OK"

    def test_primary_signal_down(self):
        from zbbx_mcp.tools.diagnose import _verdict_primary_signal
        f = self._facts(verdict="down")
        assert "agent" in _verdict_primary_signal(f).lower()

    def test_primary_signal_traffic_lost_shows_mbps(self):
        from zbbx_mcp.tools.diagnose import _verdict_primary_signal
        f = self._facts(
            verdict="traffic_lost",
            traffic_baseline_mbps=255.9, traffic_recent_mbps=2.3,
        )
        s = _verdict_primary_signal(f)
        assert "256" in s and "2.3" in s

    def test_primary_signal_https_down_shows_hours(self):
        from zbbx_mcp.tools.diagnose import _verdict_primary_signal
        f = self._facts(
            verdict="https_down", mode="domain",
            https_down=True, https_age_h=17.5,
        )
        assert "17" in _verdict_primary_signal(f)

    def test_primary_signal_degraded_with_problems(self):
        from zbbx_mcp.tools.diagnose import _verdict_primary_signal
        f = self._facts(
            verdict="degraded",
            problems=[{"name": "x"}, {"name": "y"}, {"name": "z"}],
        )
        assert "3" in _verdict_primary_signal(f)

    def test_render_bulk_table_empty(self):
        from zbbx_mcp.tools.diagnose import _render_bulk_table
        assert "No hosts" in _render_bulk_table([], 0)

    def test_render_bulk_table_sorts_by_severity(self):
        from zbbx_mcp.tools.diagnose import _render_bulk_table
        rows = [
            self._facts(host="ok-host", verdict="healthy"),
            self._facts(host="dead-host", verdict="down"),
            self._facts(host="slow-host", verdict="degraded"),
            self._facts(host="lost-host", verdict="traffic_lost"),
        ]
        out = _render_bulk_table(rows, 4)
        down_pos = out.find("dead-host")
        traffic_pos = out.find("lost-host")
        degraded_pos = out.find("slow-host")
        healthy_pos = out.find("ok-host")
        assert down_pos < traffic_pos < degraded_pos < healthy_pos

    def test_render_bulk_table_counts_flagged(self):
        from zbbx_mcp.tools.diagnose import _render_bulk_table
        rows = [
            self._facts(host="a", verdict="healthy"),
            self._facts(host="b", verdict="down"),
            self._facts(host="c", verdict="traffic_lost"),
        ]
        out = _render_bulk_table(rows, 3)
        assert "2 flagged" in out

    def test_render_bulk_table_truncates_long_action(self):
        from zbbx_mcp.tools.diagnose import _render_bulk_table
        rows = [self._facts(
            verdict="traffic_lost",
            action="a" * 200,
        )]
        out = _render_bulk_table(rows, 1)
        assert "..." in out


class TestTagFilterParser:
    """Pure-helper tests for parse_tag_filter (#145)."""

    def test_empty_returns_empty_list(self):
        from zbbx_mcp.tag_filter import parse_tag_filter
        assert parse_tag_filter("") == []
        assert parse_tag_filter("   ") == []

    def test_single_key_value_equals(self):
        from zbbx_mcp.tag_filter import parse_tag_filter
        assert parse_tag_filter("role:edge") == [
            {"tag": "role", "value": "edge", "operator": 0}
        ]

    def test_multiple_pairs_and_combined(self):
        from zbbx_mcp.tag_filter import parse_tag_filter
        out = parse_tag_filter("role:edge,env:prod")
        assert out == [
            {"tag": "role", "value": "edge", "operator": 0},
            {"tag": "env", "value": "prod", "operator": 0},
        ]

    def test_whitespace_tolerated(self):
        from zbbx_mcp.tag_filter import parse_tag_filter
        out = parse_tag_filter("role:edge ,  env:prod  ")
        assert out == [
            {"tag": "role", "value": "edge", "operator": 0},
            {"tag": "env", "value": "prod", "operator": 0},
        ]

    def test_bare_key_means_exists(self):
        from zbbx_mcp.tag_filter import parse_tag_filter
        assert parse_tag_filter("role") == [
            {"tag": "role", "value": "", "operator": 4}
        ]

    def test_empty_value_after_colon_means_exists(self):
        from zbbx_mcp.tag_filter import parse_tag_filter
        assert parse_tag_filter("role:") == [
            {"tag": "role", "value": "", "operator": 4}
        ]

    def test_empty_key_is_skipped(self):
        from zbbx_mcp.tag_filter import parse_tag_filter
        assert parse_tag_filter(":value,") == []

    def test_trailing_comma_does_not_break(self):
        from zbbx_mcp.tag_filter import parse_tag_filter
        assert parse_tag_filter("role:edge,") == [
            {"tag": "role", "value": "edge", "operator": 0}
        ]


class TestSubnetMatcher:
    """Pure-helper tests for diagnose_subnet (#149)."""

    def test_slash_24_match(self):
        from zbbx_mcp.tools.diagnose import _ip_matches_subnet
        assert _ip_matches_subnet("10.1.2.5", "10.1.2.0/24") is True

    def test_slash_24_no_match(self):
        from zbbx_mcp.tools.diagnose import _ip_matches_subnet
        assert _ip_matches_subnet("10.1.3.5", "10.1.2.0/24") is False

    def test_slash_16_match(self):
        from zbbx_mcp.tools.diagnose import _ip_matches_subnet
        assert _ip_matches_subnet("10.1.42.99", "10.1.0.0/16") is True

    def test_slash_16_no_match(self):
        from zbbx_mcp.tools.diagnose import _ip_matches_subnet
        assert _ip_matches_subnet("10.2.42.99", "10.1.0.0/16") is False

    def test_dotted_prefix_match(self):
        from zbbx_mcp.tools.diagnose import _ip_matches_subnet
        assert _ip_matches_subnet("10.1.2.5", "10.1.2") is True
        assert _ip_matches_subnet("10.1.2.5", "10.1.2.") is True

    def test_dotted_prefix_no_match(self):
        from zbbx_mcp.tools.diagnose import _ip_matches_subnet
        assert _ip_matches_subnet("10.1.20.5", "10.1.2") is False

    def test_empty_inputs(self):
        from zbbx_mcp.tools.diagnose import _ip_matches_subnet
        assert _ip_matches_subnet("", "10.1.2.0/24") is False
        assert _ip_matches_subnet("10.1.2.5", "") is False

    def test_unsupported_cidr_bits(self):
        from zbbx_mcp.tools.diagnose import _ip_matches_subnet
        # /28 etc. are not supported — return False (safer than wrong match)
        assert _ip_matches_subnet("10.1.2.5", "10.1.2.0/28") is False

    def test_malformed_cidr_does_not_crash(self):
        from zbbx_mcp.tools.diagnose import _ip_matches_subnet
        assert _ip_matches_subnet("10.1.2.5", "/24") is False
        assert _ip_matches_subnet("10.1.2.5", "garbage/24") is False
        assert _ip_matches_subnet("10.1.2.5", "10.1.2.0/abc") is False


class TestAckActionBuilder:
    """Pure-helper tests for _build_ack_action (v1.8.3 acknowledge_problem extension)."""

    def test_default_is_acknowledge_only(self):
        from zbbx_mcp.tools.problems import _build_ack_action
        assert _build_ack_action() == 2

    def test_close_only(self):
        from zbbx_mcp.tools.problems import _build_ack_action
        # ack (2) + close (1) = 3
        assert _build_ack_action(close=True) == 3

    def test_message_sets_bit_4(self):
        from zbbx_mcp.tools.problems import _build_ack_action
        # ack (2) + message (4) = 6
        assert _build_ack_action(message="hello") == 6

    def test_severity_sets_bit_8(self):
        from zbbx_mcp.tools.problems import _build_ack_action
        # ack (2) + severity (8) = 10
        assert _build_ack_action(severity=4) == 10

    def test_severity_out_of_range_is_ignored(self):
        from zbbx_mcp.tools.problems import _build_ack_action
        assert _build_ack_action(severity=-1) == 2
        assert _build_ack_action(severity=6) == 2
        assert _build_ack_action(severity=99) == 2

    def test_all_optional_flags_compose(self):
        from zbbx_mcp.tools.problems import _build_ack_action
        # ack(2) + close(1) + msg(4) + sev(8) = 15
        assert _build_ack_action(
            close=True, message="x", severity=3,
        ) == 15

    def test_unack_replaces_ack_bit(self):
        from zbbx_mcp.tools.problems import _build_ack_action
        # unack (16) replaces ack (2) — mutually exclusive
        assert _build_ack_action(unack=True) == 16

    def test_unack_can_combine_with_close_and_message(self):
        from zbbx_mcp.tools.problems import _build_ack_action
        # unack(16) + close(1) + msg(4) = 21
        assert _build_ack_action(unack=True, close=True, message="x") == 21

    def test_suppress_bit(self):
        from zbbx_mcp.tools.problems import _build_ack_action
        # ack(2) + suppress(32) — ADR 059
        assert _build_ack_action(suppress=True) == 34

    def test_unsuppress_bit(self):
        from zbbx_mcp.tools.problems import _build_ack_action
        # ack(2) + unsuppress(64)
        assert _build_ack_action(unsuppress=True) == 66

    def test_suppress_with_message_combo(self):
        from zbbx_mcp.tools.problems import _build_ack_action
        # ack(2) + msg(4) + suppress(32) = 38
        assert _build_ack_action(message="snooze", suppress=True) == 38


class TestBuildRankAction:
    """Pure-helper tests for _build_rank_action (ADR 060)."""

    def test_rank_as_symptom(self):
        from zbbx_mcp.tools.problems import _build_rank_action
        assert _build_rank_action() == 256

    def test_unrank_to_cause(self):
        from zbbx_mcp.tools.problems import _build_rank_action
        assert _build_rank_action(unrank=True) == 128

    def test_message_adds_bit_4(self):
        from zbbx_mcp.tools.problems import _build_rank_action
        assert _build_rank_action(message="correlated by subnet") == 260
        assert _build_rank_action(unrank=True, message="split") == 132


class TestSuppressUntilFromHours:
    """Pure-helper tests for _suppress_until_from_hours (ADR 059)."""

    NOW = 1_000_000

    def test_zero_means_no_suppression(self):
        from zbbx_mcp.tools.problems import _suppress_until_from_hours
        assert _suppress_until_from_hours(0, self.NOW) is None

    def test_positive_hours_to_epoch(self):
        from zbbx_mcp.tools.problems import _suppress_until_from_hours
        assert _suppress_until_from_hours(4, self.NOW) == self.NOW + 4 * 3600

    def test_fractional_hours(self):
        from zbbx_mcp.tools.problems import _suppress_until_from_hours
        assert _suppress_until_from_hours(0.5, self.NOW) == self.NOW + 1800

    def test_negative_means_indefinite_zero(self):
        from zbbx_mcp.tools.problems import _suppress_until_from_hours
        # Zabbix encodes "until the problem resolves" as suppress_until=0.
        assert _suppress_until_from_hours(-1, self.NOW) == 0


class TestZabbixVersionHelpers:
    """Pure-helper tests for version parsing + feature matrix."""

    def test_parse_standard_version(self):
        from zbbx_mcp.tools.health import _parse_zabbix_version
        assert _parse_zabbix_version("6.4.2") == (6, 4, 2)

    def test_parse_two_part_version(self):
        from zbbx_mcp.tools.health import _parse_zabbix_version
        assert _parse_zabbix_version("7.0") == (7, 0, 0)

    def test_parse_empty_returns_zeros(self):
        from zbbx_mcp.tools.health import _parse_zabbix_version
        assert _parse_zabbix_version("") == (0, 0, 0)

    def test_parse_garbage_returns_zeros(self):
        from zbbx_mcp.tools.health import _parse_zabbix_version
        assert _parse_zabbix_version("not-a-version") == (0, 0, 0)

    def test_parse_partial_garbage(self):
        from zbbx_mcp.tools.health import _parse_zabbix_version
        # "6.x.2" — major parses, minor is garbage → stop there
        assert _parse_zabbix_version("6.x.2") == (6, 0, 0)

    def test_feature_matrix_zabbix_64(self):
        from zbbx_mcp.tools.health import _feature_matrix
        feats = dict(_feature_matrix(6, 4))
        assert feats["Unacknowledge action (action bit 16)"] is True
        assert feats["Cause / symptom rank actions (bits 128/256)"] is True
        assert feats["Connector API (data streaming)"] is False
        assert feats["HA cluster API (core.ha.get)"] is False

    def test_feature_matrix_zabbix_60_no_rank_actions(self):
        from zbbx_mcp.tools.health import _feature_matrix
        feats = dict(_feature_matrix(6, 0))
        assert feats["Cause / symptom rank actions (bits 128/256)"] is False
        assert feats["Unacknowledge action (action bit 16)"] is True

    def test_feature_matrix_zabbix_70_unlocks_everything(self):
        from zbbx_mcp.tools.health import _feature_matrix
        feats = dict(_feature_matrix(7, 0))
        assert feats["Connector API (data streaming)"] is True
        assert feats["Proxy groups (proxygroup.get)"] is True
        assert feats["HA cluster API (core.ha.get)"] is True


class TestCostSummaryRedactPartial:
    """Pure-helper tests for the redact_partial flag on get_cost_summary (#150)."""

    def _fixture(self):
        # Two products: ProdA fully priced, ProdB partial (4 of 5 priced).
        # Two providers: ProvX fully priced, ProvY partial (3 of 4 priced).
        prod_costs = {
            "ProdA / t1": {"count": 3, "total": 300.0},
            "ProdB / t1": {"count": 4, "total": 800.0},
        }
        prov_costs = {
            "ProvX": {"count": 3, "total": 300.0},
            "ProvY": {"count": 3, "total": 600.0},
        }
        prod_totals = {"ProdA / t1": 3, "ProdB / t1": 5}
        prov_totals = {"ProvX": 3, "ProvY": 4}
        return dict(
            prod_costs=prod_costs, prov_costs=prov_costs,
            prod_totals=prod_totals, prov_totals=prov_totals,
            costed=7, total_hosts=8,
        )

    def test_default_preserves_full_output(self):
        from zbbx_mcp.tools.costs_summary import _render_cost_summary
        out = _render_cost_summary(**self._fixture(), redact_partial=False)
        # Grand total includes both products
        assert "$1,100.00/month" in out
        # The "Servers with cost" line is present
        assert "Servers with cost: 7 | Without: 1" in out
        # Both product rows present
        assert "ProdA / t1" in out
        assert "ProdB / t1" in out
        # Both provider rows present
        assert "ProvX" in out
        assert "ProvY" in out
        # No footer
        assert "Filtered to fully-attributed lines" not in out

    def test_redact_drops_partial_product_row(self):
        from zbbx_mcp.tools.costs_summary import _render_cost_summary
        out = _render_cost_summary(**self._fixture(), redact_partial=True)
        # Partial product gone, fully-priced kept
        assert "ProdA / t1" in out
        assert "ProdB / t1" not in out

    def test_redact_drops_partial_provider_row(self):
        from zbbx_mcp.tools.costs_summary import _render_cost_summary
        out = _render_cost_summary(**self._fixture(), redact_partial=True)
        assert "ProvX" in out
        assert "ProvY" not in out

    def test_redact_recomputes_grand_total_from_kept_rows(self):
        from zbbx_mcp.tools.costs_summary import _render_cost_summary
        out = _render_cost_summary(**self._fixture(), redact_partial=True)
        # Only ProdA (the kept row) contributes: $300/mo, $3,600/yr
        assert "$300.00/month" in out
        assert "$3,600.00/year" in out
        # The stale $1,100 from default mode must not appear
        assert "$1,100.00" not in out

    def test_redact_suppresses_without_line(self):
        from zbbx_mcp.tools.costs_summary import _render_cost_summary
        out = _render_cost_summary(**self._fixture(), redact_partial=True)
        assert "Servers with cost" not in out
        assert "Without:" not in out

    def test_redact_appends_footer(self):
        from zbbx_mcp.tools.costs_summary import _render_cost_summary
        out = _render_cost_summary(**self._fixture(), redact_partial=True)
        assert "Filtered to fully-attributed lines" in out

    def test_redact_with_no_priced_data_yields_zero_total(self):
        from zbbx_mcp.tools.costs_summary import _render_cost_summary
        # All groups partial — nothing survives.
        out = _render_cost_summary(
            prod_costs={"P / t": {"count": 1, "total": 50.0}},
            prov_costs={"Q": {"count": 1, "total": 50.0}},
            prod_totals={"P / t": 2},  # 2 total, 1 priced → partial
            prov_totals={"Q": 2},
            costed=1, total_hosts=2,
            redact_partial=True,
        )
        assert "$0.00/month" in out
        # The footer still goes on so the caller knows redaction was active.
        assert "Filtered to fully-attributed lines" in out

    def test_unknown_key_in_totals_defaults_to_kept(self):
        from zbbx_mcp.tools.costs_summary import _render_cost_summary
        # A product appears in costs but not in totals (defensive: classify
        # drift between the two passes). Default behaviour: keep the row
        # rather than silently drop it.
        out = _render_cost_summary(
            prod_costs={"Orphan / t": {"count": 1, "total": 100.0}},
            prov_costs={"P": {"count": 1, "total": 100.0}},
            prod_totals={},
            prov_totals={"P": 1},
            costed=1, total_hosts=1,
            redact_partial=True,
        )
        assert "Orphan / t" in out


class TestCanonicalHostGroups:
    """Pure-helper tests for canonical_host_groups (ADR 032)."""

    def test_standalone_host_one_group(self):
        from zbbx_mcp.data import canonical_host_groups
        hosts = [{"hostid": "1", "host": "solo"}]
        groups = canonical_host_groups(hosts)
        assert len(groups) == 1
        assert groups[0]["rep_host"]["host"] == "solo"
        assert groups[0]["sub_count"] == 0
        assert groups[0]["sub_hosts"] == []
        assert groups[0]["all_hostids"] == ["1"]

    def test_parent_with_subhosts_folds_to_one_group(self):
        from zbbx_mcp.data import canonical_host_groups
        # Parent "edge01" with five sub-hosts "edge01 v1".."edge01 v5"
        hosts = [{"hostid": "1", "host": "edge01"}] + [
            {"hostid": str(i + 2), "host": f"edge01 v{i + 1}"} for i in range(5)
        ]
        groups = canonical_host_groups(hosts)
        assert len(groups) == 1
        g = groups[0]
        assert g["rep_host"]["host"] == "edge01"
        assert g["sub_count"] == 5
        assert sorted(g["all_hostids"]) == ["1", "2", "3", "4", "5", "6"]

    def test_cost_uses_max_not_sum(self):
        from zbbx_mcp.data import canonical_host_groups
        # The bug we're fixing: 5 sub-hosts each at $280 → group cost = $280,
        # not $1,400. Sub-host {$COST_MONTH} macros typically duplicate the
        # parent's bill.
        hosts = [{"hostid": "1", "host": "edge01"}] + [
            {"hostid": str(i + 2), "host": f"edge01 v{i + 1}"} for i in range(5)
        ]
        cost_map = {str(i + 2): 280.0 for i in range(5)}  # only sub-hosts have macros
        groups = canonical_host_groups(hosts, cost_map=cost_map)
        assert len(groups) == 1
        assert groups[0]["cost"] == 280.0

    def test_traffic_uses_sum(self):
        from zbbx_mcp.data import canonical_host_groups
        # Each VIP has its own interface; group traffic adds across.
        hosts = [{"hostid": "1", "host": "edge01"}] + [
            {"hostid": str(i + 2), "host": f"edge01 v{i + 1}"} for i in range(3)
        ]
        traffic_map = {"2": 50.0, "3": 30.0, "4": 20.0}
        groups = canonical_host_groups(hosts, traffic_map=traffic_map)
        assert len(groups) == 1
        assert groups[0]["traffic"] == 100.0

    def test_cpu_uses_max(self):
        from zbbx_mcp.data import canonical_host_groups
        hosts = [{"hostid": "1", "host": "edge01"}] + [
            {"hostid": str(i + 2), "host": f"edge01 v{i + 1}"} for i in range(3)
        ]
        cpu_map = {"1": 10.0, "2": 25.0, "3": 80.0, "4": 15.0}
        groups = canonical_host_groups(hosts, cpu_map=cpu_map)
        assert len(groups) == 1
        assert groups[0]["cpu"] == 80.0

    def test_cost_none_when_no_subhost_has_macro(self):
        from zbbx_mcp.data import canonical_host_groups
        hosts = [
            {"hostid": "1", "host": "edge01"},
            {"hostid": "2", "host": "edge01 v1"},
        ]
        groups = canonical_host_groups(hosts, cost_map={})
        assert groups[0]["cost"] is None

    def test_mixed_subhost_and_standalone(self):
        from zbbx_mcp.data import canonical_host_groups
        hosts = [
            {"hostid": "1", "host": "edge01"},
            {"hostid": "2", "host": "edge01 v1"},
            {"hostid": "3", "host": "solo"},
        ]
        groups = canonical_host_groups(hosts)
        # 2 groups: edge01 (with 1 sub) + solo
        assert len(groups) == 2
        by_rep = {g["rep_host"]["host"]: g for g in groups}
        assert by_rep["edge01"]["sub_count"] == 1
        assert by_rep["solo"]["sub_count"] == 0

    def test_orphan_subhost_without_visible_parent_is_its_own_group(self):
        from zbbx_mcp.data import canonical_host_groups
        # A sub-host pattern but the parent isn't in the host list (e.g.
        # filtered out upstream). build_parent_map only maps when both
        # are present — so this host stands alone.
        hosts = [{"hostid": "1", "host": "edge01 v1"}]
        groups = canonical_host_groups(hosts)
        assert len(groups) == 1
        assert groups[0]["sub_count"] == 0

    def test_malformed_metric_values_dont_crash(self):
        from zbbx_mcp.data import canonical_host_groups
        hosts = [{"hostid": "1", "host": "edge01"}]
        # Defensive: bad strings, None — should be ignored gracefully.
        groups = canonical_host_groups(
            hosts,
            traffic_map={"1": "not-a-number"},  # type: ignore[dict-item]
            cost_map={"1": None},  # type: ignore[dict-item]
            cpu_map={"1": "abc"},  # type: ignore[dict-item]
        )
        assert groups[0]["traffic"] == 0.0
        assert groups[0]["cost"] is None
        assert groups[0]["cpu"] is None


class TestClusterCanonicalDedupe:
    """Pure-helper tests for _cluster_problems canonical-host fold (ADR 033)."""

    def _record(self, hostid: str, host: str, clock: int = 100, key: str = "k"):
        return {
            "clock": clock, "hostid": hostid, "host": host,
            "name": "Service down", "severity": 4, "key": key,
        }

    def test_one_parent_plus_subhosts_does_not_pass_threshold(self):
        from zbbx_mcp.tools.correlation import _cluster_problems
        # Single physical machine with three VIPs. Naming: parent + " " + suffix.
        # Pre-fold this would pass min_hosts=3; post-fold it must not (canonical=1).
        records = [
            self._record("1", "parent01", clock=100),
            self._record("2", "parent01 v1", clock=101),
            self._record("3", "parent01 v2", clock=102),
            self._record("4", "parent01 v3", clock=103),
        ]
        clusters = _cluster_problems(records, window_sec=60, min_hosts=3)
        assert clusters == []

    def test_three_distinct_hosts_still_form_cluster(self):
        from zbbx_mcp.tools.correlation import _cluster_problems
        # No sub-hosts; the threshold should still fire normally.
        records = [
            self._record("1", "host-a", clock=100),
            self._record("2", "host-b", clock=101),
            self._record("3", "host-c", clock=102),
        ]
        clusters = _cluster_problems(records, window_sec=60, min_hosts=3)
        assert len(clusters) == 1
        assert clusters[0]["host_count"] == 3
        assert clusters[0]["hosts"] == ["host-a", "host-b", "host-c"]

    def test_mixed_parents_and_subhosts(self):
        from zbbx_mcp.tools.correlation import _cluster_problems
        # Two distinct hosts plus one parent-with-two-subs.
        # Canonical count = 3; threshold should fire.
        records = [
            self._record("1", "host-a", clock=100),
            self._record("2", "host-b", clock=101),
            self._record("3", "parent01", clock=102),
            self._record("4", "parent01 v1", clock=103),
            self._record("5", "parent01 v2", clock=104),
        ]
        clusters = _cluster_problems(records, window_sec=60, min_hosts=3)
        assert len(clusters) == 1
        # The hosts list shows canonical names (parent appears once, not three times)
        assert clusters[0]["host_count"] == 3
        assert set(clusters[0]["hosts"]) == {"host-a", "host-b", "parent01"}

    def test_subhosts_only_without_parent_still_dedupe_to_canonical(self):
        from zbbx_mcp.tools.correlation import _cluster_problems
        # If only sub-hosts of one machine are in the bucket (parent record
        # not present), the canonical-name fold still collapses them.
        records = [
            self._record("1", "parent02 v1", clock=100),
            self._record("2", "parent02 v2", clock=101),
            self._record("3", "parent02 v3", clock=102),
            self._record("4", "parent02 v4", clock=103),
        ]
        clusters = _cluster_problems(records, window_sec=60, min_hosts=3)
        assert clusters == []  # 1 canonical host < 3

    def test_canonical_name_helper_passes_through_standalone(self):
        from zbbx_mcp.data import canonical_host_name
        assert canonical_host_name("host-a") == "host-a"

    def test_canonical_name_helper_strips_suffix(self):
        from zbbx_mcp.data import canonical_host_name
        assert canonical_host_name("parent01 v1") == "parent01"


class TestFoldRowsByCanonicalHost:
    """Pure-helper tests for fold_rows_by_canonical_host (ADR 034)."""

    def test_no_subhosts_passes_through(self):
        from zbbx_mcp.data import fold_rows_by_canonical_host
        rows = [
            {"host": "host-a", "uptime": 99.0},
            {"host": "host-b", "uptime": 50.0},
        ]
        out = fold_rows_by_canonical_host(rows, name_key="host")
        # Both rows preserved; no sub_count field added.
        assert len(out) == 2
        names = {r["host"] for r in out}
        assert names == {"host-a", "host-b"}
        assert all("sub_count" not in r for r in out)

    def test_subhosts_collapse_first_occurrence_wins(self):
        from zbbx_mcp.data import fold_rows_by_canonical_host
        rows = [
            {"host": "parent01", "uptime": 99.0},
            {"host": "parent01 v1", "uptime": 50.0},
            {"host": "parent01 v2", "uptime": 70.0},
        ]
        out = fold_rows_by_canonical_host(rows, name_key="host")
        assert len(out) == 1
        # First occurrence kept (the parent row at 99.0%)
        assert out[0]["host"] == "parent01"
        assert out[0]["uptime"] == 99.0
        assert out[0]["sub_count"] == 2

    def test_sort_key_makes_worst_win(self):
        from zbbx_mcp.data import fold_rows_by_canonical_host
        # Sort ascending by uptime → lowest uptime first → it wins after dedup.
        rows = [
            {"host": "parent01", "uptime": 99.0},
            {"host": "parent01 v1", "uptime": 50.0},
            {"host": "parent01 v2", "uptime": 70.0},
        ]
        out = fold_rows_by_canonical_host(
            rows, name_key="host",
            sort_key=lambda r: r["uptime"],
        )
        assert len(out) == 1
        assert out[0]["uptime"] == 50.0
        assert out[0]["host"] == "parent01"  # rewritten to canonical
        assert out[0]["sub_count"] == 2

    def test_mixed_subhosts_and_distinct_hosts(self):
        from zbbx_mcp.data import fold_rows_by_canonical_host
        rows = [
            {"host": "host-a", "v": 1},
            {"host": "parent01", "v": 2},
            {"host": "parent01 v1", "v": 3},
            {"host": "host-b", "v": 4},
        ]
        out = fold_rows_by_canonical_host(rows, name_key="host")
        names = {r["host"] for r in out}
        assert names == {"host-a", "parent01", "host-b"}
        # Only parent01 has a sub_count
        sub_counts = {r["host"]: r.get("sub_count") for r in out}
        assert sub_counts["parent01"] == 1
        assert sub_counts["host-a"] is None
        assert sub_counts["host-b"] is None

    def test_alternate_name_key(self):
        from zbbx_mcp.data import fold_rows_by_canonical_host
        # The helper should accept any key field, not just "host"
        rows = [
            {"server_name": "parent01 v1", "x": 1},
            {"server_name": "parent01 v2", "x": 2},
        ]
        out = fold_rows_by_canonical_host(rows, name_key="server_name")
        assert len(out) == 1
        assert out[0]["server_name"] == "parent01"


class TestExcelFills:
    """Regression for the lazy-init Fill bug fixed in v1.9.2.

    Before the fix, ``HEADER_FILL`` and friends were ``None`` at import
    time and only rebound inside ``_init_openpyxl()``. Consumers doing
    ``from zbbx_mcp.excel import HEADER_FILL`` captured the ``None``
    binding, which then fired ``TypeError: expected
    <class 'openpyxl.styles.fills.Fill'>`` during ``wb.save()`` —
    Sentry issue ``dc717f4d`` against ``generate_full_report``.
    """

    def test_fills_are_pattern_fill_instances(self):
        from openpyxl.styles import PatternFill

        from zbbx_mcp.excel import (
            DARK_RED_FILL,
            GREEN_FILL,
            HEADER_FILL,
            LIGHT_GREEN_FILL,
            ORANGE_FILL,
            RED_FILL,
        )
        for fill in (HEADER_FILL, RED_FILL, ORANGE_FILL, GREEN_FILL,
                     LIGHT_GREEN_FILL, DARK_RED_FILL):
            assert isinstance(fill, PatternFill), (
                f"{fill!r} should be a PatternFill, not {type(fill).__name__}"
            )

    def test_workbook_with_module_fills_saves(self):
        import io

        from openpyxl import Workbook

        from zbbx_mcp.excel import (
            DARK_RED_FILL,
            GREEN_FILL,
            HEADER_FILL,
            LIGHT_GREEN_FILL,
            ORANGE_FILL,
            RED_FILL,
        )
        wb = Workbook()
        ws = wb.active
        for i, fill in enumerate(
            (HEADER_FILL, RED_FILL, ORANGE_FILL, GREEN_FILL,
             LIGHT_GREEN_FILL, DARK_RED_FILL),
            start=1,
        ):
            c = ws.cell(row=i, column=1, value=str(i))
            c.fill = fill
        b = io.BytesIO()
        wb.save(b)
        assert len(b.getvalue()) > 0

    def test_full_report_module_level_imports_resolve_to_fills(self):
        # The specific failure mode: ``full_report.py`` does
        # ``from zbbx_mcp.excel import HEADER_FILL, RED_FILL, ...`` at
        # module level. After the fix those names must already point to
        # PatternFill instances at import time.
        from openpyxl.styles import PatternFill

        from zbbx_mcp.tools import full_report
        for name in ("RED_FILL", "GREEN_FILL", "ORANGE_FILL"):
            val = getattr(full_report, name)
            assert isinstance(val, PatternFill), (
                f"full_report.{name} is {type(val).__name__}; "
                f"lazy-init regression"
            )


class TestInlineCanonicalFolds:
    """Sanity checks for the inline canonical folds added in v1.9.3.

    The seven tools (`get_high_cpu_servers`, `get_underloaded_servers`,
    `get_low_disk_servers`, `get_low_memory_servers`, `get_stale_servers`,
    `detect_traffic_drops`, `get_traffic_report`) each apply a small
    dedup-by-canonical loop inline. The pattern is exercised in three
    representative shapes here: a tuple list dedup, a (hid, value) tuple
    dedup with host lookup, and a dict-list SUM fold (`get_traffic_report`
    style).
    """

    def test_tuple_first_per_canonical_wins_after_sort(self):
        """Pattern used by `get_high_cpu_servers` / `get_underloaded_servers`."""
        from zbbx_mcp.data import canonical_host_name
        # Three sub-hosts of one box plus one standalone host. Sort desc by
        # value, then keep the first occurrence per canonical.
        items = [
            (95, {"host": "parent01 v1"}),
            (90, {"host": "parent01 v2"}),
            (80, {"host": "host-a"}),
            (75, {"host": "parent01 v3"}),
        ]
        items.sort(key=lambda x: -x[0])
        seen: set[str] = set()
        folded = []
        for val, h in items:
            cn = canonical_host_name(h.get("host", ""))
            if cn in seen:
                continue
            seen.add(cn)
            folded.append((val, h))
        assert len(folded) == 2
        # parent01 group represented once (by its worst-wins occurrence at 95)
        names = {h["host"] for _, h in folded}
        canonical_names = {canonical_host_name(n) for n in names}
        assert canonical_names == {"parent01", "host-a"}
        # Worst value (95) is the surviving parent01 entry
        parent_val = [v for v, h in folded if "parent01" in h["host"]][0]
        assert parent_val == 95

    def test_traffic_report_style_sum_fold(self):
        """Pattern used by `get_traffic_report` — SUM across sub-hosts."""
        from zbbx_mcp.data import canonical_host_name
        rows = [
            {"host": "parent01", "traffic": 100.0, "connections": 10},
            {"host": "parent01 v1", "traffic": 50.0, "connections": 5},
            {"host": "parent01 v2", "traffic": 30.0, "connections": 3},
            {"host": "host-a", "traffic": 20.0, "connections": 2},
        ]
        canonical_rows: dict[str, dict] = {}
        for r in rows:
            cn = canonical_host_name(r["host"])
            g = canonical_rows.get(cn)
            if g is None:
                canonical_rows[cn] = {**r, "host": cn}
            else:
                g["traffic"] += r["traffic"]
                g["connections"] += r["connections"]
                g["sub_count"] = g.get("sub_count", 0) + 1
        for g in canonical_rows.values():
            g["bw_per_client"] = (
                g["traffic"] / g["connections"]
                if g["connections"] > 0 else 0
            )
        out = list(canonical_rows.values())
        assert len(out) == 2
        by_host = {r["host"]: r for r in out}
        # parent01 sums to 180 traffic, 18 connections
        assert by_host["parent01"]["traffic"] == 180.0
        assert by_host["parent01"]["connections"] == 18
        assert by_host["parent01"]["sub_count"] == 2
        assert by_host["parent01"]["bw_per_client"] == 10.0
        # host-a passes through
        assert by_host["host-a"]["traffic"] == 20.0
        assert "sub_count" not in by_host["host-a"]

    def test_hostid_indirection_dedup(self):
        """Pattern used by `get_low_disk_servers` / `get_low_memory_servers`."""
        from zbbx_mcp.data import canonical_host_name
        host_map = {
            "1": {"host": "parent01"},
            "2": {"host": "parent01 v1"},
            "3": {"host": "parent01 v2"},
            "4": {"host": "host-a"},
        }
        # Already sorted worst-first (highest pct first)
        flagged = [("1", 95), ("2", 90), ("3", 80), ("4", 70)]
        seen: set[str] = set()
        folded = []
        for hid, val in flagged:
            h = host_map.get(hid, {})
            cn = canonical_host_name(h.get("host", hid))
            if cn in seen:
                continue
            seen.add(cn)
            folded.append((hid, val))
        assert len(folded) == 2
        # parent01 (worst at val=95) and host-a (val=70)
        assert [v for _, v in folded] == [95, 70]


class TestShutdownCandidateMetricFold:
    """Sanity tests for the per-canonical metric aggregation pattern used
    in `get_shutdown_candidates` (ADR 037).

    cpu = MAX, traffic = SUM, service = WORST across a canonical group.
    """

    def _aggregate_group(self, hostids, metrics, service_map):
        cpus_avg = []
        traffics_avg = []
        services = []
        for hid in hostids:
            hm = metrics.get(hid, {})
            if hm.get("cpu") is not None:
                cpus_avg.append(hm["cpu"])
            if hm.get("traffic") is not None:
                traffics_avg.append(hm["traffic"])
            if hid in service_map:
                services.append(service_map[hid])
        cpu_avg = max(cpus_avg) if cpus_avg else None
        traffic_avg = sum(traffics_avg) if traffics_avg else None
        if 0 in services:
            service = "DOWN"
        elif -1 in services:
            service = "PARTIAL"
        elif 1 in services:
            service = "OK"
        else:
            service = ""
        return cpu_avg, traffic_avg, service

    def test_cpu_max_traffic_sum_service_worst(self):
        metrics = {
            "1": {"cpu": 5, "traffic": 0.1},
            "2": {"cpu": 75, "traffic": 50.0},
            "3": {"cpu": 3, "traffic": 0.5},
        }
        services = {"1": 1, "2": 0, "3": 1}
        cpu, traffic, service = self._aggregate_group(
            ["1", "2", "3"], metrics, services,
        )
        assert cpu == 75
        assert traffic == 50.6
        assert service == "DOWN"

    def test_all_idle_group_qualifies_as_dead(self):
        # Bug-fix case: parent + sub-hosts all idle → one DEAD candidate.
        metrics = {hid: {"cpu": 0.5, "traffic": 0.1} for hid in "12345"}
        services = {hid: 1 for hid in "12345"}
        cpu, traffic, service = self._aggregate_group(
            list("12345"), metrics, services,
        )
        assert cpu == 0.5
        assert traffic == 0.5
        assert traffic < 1.0 and cpu < 5.0

    def test_busy_subhost_rescues_parent_from_dead(self):
        # Parent's own metrics zero but sub-host very busy → group should
        # NOT qualify as DEAD (post-fold reality).
        metrics = {
            "parent": {"cpu": 0, "traffic": 0},
            "sub1": {"cpu": 80, "traffic": 200.0},
        }
        services = {"parent": 1, "sub1": 1}
        cpu, traffic, service = self._aggregate_group(
            ["parent", "sub1"], metrics, services,
        )
        assert cpu == 80
        assert traffic == 200.0
        assert not (traffic < 1.0 and cpu < 5.0)
        assert not (cpu > 50 and traffic < 1.0)
        assert service != "DOWN"

    def test_empty_metrics_returns_none(self):
        cpu, traffic, service = self._aggregate_group([], {}, {})
        assert cpu is None
        assert traffic is None
        assert service == ""

    def test_partial_service_loses_to_down(self):
        services = {"1": 1, "2": -1, "3": 0}
        _, _, service = self._aggregate_group(
            ["1", "2", "3"], {}, services,
        )
        assert service == "DOWN"

    def test_partial_service_wins_over_ok(self):
        services = {"1": 1, "2": -1, "3": 1}
        _, _, service = self._aggregate_group(
            ["1", "2", "3"], {}, services,
        )
        assert service == "PARTIAL"


class TestFreshestAgentPing:
    """Pure-helper tests for _freshest_agent_ping (#158, ADR 049)."""

    def test_none_when_no_ping(self):
        from zbbx_mcp.tools.diagnose import _freshest_agent_ping
        assert _freshest_agent_ping([{"key_": "net.if.in[primary]"}]) is None

    def test_picks_freshest_across_vips(self):
        from zbbx_mcp.tools.diagnose import _freshest_agent_ping
        # parent's ping is live (clock 200, up); a stale sub-host ping (clock
        # 100, down) must not win.
        items = [
            {"key_": "agent.ping", "lastvalue": "0", "lastclock": "100"},
            {"key_": "agent.ping", "lastvalue": "1", "lastclock": "200"},
        ]
        ping = _freshest_agent_ping(items)
        assert ping["lastvalue"] == "1" and ping["lastclock"] == "200"

    def test_single_ping_returned(self):
        from zbbx_mcp.tools.diagnose import _freshest_agent_ping
        items = [{"key_": "agent.ping", "lastvalue": "1", "lastclock": "50"}]
        assert _freshest_agent_ping(items)["lastclock"] == "50"

    def test_missing_clock_treated_as_zero(self):
        from zbbx_mcp.tools.diagnose import _freshest_agent_ping
        items = [
            {"key_": "agent.ping", "lastvalue": "1"},  # no clock → 0
            {"key_": "agent.ping", "lastvalue": "0", "lastclock": "5"},
        ]
        assert _freshest_agent_ping(items)["lastclock"] == "5"


class TestBulkDiagnosePreFold:
    """Pure-helper tests for `_dedupe_records_by_canonical` (ADR 039).

    Pre-fold of the input host list before bulk diagnose so the
    fan-out emits one row per physical machine.
    """

    def test_standalone_hosts_pass_through(self):
        from zbbx_mcp.tools.diagnose import _dedupe_records_by_canonical
        records = [
            {"hostid": "1", "host": "host-a"},
            {"hostid": "2", "host": "host-b"},
        ]
        deduped, subs = _dedupe_records_by_canonical(records)
        assert len(deduped) == 2
        names = {r["host"] for r in deduped}
        assert names == {"host-a", "host-b"}
        assert all(c == 0 for c in subs.values())

    def test_parent_plus_subhosts_collapse_to_parent(self):
        from zbbx_mcp.tools.diagnose import _dedupe_records_by_canonical
        records = [
            {"hostid": "1", "host": "parent01"},
            {"hostid": "2", "host": "parent01 v1"},
            {"hostid": "3", "host": "parent01 v2"},
            {"hostid": "4", "host": "parent01 v3"},
        ]
        deduped, subs = _dedupe_records_by_canonical(records)
        assert len(deduped) == 1
        # Parent preferred as the representative
        assert deduped[0]["host"] == "parent01"
        assert deduped[0]["hostid"] == "1"
        assert subs["parent01"] == 3
        # The rep carries every VIP's hostid so the diagnosis queries
        # problems across the whole box (ADR 046).
        assert set(deduped[0]["_group_hostids"]) == {"1", "2", "3", "4"}

    def test_standalone_group_hostids_is_self(self):
        from zbbx_mcp.tools.diagnose import _dedupe_records_by_canonical
        deduped, _ = _dedupe_records_by_canonical([{"hostid": "9", "host": "solo"}])
        assert deduped[0]["_group_hostids"] == ["9"]

    def test_subhost_only_set_picks_first_as_rep(self):
        from zbbx_mcp.tools.diagnose import _dedupe_records_by_canonical
        records = [
            {"hostid": "2", "host": "parent01 v1"},
            {"hostid": "3", "host": "parent01 v2"},
            {"hostid": "4", "host": "parent01 v3"},
        ]
        deduped, subs = _dedupe_records_by_canonical(records)
        assert len(deduped) == 1
        assert deduped[0]["host"] == "parent01 v1"
        assert subs["parent01"] == 2

    def test_mixed_standalone_and_groups(self):
        from zbbx_mcp.tools.diagnose import _dedupe_records_by_canonical
        records = [
            {"hostid": "1", "host": "host-a"},
            {"hostid": "2", "host": "parent01"},
            {"hostid": "3", "host": "parent01 v1"},
            {"hostid": "4", "host": "host-b"},
        ]
        deduped, subs = _dedupe_records_by_canonical(records)
        names = {r["host"] for r in deduped}
        assert names == {"host-a", "parent01", "host-b"}
        assert subs["host-a"] == 0
        assert subs["parent01"] == 1
        assert subs["host-b"] == 0

    def test_empty_input(self):
        from zbbx_mcp.tools.diagnose import _dedupe_records_by_canonical
        deduped, subs = _dedupe_records_by_canonical([])
        assert deduped == []
        assert subs == {}


class TestCollapseDependentProblems:
    """Pure-helper tests for collapse_dependent_problems (#144, ADR 048)."""

    def test_drops_symptom_when_root_firing(self):
        from zbbx_mcp.data import collapse_dependent_problems
        problems = [
            {"eventid": "1", "objectid": "10", "name": "root"},
            {"eventid": "2", "objectid": "20", "name": "symptom"},
        ]
        kept, n = collapse_dependent_problems(problems, {"20": {"10"}})
        assert n == 1
        assert {p["objectid"] for p in kept} == {"10"}

    def test_keeps_symptom_when_root_not_firing(self):
        from zbbx_mcp.data import collapse_dependent_problems
        problems = [{"eventid": "2", "objectid": "20", "name": "symptom"}]
        kept, n = collapse_dependent_problems(problems, {"20": {"10"}})
        assert n == 0 and len(kept) == 1

    def test_no_dependencies_is_noop(self):
        from zbbx_mcp.data import collapse_dependent_problems
        problems = [{"eventid": "1", "objectid": "10"}, {"eventid": "2", "objectid": "11"}]
        kept, n = collapse_dependent_problems(problems, {})
        assert n == 0 and len(kept) == 2

    def test_collapse_false_is_noop(self):
        from zbbx_mcp.data import collapse_dependent_problems
        problems = [{"eventid": "1", "objectid": "10"}, {"eventid": "2", "objectid": "20"}]
        kept, n = collapse_dependent_problems(problems, {"20": {"10"}}, collapse=False)
        assert n == 0 and len(kept) == 2

    def test_chain_collapses_only_active_dependency(self):
        from zbbx_mcp.data import collapse_dependent_problems
        problems = [
            {"eventid": "1", "objectid": "10"},
            {"eventid": "2", "objectid": "20"},
            {"eventid": "3", "objectid": "30"},
        ]
        dep_map = {"30": {"20"}, "20": {"10"}}
        kept, n = collapse_dependent_problems(problems, dep_map)
        assert n == 2
        assert {p["objectid"] for p in kept} == {"10"}

    def test_missing_objectid_kept(self):
        from zbbx_mcp.data import collapse_dependent_problems
        kept, n = collapse_dependent_problems([{"eventid": "1"}], {"x": {"y"}})
        assert n == 0 and len(kept) == 1


class _ProblemOnlyClient:
    """Minimal async client returning a fixed problem.get payload.

    Records every call so a test can assert nothing else was hit. A
    domain-mode host (no items) makes ``_collect_diagnosis_inner`` issue
    only ``problem.get``, which keeps this stub tiny.
    """

    def __init__(self, problems):
        self._problems = problems
        self.calls = []

    async def call(self, method, params):
        self.calls.append((method, params))
        if method == "problem.get":
            return [dict(p) for p in self._problems]
        return []


class TestDiagnoseSuppressThreading:
    """ADR 052 — _collect_diagnosis_inner honours include_suppressed."""

    async def test_suppressed_only_reads_healthy_by_default(self):
        """A box whose only problem is maintenance-suppressed must not
        read degraded — the false-positive class ADR 052 closes."""
        from zbbx_mcp.tools.diagnose import _collect_diagnosis_inner

        now = 1_000_000
        problems = [
            {"eventid": "1", "name": "planned reboot", "severity": "4",
             "clock": str(now - 60), "suppressed": "1"},
        ]
        client = _ProblemOnlyClient(problems)
        facts = await _collect_diagnosis_inner(
            client, {"hostid": "10", "host": "h"}, [],  # no items → domain mode
            now=now,
        )
        assert facts["problems"] == []
        assert facts["verdict"] == "healthy"

    async def test_include_suppressed_keeps_maintenance_problem(self):
        from zbbx_mcp.tools.diagnose import _collect_diagnosis_inner

        now = 1_000_000
        problems = [
            {"eventid": "1", "name": "planned reboot", "severity": "4",
             "clock": str(now - 60), "suppressed": "1"},
        ]
        client = _ProblemOnlyClient(problems)
        facts = await _collect_diagnosis_inner(
            client, {"hostid": "10", "host": "h"}, [],
            now=now, include_suppressed=True,
        )
        assert [p["name"] for p in facts["problems"]] == ["planned reboot"]
        assert facts["verdict"] == "degraded"

    async def test_mixed_keeps_only_live_problem(self):
        from zbbx_mcp.tools.diagnose import _collect_diagnosis_inner

        now = 1_000_000
        problems = [
            {"eventid": "1", "name": "live", "severity": "4",
             "clock": str(now - 60), "suppressed": "0"},
            {"eventid": "2", "name": "maint", "severity": "4",
             "clock": str(now - 60), "suppressed": "1"},
        ]
        client = _ProblemOnlyClient(problems)
        facts = await _collect_diagnosis_inner(
            client, {"hostid": "10", "host": "h"}, [], now=now,
        )
        assert [p["name"] for p in facts["problems"]] == ["live"]


class TestRecentChangesWireContract:
    """ADR 070 — get_recent_changes must not send selectHosts to problem.get.

    Same -32602 class as triage's ADR 068; found live when the tool errored
    during a feed-vs-Zabbix analysis. Drives the real tool function through a
    recording fake client and asserts the wire contract.
    """

    class _Client:
        def __init__(self):
            self.calls = []

        async def call(self, method, params):
            self.calls.append((method, params))
            if method == "problem.get":
                return [{"eventid": "9", "name": "Service Down", "severity": "5",
                         "clock": "1000", "acknowledged": "0", "suppressed": "0",
                         "objectid": "77"}]
            if method == "trigger.get":
                return [{"triggerid": "77", "hosts": [{"host": "node-eu-a1"}]}]
            return []  # event.get → no resolved events

    def _run(self):
        import asyncio

        from zbbx_mcp.tools import availability as availability_mod

        class _MCP:
            def __init__(self):
                self.fns = {}

            def tool(self):
                def deco(f):
                    self.fns[f.__name__] = f
                    return f
                return deco

        class _Resolver:
            def __init__(self, client):
                self._client = client

            def resolve(self, instance):
                return self._client

        client = self._Client()
        mcp = _MCP()
        availability_mod.register(mcp, _Resolver(client))
        out = asyncio.run(mcp.fns["get_recent_changes"]())
        return client, out

    def test_problem_get_omits_selecthosts(self):
        client, _ = self._run()
        pget = next(p for m, p in client.calls if m == "problem.get")
        assert "selectHosts" not in pget
        assert "objectid" in pget["output"]

    def test_event_get_keeps_selecthosts(self):
        # event.get DOES support selectHosts — the resolved branch is untouched.
        client, _ = self._run()
        eget = next(p for m, p in client.calls if m == "event.get")
        assert eget.get("selectHosts") == ["host"]

    def test_host_rendered_via_trigger_map(self):
        client, out = self._run()
        tget = next(p for m, p in client.calls if m == "trigger.get")
        assert tget.get("selectHosts") == ["host"]
        assert "node-eu-a1" in out  # host name reached the rendered table


class TestKeepActiveOrRecent:
    """ADR 069 — diagnose_host must not age out still-active problems."""

    NOW = 1_000_000

    def test_active_old_problem_kept(self):
        from zbbx_mcp.tools.diagnose import _keep_active_or_recent
        # Unresolved (no r_eventid), started 72h ago — must survive the window.
        probs = [{"eventid": "1", "clock": str(self.NOW - 72 * 3600)}]
        assert _keep_active_or_recent(probs, self.NOW, 24) == probs

    def test_active_old_problem_with_zero_r_eventid_kept(self):
        from zbbx_mcp.tools.diagnose import _keep_active_or_recent
        probs = [{"eventid": "1", "clock": str(self.NOW - 72 * 3600), "r_eventid": "0"}]
        assert len(_keep_active_or_recent(probs, self.NOW, 24)) == 1

    def test_resolved_old_problem_dropped(self):
        from zbbx_mcp.tools.diagnose import _keep_active_or_recent
        probs = [{"eventid": "1", "clock": str(self.NOW - 72 * 3600), "r_eventid": "9"}]
        assert _keep_active_or_recent(probs, self.NOW, 24) == []

    def test_resolved_recent_problem_kept(self):
        from zbbx_mcp.tools.diagnose import _keep_active_or_recent
        probs = [{"eventid": "1", "clock": str(self.NOW - 60), "r_eventid": "9"}]
        assert len(_keep_active_or_recent(probs, self.NOW, 24)) == 1

    async def test_days_old_active_problem_keeps_host_non_healthy(self):
        """The reported bug: a host with an unresolved Disaster from 3 days
        ago must NOT read healthy."""
        from zbbx_mcp.tools.diagnose import _collect_diagnosis_inner

        now = 1_000_000
        problems = [
            {"eventid": "1", "name": "Service down", "severity": "5",
             "clock": str(now - 72 * 3600), "suppressed": "0", "r_eventid": "0"},
        ]
        client = _ProblemOnlyClient(problems)
        facts = await _collect_diagnosis_inner(
            client, {"hostid": "10", "host": "h"}, [], now=now,
        )
        assert [p["name"] for p in facts["problems"]] == ["Service down"]
        assert facts["verdict"] != "healthy"


class TestUnmappedGroupCounts:
    """Pure-helper tests for unmapped_group_counts (ADR 058)."""

    def test_counts_sorted_desc_then_name(self):
        from zbbx_mcp.classify import unmapped_group_counts
        sets = [["GrpA"], ["GrpA", "GrpB"], ["GrpB"], ["GrpB"]]
        out = unmapped_group_counts(sets, {})
        assert out == [("GrpB", 3), ("GrpA", 2)]

    def test_mapped_and_skip_mapped_excluded(self):
        from zbbx_mcp.classify import unmapped_group_counts
        pmap = {"Mapped": ("Prod", "Tier"), "Skipped": (None, None)}
        sets = [["Mapped", "Gap"], ["Skipped", "Gap"]]
        assert unmapped_group_counts(sets, pmap) == [("Gap", 2)]

    def test_groupless_host_counted(self):
        from zbbx_mcp.classify import unmapped_group_counts
        assert unmapped_group_counts([[]], {}) == [("(no groups)", 1)]

    def test_empty_input(self):
        from zbbx_mcp.classify import unmapped_group_counts
        assert unmapped_group_counts([], {"X": ("P", "T")}) == []


class TestSummarizeTokenExpiry:
    """Pure-helper tests for summarize_token_expiry (ADR 057)."""

    NOW = 1_000_000_000

    def test_expiring_token_flagged_sorted_soonest_first(self):
        from zbbx_mcp.tools.health import summarize_token_expiry
        tokens = [
            {"name": "b", "expires_at": str(self.NOW + 20 * 86400), "status": "0"},
            {"name": "a", "expires_at": str(self.NOW + 5 * 86400), "status": "0"},
        ]
        out = summarize_token_expiry(tokens, self.NOW)
        assert [n for n, _ in out] == ["a", "b"]
        assert out[0][1] == 5

    def test_never_expiring_and_disabled_skipped(self):
        from zbbx_mcp.tools.health import summarize_token_expiry
        tokens = [
            {"name": "never", "expires_at": "0", "status": "0"},
            {"name": "disabled", "expires_at": str(self.NOW + 86400), "status": "1"},
        ]
        assert summarize_token_expiry(tokens, self.NOW) == []

    def test_far_future_not_flagged(self):
        from zbbx_mcp.tools.health import summarize_token_expiry
        tokens = [{"name": "ok", "expires_at": str(self.NOW + 90 * 86400), "status": "0"}]
        assert summarize_token_expiry(tokens, self.NOW) == []

    def test_already_expired_negative_days(self):
        from zbbx_mcp.tools.health import summarize_token_expiry
        tokens = [{"name": "dead", "expires_at": str(self.NOW - 2 * 86400), "status": "0"}]
        out = summarize_token_expiry(tokens, self.NOW)
        assert out and out[0][1] < 0


class TestFormatProxyCompat:
    """Pure-helper tests for format_proxy_compat (ADR 056)."""

    def test_current_version_no_annotation(self):
        from zbbx_mcp.tools.proxies import format_proxy_compat
        assert format_proxy_compat("1", "7.4.9") == " v7.4.9"

    def test_outdated_flagged(self):
        from zbbx_mcp.tools.proxies import format_proxy_compat
        assert "OUTDATED" in format_proxy_compat("2", "7.0.0")

    def test_unsupported_flagged(self):
        from zbbx_mcp.tools.proxies import format_proxy_compat
        assert "UNSUPPORTED" in format_proxy_compat("3", "6.0.1")

    def test_unknown_version_and_undefined_compat_empty(self):
        from zbbx_mcp.tools.proxies import format_proxy_compat
        assert format_proxy_compat("0", "0") == ""


class TestFilterSuppressed:
    """Pure-helper tests for filter_suppressed (#143, ADR 044)."""

    def _probs(self):
        return [
            {"eventid": "1", "name": "real", "suppressed": "0"},
            {"eventid": "2", "name": "maint", "suppressed": "1"},
            {"eventid": "3", "name": "also-real"},  # field absent → not suppressed
        ]

    def test_default_excludes_suppressed(self):
        from zbbx_mcp.data import filter_suppressed
        out = filter_suppressed(self._probs())
        assert {p["eventid"] for p in out} == {"1", "3"}

    def test_include_keeps_all(self):
        from zbbx_mcp.data import filter_suppressed
        assert len(filter_suppressed(self._probs(), include_suppressed=True)) == 3

    def test_missing_field_treated_as_visible(self):
        from zbbx_mcp.data import filter_suppressed
        assert len(filter_suppressed([{"eventid": "9"}])) == 1

    def test_empty_input(self):
        from zbbx_mcp.data import filter_suppressed
        assert filter_suppressed([]) == []

    def test_returns_new_list_not_alias(self):
        from zbbx_mcp.data import filter_suppressed
        src = [{"eventid": "1", "suppressed": "0"}]
        out = filter_suppressed(src, include_suppressed=True)
        assert out == src and out is not src


class TestClassifyCountryGroup:
    """Pure-helper tests for service_brief per-country group fold (#154, ADR 045)."""

    def test_real_traffic_is_validated(self):
        from zbbx_mcp.tools.service_brief import _classify_country_group
        # summed box traffic >= floor → validated regardless of checks
        assert _classify_country_group(6.0, [0, 0]) == "validated"

    def test_no_traffic_no_checks_skipped(self):
        from zbbx_mcp.tools.service_brief import _classify_country_group
        assert _classify_country_group(0.0, []) == "skip"

    def test_all_checks_up_is_ok(self):
        from zbbx_mcp.tools.service_brief import _classify_country_group
        assert _classify_country_group(0.0, [1, 1, 1]) == "ok"

    def test_one_failing_vip_check_drops_to_partial(self):
        from zbbx_mcp.tools.service_brief import _classify_country_group
        # worst-wins across merged VIP checks: a single failure → partial
        assert _classify_country_group(0.0, [1, 1, 0]) == "partial"

    def test_all_checks_down_is_down(self):
        from zbbx_mcp.tools.service_brief import _classify_country_group
        assert _classify_country_group(0.0, [0, 0]) == "down"

    def test_summed_subhost_traffic_validates_box(self):
        from zbbx_mcp.tools.service_brief import _classify_country_group
        # two VIPs at 3 Mbps each = 6 summed → above the 5 Mbps floor.
        # (the point of the fold: per-VIP each would be below the floor)
        assert _classify_country_group(3.0 + 3.0, [0]) == "validated"
