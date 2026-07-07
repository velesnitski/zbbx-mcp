"""Health, telemetry, stale-gate, and misc formatter tests (split from test_analytics, ADR 074)."""



class TestServiceCheckStaleGate:
    """Pure-helper tests for is_service_check_stale (#130)."""

    def test_state_one_is_stale(self):
        from zbbx_mcp.fetch import is_service_check_stale

        # state=1 means Zabbix flagged the item unsupported.
        item = {"state": "1", "lastclock": str(1_700_000_000), "lastvalue": "0"}
        assert is_service_check_stale(item, now=1_700_000_300) is True

    def test_lastclock_inside_window_is_fresh(self):
        from zbbx_mcp.fetch import is_service_check_stale

        item = {"state": "0", "lastclock": str(1_700_000_000), "lastvalue": "1"}
        # 5 minutes old — well within the default 30min window.
        assert is_service_check_stale(item, now=1_700_000_300) is False

    def test_lastclock_outside_window_is_stale(self):
        from zbbx_mcp.fetch import is_service_check_stale

        item = {"state": "0", "lastclock": str(1_700_000_000), "lastvalue": "1"}
        # 31 minutes old — past the default 30min window.
        assert is_service_check_stale(item, now=1_700_000_000 + 31 * 60) is True

    def test_zero_lastclock_is_stale(self):
        from zbbx_mcp.fetch import is_service_check_stale

        # Item never polled → lastclock=0. Treat as stale.
        assert is_service_check_stale(
            {"state": "0", "lastclock": "0"}, now=1_700_000_300,
        ) is True

    def test_missing_lastclock_is_stale(self):
        from zbbx_mcp.fetch import is_service_check_stale

        assert is_service_check_stale(
            {"state": "0"}, now=1_700_000_300,
        ) is True

    def test_garbage_lastclock_is_stale(self):
        from zbbx_mcp.fetch import is_service_check_stale

        assert is_service_check_stale(
            {"state": "0", "lastclock": "not-a-number"}, now=1_700_000_300,
        ) is True

    def test_custom_stale_window(self):
        from zbbx_mcp.fetch import is_service_check_stale

        item = {"state": "0", "lastclock": str(1_700_000_000), "lastvalue": "1"}
        # 5 minutes old; tighten the window to 60s and it becomes stale.
        assert is_service_check_stale(item, now=1_700_000_300, stale_sec=60) is True
        # Loosen to 1h and the same item is fresh.
        assert is_service_check_stale(item, now=1_700_000_300, stale_sec=3600) is False

    def test_state_one_overrides_fresh_lastclock(self):
        from zbbx_mcp.fetch import is_service_check_stale

        # Even a recent lastclock cannot rescue an unsupported item.
        item = {"state": "1", "lastclock": str(1_700_000_290), "lastvalue": "0"}
        assert is_service_check_stale(item, now=1_700_000_300) is True

class TestStaleItemsCascade:
    """Pure-helper tests for cascade collapse (#133)."""

    def test_no_master_passes_through(self):
        from zbbx_mcp.tools.items import _collapse_dependent_chain

        stale = [{"itemid": "i1", "master_itemid": ""}]
        out = _collapse_dependent_chain(stale)
        assert len(out) == 1
        assert out[0]["affected_count"] == 0

    def test_child_with_stale_master_collapsed(self):
        from zbbx_mcp.tools.items import _collapse_dependent_chain

        stale = [
            {"itemid": "i1", "master_itemid": ""},   # root
            {"itemid": "i2", "master_itemid": "i1"}, # child of root
        ]
        out = _collapse_dependent_chain(stale)
        assert len(out) == 1
        assert out[0]["itemid"] == "i1"
        assert out[0]["affected_count"] == 1

    def test_child_with_non_stale_master_kept(self):
        from zbbx_mcp.tools.items import _collapse_dependent_chain

        # Child references master_itemid that is NOT in the stale list —
        # treat the child as its own root (its master is healthy).
        stale = [{"itemid": "i2", "master_itemid": "i_healthy"}]
        out = _collapse_dependent_chain(stale)
        assert len(out) == 1
        assert out[0]["itemid"] == "i2"
        assert out[0]["affected_count"] == 0

    def test_two_hop_chain_collapses_to_root(self):
        from zbbx_mcp.tools.items import _collapse_dependent_chain

        stale = [
            {"itemid": "root", "master_itemid": ""},
            {"itemid": "mid", "master_itemid": "root"},
            {"itemid": "leaf", "master_itemid": "mid"},
        ]
        out = _collapse_dependent_chain(stale)
        assert len(out) == 1
        assert out[0]["itemid"] == "root"
        assert out[0]["affected_count"] == 2  # mid + leaf

    def test_multiple_children_share_one_root(self):
        from zbbx_mcp.tools.items import _collapse_dependent_chain

        stale = [
            {"itemid": "root", "master_itemid": ""},
            {"itemid": "c1", "master_itemid": "root"},
            {"itemid": "c2", "master_itemid": "root"},
            {"itemid": "c3", "master_itemid": "root"},
        ]
        out = _collapse_dependent_chain(stale)
        assert len(out) == 1
        assert out[0]["affected_count"] == 3

    def test_circular_reference_does_not_loop(self):
        from zbbx_mcp.tools.items import _collapse_dependent_chain

        # Pathological: A → B → A. The root_of walk must terminate.
        stale = [
            {"itemid": "a", "master_itemid": "b"},
            {"itemid": "b", "master_itemid": "a"},
        ]
        out = _collapse_dependent_chain(stale)
        # Both end up rooted at one of the cycle members; count should be
        # bounded and the function must return.
        assert sum(s.get("affected_count", 0) for s in out) >= 0

    def test_input_not_mutated(self):
        from zbbx_mcp.tools.items import _collapse_dependent_chain

        stale = [{"itemid": "i1", "master_itemid": ""}]
        original = dict(stale[0])
        _ = _collapse_dependent_chain(stale)
        assert stale[0] == original

class TestTelemetrySummary:
    """Pure-helper tests for _summarise_records (#7)."""

    def _rec(self, tool="x", status="ok", duration_ms=10, response_size=100, ts=None):
        r = {
            "tool": tool,
            "status": status,
            "duration_ms": duration_ms,
            "response_size": response_size,
        }
        if ts is not None:
            r["ts"] = ts
        return r

    def test_per_tool_counts_and_avg(self):
        from zbbx_mcp.tools.telemetry import _summarise_records

        records = [
            self._rec(tool="search_hosts", duration_ms=20),
            self._rec(tool="search_hosts", duration_ms=40),
            self._rec(tool="get_problems", duration_ms=200),
        ]
        out = _summarise_records(records)
        by_tool = {r["tool"]: r for r in out}
        assert by_tool["search_hosts"]["calls"] == 2
        assert by_tool["search_hosts"]["avg_ms"] == 30.0
        assert by_tool["get_problems"]["calls"] == 1
        assert by_tool["get_problems"]["avg_ms"] == 200.0

    def test_error_rate_pct(self):
        from zbbx_mcp.tools.telemetry import _summarise_records

        records = (
            [self._rec(status="ok") for _ in range(7)]
            + [self._rec(status="error") for _ in range(3)]
        )
        out = _summarise_records(records)
        assert out[0]["errors"] == 3
        assert out[0]["error_pct"] == 30.0

    def test_sorted_by_calls_desc(self):
        from zbbx_mcp.tools.telemetry import _summarise_records

        records = (
            [self._rec(tool="a") for _ in range(2)]
            + [self._rec(tool="b") for _ in range(5)]
            + [self._rec(tool="c") for _ in range(3)]
        )
        out = _summarise_records(records)
        assert [r["tool"] for r in out] == ["b", "c", "a"]

    def test_max_ms_tracked(self):
        from zbbx_mcp.tools.telemetry import _summarise_records

        records = [
            self._rec(duration_ms=10),
            self._rec(duration_ms=500),
            self._rec(duration_ms=50),
        ]
        out = _summarise_records(records)
        assert out[0]["max_ms"] == 500

    def test_garbage_duration_treated_as_zero(self):
        from zbbx_mcp.tools.telemetry import _summarise_records

        records = [
            {"tool": "x", "status": "ok", "duration_ms": "not-a-number"},
            {"tool": "x", "status": "ok", "duration_ms": 100},
        ]
        out = _summarise_records(records)
        assert out[0]["calls"] == 2
        assert out[0]["avg_ms"] == 50.0  # (0 + 100) / 2

    def test_since_ts_filter_drops_old_records(self):
        from zbbx_mcp.tools.telemetry import _summarise_records

        records = [
            self._rec(tool="a", ts=1000),
            self._rec(tool="a", ts=2000),
            self._rec(tool="a", ts=3000),
        ]
        out = _summarise_records(records, since_ts=2000)
        assert out[0]["calls"] == 2  # 2000 and 3000 kept; 1000 dropped

    def test_iso_timestamp_filter(self):
        from zbbx_mcp.tools.telemetry import _summarise_records

        records = [
            self._rec(tool="a", ts="2026-05-04T12:00:00Z"),
            self._rec(tool="a", ts="2026-05-05T12:00:00Z"),
        ]
        # 2026-05-05 00:00 UTC = 1777939200
        out = _summarise_records(records, since_ts=1777939200)
        assert out[0]["calls"] == 1  # only the May 5 record

class TestTagFilterParser:
    """Pure-helper tests for parse_tag_filter (#145)."""

    def test_empty_returns_empty_list(self):
        from zbbx_mcp.tag_filter import parse_tag_filter
        assert parse_tag_filter("") == []
        assert parse_tag_filter("   ") == []

    def test_single_key_value_equals(self):
        from zbbx_mcp.tag_filter import parse_tag_filter
        assert parse_tag_filter("role:edge") == [
            {"tag": "role", "value": "edge", "operator": 0}
        ]

    def test_multiple_pairs_and_combined(self):
        from zbbx_mcp.tag_filter import parse_tag_filter
        out = parse_tag_filter("role:edge,env:prod")
        assert out == [
            {"tag": "role", "value": "edge", "operator": 0},
            {"tag": "env", "value": "prod", "operator": 0},
        ]

    def test_whitespace_tolerated(self):
        from zbbx_mcp.tag_filter import parse_tag_filter
        out = parse_tag_filter("role:edge ,  env:prod  ")
        assert out == [
            {"tag": "role", "value": "edge", "operator": 0},
            {"tag": "env", "value": "prod", "operator": 0},
        ]

    def test_bare_key_means_exists(self):
        from zbbx_mcp.tag_filter import parse_tag_filter
        assert parse_tag_filter("role") == [
            {"tag": "role", "value": "", "operator": 4}
        ]

    def test_empty_value_after_colon_means_exists(self):
        from zbbx_mcp.tag_filter import parse_tag_filter
        assert parse_tag_filter("role:") == [
            {"tag": "role", "value": "", "operator": 4}
        ]

    def test_empty_key_is_skipped(self):
        from zbbx_mcp.tag_filter import parse_tag_filter
        assert parse_tag_filter(":value,") == []

    def test_trailing_comma_does_not_break(self):
        from zbbx_mcp.tag_filter import parse_tag_filter
        assert parse_tag_filter("role:edge,") == [
            {"tag": "role", "value": "edge", "operator": 0}
        ]

class TestZabbixVersionHelpers:
    """Pure-helper tests for version parsing + feature matrix."""

    def test_parse_standard_version(self):
        from zbbx_mcp.tools.health import _parse_zabbix_version
        assert _parse_zabbix_version("6.4.2") == (6, 4, 2)

    def test_parse_two_part_version(self):
        from zbbx_mcp.tools.health import _parse_zabbix_version
        assert _parse_zabbix_version("7.0") == (7, 0, 0)

    def test_parse_empty_returns_zeros(self):
        from zbbx_mcp.tools.health import _parse_zabbix_version
        assert _parse_zabbix_version("") == (0, 0, 0)

    def test_parse_garbage_returns_zeros(self):
        from zbbx_mcp.tools.health import _parse_zabbix_version
        assert _parse_zabbix_version("not-a-version") == (0, 0, 0)

    def test_parse_partial_garbage(self):
        from zbbx_mcp.tools.health import _parse_zabbix_version
        # "6.x.2" — major parses, minor is garbage → stop there
        assert _parse_zabbix_version("6.x.2") == (6, 0, 0)

    def test_feature_matrix_zabbix_64(self):
        from zbbx_mcp.tools.health import _feature_matrix
        feats = dict(_feature_matrix(6, 4))
        assert feats["Unacknowledge action (action bit 16)"] is True
        assert feats["Cause / symptom rank actions (bits 128/256)"] is True
        assert feats["Connector API (data streaming)"] is False
        assert feats["HA cluster API (core.ha.get)"] is False

    def test_feature_matrix_zabbix_60_no_rank_actions(self):
        from zbbx_mcp.tools.health import _feature_matrix
        feats = dict(_feature_matrix(6, 0))
        assert feats["Cause / symptom rank actions (bits 128/256)"] is False
        assert feats["Unacknowledge action (action bit 16)"] is True

    def test_feature_matrix_zabbix_70_unlocks_everything(self):
        from zbbx_mcp.tools.health import _feature_matrix
        feats = dict(_feature_matrix(7, 0))
        assert feats["Connector API (data streaming)"] is True
        assert feats["Proxy groups (proxygroup.get)"] is True
        assert feats["HA cluster API (core.ha.get)"] is True

class TestExcelFills:
    """Regression for the lazy-init Fill bug fixed in v1.9.2.

    Before the fix, ``HEADER_FILL`` and friends were ``None`` at import
    time and only rebound inside ``_init_openpyxl()``. Consumers doing
    ``from zbbx_mcp.excel import HEADER_FILL`` captured the ``None``
    binding, which then fired ``TypeError: expected
    <class 'openpyxl.styles.fills.Fill'>`` during ``wb.save()`` —
    Sentry issue ``dc717f4d`` against ``generate_full_report``.
    """

    def test_fills_are_pattern_fill_instances(self):
        from openpyxl.styles import PatternFill

        from zbbx_mcp.excel import (
            DARK_RED_FILL,
            GREEN_FILL,
            HEADER_FILL,
            LIGHT_GREEN_FILL,
            ORANGE_FILL,
            RED_FILL,
        )
        for fill in (HEADER_FILL, RED_FILL, ORANGE_FILL, GREEN_FILL,
                     LIGHT_GREEN_FILL, DARK_RED_FILL):
            assert isinstance(fill, PatternFill), (
                f"{fill!r} should be a PatternFill, not {type(fill).__name__}"
            )

    def test_workbook_with_module_fills_saves(self):
        import io

        from openpyxl import Workbook

        from zbbx_mcp.excel import (
            DARK_RED_FILL,
            GREEN_FILL,
            HEADER_FILL,
            LIGHT_GREEN_FILL,
            ORANGE_FILL,
            RED_FILL,
        )
        wb = Workbook()
        ws = wb.active
        for i, fill in enumerate(
            (HEADER_FILL, RED_FILL, ORANGE_FILL, GREEN_FILL,
             LIGHT_GREEN_FILL, DARK_RED_FILL),
            start=1,
        ):
            c = ws.cell(row=i, column=1, value=str(i))
            c.fill = fill
        b = io.BytesIO()
        wb.save(b)
        assert len(b.getvalue()) > 0

    def test_full_report_module_level_imports_resolve_to_fills(self):
        # The specific failure mode: ``full_report.py`` does
        # ``from zbbx_mcp.excel import HEADER_FILL, RED_FILL, ...`` at
        # module level. After the fix those names must already point to
        # PatternFill instances at import time.
        from openpyxl.styles import PatternFill

        from zbbx_mcp.tools import full_report
        for name in ("RED_FILL", "GREEN_FILL", "ORANGE_FILL"):
            val = getattr(full_report, name)
            assert isinstance(val, PatternFill), (
                f"full_report.{name} is {type(val).__name__}; "
                f"lazy-init regression"
            )

class TestStaleBuildWarning:
    """Pure-helper tests for stale-build detection (ADR 073)."""

    def test_mismatch_warns(self):
        from zbbx_mcp.tools.health import stale_build_warning
        out = stale_build_warning("1.16.2", "1.16.5")
        assert "v1.16.2" in out and "v1.16.5" in out and "reconnect" in out

    def test_match_silent(self):
        from zbbx_mcp.tools.health import stale_build_warning
        assert stale_build_warning("1.16.5", "1.16.5") == ""

    def test_unknown_sides_silent(self):
        from zbbx_mcp.tools.health import stale_build_warning
        assert stale_build_warning("1.16.5", "") == ""          # wheel install
        assert stale_build_warning("", "1.16.5") == ""
        assert stale_build_warning("0.0.0+unknown", "1.16.5") == ""

    def test_source_tree_version_reads_checkout(self, tmp_path):
        from zbbx_mcp.tools.health import source_tree_version
        (tmp_path / "pyproject.toml").write_text(
            '[project]\nname = "x"\nversion = "9.9.9"\n')
        pkg = tmp_path / "src" / "zbbx_mcp"
        pkg.mkdir(parents=True)
        init = pkg / "__init__.py"
        init.write_text("")
        assert source_tree_version(str(init)) == "9.9.9"

    def test_source_tree_version_absent(self, tmp_path):
        from zbbx_mcp.tools.health import source_tree_version
        pkg = tmp_path / "src" / "zbbx_mcp"
        pkg.mkdir(parents=True)
        init = pkg / "__init__.py"
        init.write_text("")
        assert source_tree_version(str(init)) == ""  # no pyproject → wheel-like

    def test_check_connection_carries_warning_on_mismatch(self, monkeypatch):
        import zbbx_mcp
        from tests.wiretest import RecordingClient, run_tool
        from zbbx_mcp.tools import health as health_mod

        monkeypatch.setattr(health_mod, "source_tree_version", lambda _: "9.9.9")
        monkeypatch.setattr(zbbx_mcp, "__version__", "1.0.0")
        out = run_tool(health_mod, "check_connection",
                       RecordingClient({"apiinfo.version": "7.4.9"}))
        assert "Connected. Zabbix version: 7.4.9" in out
        assert "Running build v1.0.0" in out and "v9.9.9" in out

class TestTelemetryTokenFooter:
    """Pure-helper tests for the telemetry token estimate (ADR 073)."""

    def test_footer_math(self):
        from zbbx_mcp.tools.telemetry import _token_footer
        rows = [
            {"calls": 3, "response_chars_total": 1200},
            {"calls": 1, "response_chars_total": 800},
        ]
        out = _token_footer(rows)
        # 2000 chars ≈ 500 tokens across 4 calls → 125 tokens/call
        assert "2,000 chars" in out and "500 tokens" in out and "125 tokens/call" in out

    def test_footer_empty_when_no_sizes(self):
        from zbbx_mcp.tools.telemetry import _token_footer
        assert _token_footer([]) == ""
        assert _token_footer([{"calls": 2, "response_chars_total": 0}]) == ""

    def test_summarise_exposes_totals(self):
        from zbbx_mcp.tools.telemetry import _summarise_records
        rows = _summarise_records([
            {"tool": "a", "status": "ok", "duration_ms": 5, "response_size": 100},
            {"tool": "a", "status": "ok", "duration_ms": 5, "response_size": 300},
        ])
        assert rows[0]["response_chars_total"] == 400

    def test_tool_renders_footer(self, tmp_path):
        import json as _json

        from tests.wiretest import RecordingClient, run_tool
        from zbbx_mcp.tools import telemetry as telemetry_mod

        log = tmp_path / "analytics.log"
        log.write_text("\n".join(
            _json.dumps({"tool": "a", "status": "ok", "duration_ms": 5,
                         "response_size": 400})
            for _ in range(2)
        ))
        out = run_tool(telemetry_mod, "get_telemetry_summary",
                       RecordingClient(), log_path=str(log))
        assert "Σ responses: 800 chars ≈ 200 tokens" in out

class TestRecentChangesWireContract:
    """ADR 070 — get_recent_changes must not send selectHosts to problem.get.

    Same -32602 class as triage's ADR 068; found live when the tool errored
    during a feed-vs-Zabbix analysis. Drives the real tool function through a
    recording fake client and asserts the wire contract.
    """

    def _run(self):
        from tests.wiretest import RecordingClient, run_tool
        from zbbx_mcp.tools import availability as availability_mod

        client = RecordingClient({
            "problem.get": [{"eventid": "9", "name": "Service Down",
                             "severity": "5", "clock": "1000",
                             "acknowledged": "0", "suppressed": "0",
                             "objectid": "77"}],
            "trigger.get": [{"triggerid": "77",
                             "hosts": [{"host": "node-eu-a1"}]}],
        })  # event.get → no resolved events (default [])
        out = run_tool(availability_mod, "get_recent_changes", client)
        return client, out

    def test_problem_get_omits_selecthosts(self):
        client, _ = self._run()
        pget = next(p for m, p in client.calls if m == "problem.get")
        assert "selectHosts" not in pget
        assert "objectid" in pget["output"]

    def test_event_get_keeps_selecthosts(self):
        # event.get DOES support selectHosts — the resolved branch is untouched.
        client, _ = self._run()
        eget = next(p for m, p in client.calls if m == "event.get")
        assert eget.get("selectHosts") == ["host"]

    def test_host_rendered_via_trigger_map(self):
        client, out = self._run()
        tget = next(p for m, p in client.calls if m == "trigger.get")
        assert tget.get("selectHosts") == ["host"]
        assert "node-eu-a1" in out  # host name reached the rendered table

class TestSummarizeTokenExpiry:
    """Pure-helper tests for summarize_token_expiry (ADR 057)."""

    NOW = 1_000_000_000

    def test_expiring_token_flagged_sorted_soonest_first(self):
        from zbbx_mcp.tools.health import summarize_token_expiry
        tokens = [
            {"name": "b", "expires_at": str(self.NOW + 20 * 86400), "status": "0"},
            {"name": "a", "expires_at": str(self.NOW + 5 * 86400), "status": "0"},
        ]
        out = summarize_token_expiry(tokens, self.NOW)
        assert [n for n, _ in out] == ["a", "b"]
        assert out[0][1] == 5

    def test_never_expiring_and_disabled_skipped(self):
        from zbbx_mcp.tools.health import summarize_token_expiry
        tokens = [
            {"name": "never", "expires_at": "0", "status": "0"},
            {"name": "disabled", "expires_at": str(self.NOW + 86400), "status": "1"},
        ]
        assert summarize_token_expiry(tokens, self.NOW) == []

    def test_far_future_not_flagged(self):
        from zbbx_mcp.tools.health import summarize_token_expiry
        tokens = [{"name": "ok", "expires_at": str(self.NOW + 90 * 86400), "status": "0"}]
        assert summarize_token_expiry(tokens, self.NOW) == []

    def test_already_expired_negative_days(self):
        from zbbx_mcp.tools.health import summarize_token_expiry
        tokens = [{"name": "dead", "expires_at": str(self.NOW - 2 * 86400), "status": "0"}]
        out = summarize_token_expiry(tokens, self.NOW)
        assert out and out[0][1] < 0

class TestFormatProxyCompat:
    """Pure-helper tests for format_proxy_compat (ADR 056)."""

    def test_current_version_no_annotation(self):
        from zbbx_mcp.tools.proxies import format_proxy_compat
        assert format_proxy_compat("1", "7.4.9") == " v7.4.9"

    def test_outdated_flagged(self):
        from zbbx_mcp.tools.proxies import format_proxy_compat
        assert "OUTDATED" in format_proxy_compat("2", "7.0.0")

    def test_unsupported_flagged(self):
        from zbbx_mcp.tools.proxies import format_proxy_compat
        assert "UNSUPPORTED" in format_proxy_compat("3", "6.0.1")

    def test_unknown_version_and_undefined_compat_empty(self):
        from zbbx_mcp.tools.proxies import format_proxy_compat
        assert format_proxy_compat("0", "0") == ""
