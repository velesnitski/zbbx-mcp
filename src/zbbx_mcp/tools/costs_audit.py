"""Cost audit tools — analyse, reconcile, and export billing matches.

Extracted from the former monolithic costs.py.
"""

import asyncio
import csv as _csv
import json
import os
import re

import httpx

from zbbx_mcp.data import host_ip
from zbbx_mcp.resolver import InstanceResolver
from zbbx_mcp.tools.costs_common import (
    _IP_RE,
    _load_billing_csv,
    _provider_medians,
)
from zbbx_mcp.utils import confined_input_path, confined_output_path, safe_output_path


def register(mcp, resolver: InstanceResolver, skip: set[str] = frozenset()) -> None:

    if "export_cost_audit" not in skip:

        @mcp.tool()
        async def export_cost_audit(
            output_xlsx: str = "~/Downloads/cost_audit.xlsx",
            mode: str = "estimated",
            source_xlsx: str = "",
            instance: str = "",
        ) -> str:
            """Export hosts with {$COST_MONTH} source for accounting review.

            mode='estimated' (default) returns only hosts not backed by an
            exact billing match; mode='all' returns every costed host. See
            ADR 005 for column layout.
            """
            try:
                import openpyxl
                from openpyxl.styles import Alignment, Font, PatternFill
            except ImportError:
                return "openpyxl not installed"

            from zbbx_mcp.classify import classify_host as _classify_host
            from zbbx_mcp.classify import detect_provider
            from zbbx_mcp.data import extract_country

            try:
                client = resolver.resolve(instance)
                hosts, macros = await asyncio.gather(
                    client.call("host.get", {
                        "output": ["hostid", "host"],
                        "selectGroups": ["name"],
                        "selectInterfaces": ["ip"],
                        "filter": {"status": "0"},
                    }),
                    client.call("usermacro.get", {
                        "output": ["hostid", "value", "description"],
                        "filter": {"macro": "{$COST_MONTH}"},
                    }),
                )
            except (httpx.HTTPError, ValueError) as e:
                return f"Error: {e}"

            # Optional: parse source-of-truth XLSX and record sheet+row for each IP.
            source_ip_locations: dict[str, list[tuple[str, int]]] = {}
            source_loaded = False
            source_abs_path = ""
            if source_xlsx:
                try:
                    import openpyxl
                    source_abs_path = confined_input_path(source_xlsx)
                    wb_src = openpyxl.load_workbook(source_abs_path, data_only=True)
                    ip_re_local = _IP_RE
                    for sheet_name in wb_src.sheetnames:
                        ws_src = wb_src[sheet_name]
                        for r_idx, row in enumerate(ws_src.iter_rows(values_only=True), start=1):
                            for cell in row:
                                if cell is None:
                                    continue
                                for ip in ip_re_local.findall(str(cell)):
                                    source_ip_locations.setdefault(ip, []).append(
                                        (sheet_name, r_idx)
                                    )
                    source_loaded = True
                except (ImportError, OSError, ValueError) as e:
                    return f"Failed to load source_xlsx: {e}"

            host_by_id = {h["hostid"]: h for h in hosts}
            # Explicit "estimated" markers — anything else is assumed billing-backed
            # (legacy macros with generic descriptions pre-date the provenance tags
            # and were almost always written from billing matches).
            estimated_prefixes = (
                "src:bulk_pattern",
                "src:product_median",
                "src:provider_median",
                "estimated from",
                "twin cluster",
            )

            rows = []
            for m in macros:
                try:
                    cost = float(m.get("value") or 0)
                except (ValueError, TypeError):
                    continue
                if cost <= 0:
                    continue
                h = host_by_id.get(m["hostid"])
                if not h:
                    continue
                desc = (m.get("description") or "").strip()
                is_estimated = any(desc.startswith(p) for p in estimated_prefixes)
                # Heuristic: our only bulk import in this cycle used $210.82 with a
                # generic description — catch it explicitly.
                if not is_estimated and abs(cost - 210.82) < 0.01 and not desc.startswith(("src:billing_", "base ")):
                    is_estimated = True
                backed = not is_estimated
                if mode == "estimated" and backed:
                    continue
                ip = host_ip(h) or ""
                prov = detect_provider(ip) if ip else "?"
                prod, tier = _classify_host(h.get("groups", []))
                cc = extract_country(h["host"]) or ""
                src_locs = source_ip_locations.get(ip, []) if source_loaded else []
                rows.append({
                    "host": h["host"],
                    "host_id": h["hostid"],
                    "ip": ip,
                    "provider": prov,
                    "product": prod,
                    "tier": tier,
                    "country": cc,
                    "cost": cost,
                    "source": desc or "(no tag)",
                    "billing_backed": "yes" if backed else "no",
                    "in_source_of_truth": (
                        ("yes" if src_locs else "no") if source_loaded else ""
                    ),
                    "source_row": "; ".join(f"{s}!{r}" for s, r in src_locs[:3]),
                })

            rows.sort(key=lambda r: (-r["cost"], r["host"]))

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "cost_audit"
            hdr_font = Font(bold=True, color="FFFFFF")
            hdr_fill = PatternFill("solid", fgColor="366092")
            center = Alignment(horizontal="center")

            headers = ["Host", "Host ID", "IP", "Provider", "Product", "Tier",
                       "Country", "Cost $/mo", "Source", "Billing-backed"]
            if source_loaded:
                headers.extend(["In source of truth", "Source row"])
            for i, h in enumerate(headers, start=1):
                c = ws.cell(row=1, column=i, value=h)
                c.font = hdr_font
                c.fill = hdr_fill
                c.alignment = center

            green_fill = PatternFill("solid", fgColor="D9EAD3")
            yes_col_idx = 11  # "In source of truth" column index when present

            for i, r in enumerate(rows, start=2):
                vals = [r["host"], r["host_id"], r["ip"], r["provider"],
                        r["product"], r["tier"], r["country"], round(r["cost"], 2),
                        r["source"], r["billing_backed"]]
                if source_loaded:
                    vals.append(r["in_source_of_truth"])
                    vals.append(r["source_row"])
                for j, v in enumerate(vals, start=1):
                    cell = ws.cell(row=i, column=j, value=v)
                    # Hyperlink the source_row cell back into the workbook
                    if (source_loaded and j == yes_col_idx + 1
                            and r["source_row"] and source_abs_path):
                        first = r["source_row"].split(";")[0].strip()
                        if "!" in first:
                            sheet_part, row_part = first.split("!", 1)
                            cell.hyperlink = (
                                f"file://{source_abs_path}#'{sheet_part}'!A{row_part}"
                            )
                            cell.font = Font(color="0563C1", underline="single")
                # Highlight row green when present in source of truth
                if source_loaded and r["in_source_of_truth"] == "yes":
                    for j in range(1, len(vals) + 1):
                        ws.cell(row=i, column=j).fill = green_fill

            total_row = len(rows) + 3
            ws.cell(row=total_row, column=1, value=f"TOTAL ({len(rows)} hosts)").font = Font(bold=True)
            ws.cell(row=total_row, column=8, value=round(sum(r["cost"] for r in rows), 2)).font = Font(bold=True)

            widths = [30, 10, 16, 14, 18, 14, 10, 12, 38, 14]
            cols = "ABCDEFGHIJ"
            if source_loaded:
                widths.extend([18, 28])
                cols = "ABCDEFGHIJKL"
            for col, w in zip(cols, widths, strict=True):
                ws.column_dimensions[col].width = w

            try:
                out = confined_output_path(output_xlsx)
            except (OSError, ValueError) as e:
                return f"Failed to write XLSX: {e}"
            wb.save(out)
            total = sum(r["cost"] for r in rows)
            lines = [
                f"Exported {len(rows)} hosts (mode={mode}), ${total:,.2f}/mo",
                f"Wrote: {out}",
            ]
            if source_loaded:
                yes = sum(1 for r in rows if r["in_source_of_truth"] == "yes")
                lines.append(
                    f"In source of truth: {yes} yes (green) / {len(rows) - yes} no — review needed"
                )
            return "\n".join(lines)


    if "detect_cost_anomalies" not in skip:

        @mcp.tool()
        async def detect_cost_anomalies(
            high_factor: float = 2.5,
            low_factor: float = 0.3,
            max_results: int = 50,
            instance: str = "",
        ) -> str:
            """Flag hosts whose {$COST_MONTH} is far from its provider median.

            Useful after an import to catch mis-allocations (e.g. a bulk
            aggregate applied as a per-server rate). Compares each costed
            host's value to its provider median and reports the outliers.

            Args:
                high_factor: flag hosts at or above this multiple of median (default: 2.5)
                low_factor: flag hosts at or below this multiple of median (default: 0.3)
                max_results: cap on rows shown
                instance: Zabbix instance (optional)
            """
            from zbbx_mcp.classify import detect_provider

            try:
                client = resolver.resolve(instance)
                medians = await _provider_medians(client)
                hosts, macros = await asyncio.gather(
                    client.call("host.get", {
                        "output": ["hostid", "host"],
                        "selectInterfaces": ["ip"],
                        "filter": {"status": "0"},
                    }),
                    client.call("usermacro.get", {
                        "output": ["hostid", "value", "description"],
                        "filter": {"macro": "{$COST_MONTH}"},
                    }),
                )
                host_by_id = {h["hostid"]: h for h in hosts}
                anomalies = []
                for m in macros:
                    try:
                        v = float(m.get("value") or 0)
                    except (ValueError, TypeError):
                        continue
                    if v <= 0:
                        continue
                    h = host_by_id.get(m["hostid"])
                    if not h:
                        continue
                    ip = host_ip(h)
                    if not ip:
                        continue
                    prov = detect_provider(ip)
                    med = medians.get(prov)
                    if not med:
                        continue
                    ratio = v / med
                    if ratio >= high_factor or ratio <= low_factor:
                        anomalies.append({
                            "host": h["host"],
                            "provider": prov,
                            "cost": v,
                            "median": med,
                            "ratio": ratio,
                            "source": (m.get("description") or "").strip() or "(no tag)",
                        })
                anomalies.sort(key=lambda a: -abs(a["ratio"] - 1))
                if not anomalies:
                    return "No cost anomalies detected."
                lines = [
                    f"**{len(anomalies)} cost anomalies detected**",
                    f"Thresholds: >{high_factor}× or <{low_factor}× provider median\n",
                    "| Host | Provider | $/mo | Median | Ratio | Source |",
                    "|------|----------|------|--------|-------|--------|",
                ]
                for a in anomalies[:max_results]:
                    lines.append(
                        f"| {a['host']} | {a['provider']} | ${a['cost']:.2f} | "
                        f"${a['median']:.2f} | {a['ratio']:.1f}× | {a['source']} |"
                    )
                if len(anomalies) > max_results:
                    lines.append(f"*+{len(anomalies) - max_results} more*")
                return "\n".join(lines)
            except (httpx.HTTPError, ValueError) as e:
                return f"Error: {e}"


    if "analyze_cost_import" not in skip:

        @mcp.tool()
        async def analyze_cost_import(
            file_path: str,
            output_csv: str = "",
            output_json: str = "",
            instance: str = "",
        ) -> str:
            """Tiered probability analysis of unmatched cost entries.

            Outputs HIGH/MEDIUM/LOW/UNKNOWN tiers for each unmatched IP based
            on subnet, name-fuzzy, and provider-CIDR signals. See ADR 004.

            Args:
                file_path: Path to JSON cost file {ip: price} or {ip: {name, price, ...}}
                output_csv: Path for CSV output (default: ~/Downloads/cost_import_analysis.csv)
                output_json: Path for JSON output (default: ~/Downloads/cost_import_analysis.json)
                instance: Zabbix instance name (optional)
            """
            from collections import defaultdict

            from zbbx_mcp.classify import detect_provider

            try:
                path = confined_input_path(file_path)
                with open(path) as f:
                    raw = json.load(f)

                # Normalize input: accept both {ip: price} and {ip: {name, price, ...}}
                costs: dict[str, dict] = {}
                for ip, val in raw.items():
                    if isinstance(val, dict):
                        price = val.get("price") or val.get("cost") or 0
                        name = val.get("name", "")
                    else:
                        try:
                            price = float(val)
                            name = ""
                        except (ValueError, TypeError):
                            continue
                    if price > 0:
                        costs[ip.strip()] = {"name": str(name), "price": float(price)}

                if not costs:
                    return "No valid cost entries found in file."

                client = resolver.resolve(instance)
                hosts = await client.call("host.get", {
                    "output": ["hostid", "host", "name"],
                    "selectInterfaces": ["ip"],
                })

                zabbix_ips: dict[str, str] = {}
                zabbix_by_24: dict[str, list] = defaultdict(list)
                zabbix_names: list[str] = []
                for h in hosts:
                    zabbix_names.append(h["host"].lower())
                    for iface in h.get("interfaces", []):
                        ip = iface.get("ip", "")
                        if ip and ip != "127.0.0.1":
                            zabbix_ips[ip] = h["host"]
                            parts = ip.split(".")
                            if len(parts) == 4:
                                zabbix_by_24[f"{parts[0]}.{parts[1]}.{parts[2]}"].append((ip, h["host"]))

                def _tokens(s):
                    if not s:
                        return []
                    tokens = re.findall(r"[a-z]{2,}\d*|\d{2,}", s.lower())
                    stop = {"dedicated", "server", "custom", "dual", "standard", "gold", "price", "incl", "excl", "taxes"}
                    return [t for t in tokens if t not in stop and len(t) >= 2]

                results = []
                _PRIV = ("10.", "192.168.", "172.16.", "172.17.", "172.18.", "172.19.",
                         "172.20.", "172.21.", "172.22.", "172.23.", "172.24.", "172.25.",
                         "172.26.", "172.27.", "172.28.", "172.29.", "172.30.", "172.31.")

                for ip, info in costs.items():
                    if ip.startswith(_PRIV):
                        continue
                    if ip in zabbix_ips:
                        continue  # already matched

                    score = 0
                    signals = []

                    parts = ip.split(".")
                    if len(parts) == 4:
                        subnet = zabbix_by_24.get(f"{parts[0]}.{parts[1]}.{parts[2]}", [])
                        if subnet:
                            score += 40
                            signals.append(f"/24 match: {subnet[0][1]} (+{len(subnet) - 1} more)")

                    name_tokens = _tokens(info["name"])
                    matched_hosts = []
                    if name_tokens:
                        for hn in zabbix_names:
                            if all(t in hn for t in name_tokens[:2]):
                                matched_hosts.append(hn)
                                if len(matched_hosts) >= 3:
                                    break
                        if matched_hosts:
                            score += 35
                            signals.append(f"name match: {', '.join(matched_hosts[:2])}")
                        else:
                            for hn in zabbix_names:
                                for t in name_tokens:
                                    if len(t) >= 4 and t in hn:
                                        matched_hosts.append(hn)
                                        break
                                if len(matched_hosts) >= 2:
                                    break
                            if matched_hosts:
                                score += 15
                                signals.append(f"partial name: {matched_hosts[0]}")

                    prov = detect_provider(ip)
                    if prov not in ("Unknown", "Other"):
                        score += 15
                        signals.append(f"provider: {prov}")

                    if score >= 70:
                        tier = "HIGH"
                        sug = "Likely same server (different IP or hostname) — verify then add or update"
                    elif score >= 40:
                        tier = "MEDIUM"
                        sug = "Possible match — review subnet/name before importing"
                    elif score >= 15:
                        tier = "LOW"
                        sug = "Weak signal — new server not onboarded, or decommissioned"
                    else:
                        tier = "UNKNOWN"
                        sug = "No signal — external infra, abandoned, or needs manual investigation"

                    results.append({
                        "ip": ip, "name": info["name"], "price_monthly": info["price"],
                        "tier": tier, "confidence": score, "signals": signals, "suggestion": sug,
                    })

                tier_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2, "UNKNOWN": 3}
                results.sort(key=lambda x: (tier_order[x["tier"]], -x["price_monthly"]))

                by_tier: dict[str, dict] = defaultdict(lambda: {"count": 0, "cost": 0.0})
                for r in results:
                    by_tier[r["tier"]]["count"] += 1
                    by_tier[r["tier"]]["cost"] += r["price_monthly"]

                downloads = os.path.expanduser("~/Downloads")
                out_json = confined_output_path(output_json) if output_json else os.path.join(downloads, "cost_import_analysis.json")
                out_csv = confined_output_path(output_csv) if output_csv else os.path.join(downloads, "cost_import_analysis.csv")

                with open(out_json, "w") as f:
                    json.dump(results, f, indent=2)

                import csv as _csv
                with open(out_csv, "w", newline="") as f:
                    w = _csv.writer(f)
                    w.writerow(["Tier", "Confidence", "IP", "Billing Name", "Monthly $",
                                "Annual $", "Signals", "Suggestion"])
                    for r in results:
                        w.writerow([r["tier"], r["confidence"], r["ip"], r["name"],
                                    f"{r['price_monthly']:.2f}",
                                    f"{r['price_monthly'] * 12:.2f}",
                                    " | ".join(r["signals"]), r["suggestion"]])

                total_unmatched = sum(r["price_monthly"] for r in results)
                lines = [
                    f"**Cost Import Analysis: {len(results)} unmatched IPs = ${total_unmatched:,.2f}/mo**\n",
                    "| Tier | Count | Monthly $ | Annual $ |",
                    "|------|-------|-----------|----------|",
                ]
                for t in ["HIGH", "MEDIUM", "LOW", "UNKNOWN"]:
                    d = by_tier[t]
                    lines.append(f"| {t} | {d['count']} | ${d['cost']:,.2f} | ${d['cost'] * 12:,.2f} |")
                lines.append(f"\nSaved: `{out_json}`")
                lines.append(f"Saved: `{out_csv}`")

                return "\n".join(lines)
            except (httpx.HTTPError, ValueError, OSError) as e:
                return f"Error: {e}"


    if "reconcile_billing_audit" not in skip:

        @mcp.tool()
        async def reconcile_billing_audit(
            file_path: str,
            output_dir: str = "",
            instance: str = "",
        ) -> str:
            """Categorize billing entries into finance action buckets.

            Buckets: importable, already_costed, stale_ip, subnet_match, onboard, cancel.

            Args:
                file_path: CSV with ip, billing_name, price_monthly columns
                output_dir: Directory to write per-bucket CSVs (default: same as input)
                instance: Zabbix instance (optional)
            """
            try:
                rows = _load_billing_csv(file_path)
            except (FileNotFoundError, OSError, ValueError) as e:
                return f"Failed to read CSV: {e}"
            if not rows:
                return "No valid billing rows found."

            try:
                client = resolver.resolve(instance)
                hosts = await client.call("host.get", {
                    "output": ["hostid", "host"],
                    "selectInterfaces": ["ip"],
                    "filter": {"status": "0"},
                })
                macros = await client.call("usermacro.get", {
                    "output": ["hostid", "value"],
                    "filter": {"macro": "{$COST_MONTH}"},
                })
                has_cost = {m["hostid"] for m in macros if m.get("value") and m["value"] != "0"}
            except httpx.HTTPError as e:
                return f"Zabbix error: {e}"

            ip_to_host: dict[str, dict] = {}
            subnet_to_host: dict[str, list] = {}
            name_to_host: dict[str, dict] = {}
            for h in hosts:
                hip = host_ip(h)
                if hip:
                    ip_to_host[hip] = h
                    subnet_to_host.setdefault(".".join(hip.split(".")[:3]), []).append(h)
                name_to_host[h["host"].lower()] = h

            buckets: dict[str, list] = {
                "importable": [], "already_costed": [], "stale_ip": [],
                "subnet_match": [], "onboard": [], "cancel": [],
            }

            for r in rows:
                ip, name = r["ip"], r["name"]
                h = ip_to_host.get(ip)
                if h:
                    entry = {**r, "zabbix_host": h["host"]}
                    buckets["already_costed" if h["hostid"] in has_cost else "importable"].append(entry)
                    continue
                low = name.lower()
                matched = None
                if low:
                    for key, hh in name_to_host.items():
                        first = low.split()[0] if low else ""
                        if key == low or (first and key == first):
                            matched = hh
                            break
                if matched:
                    buckets["stale_ip"].append({
                        **r, "zabbix_host": matched["host"], "zabbix_ip": host_ip(matched),
                    })
                    continue
                subnet = ".".join(ip.split(".")[:3])
                if subnet in subnet_to_host:
                    sample = subnet_to_host[subnet][0]
                    buckets["subnet_match"].append({
                        **r, "zabbix_host": sample["host"], "zabbix_ip": host_ip(sample),
                    })
                    continue
                if name and any(c.isalpha() for c in name) and "." not in name[:5]:
                    buckets["onboard"].append(r)
                else:
                    buckets["cancel"].append(r)

            if not output_dir:
                output_dir = os.path.dirname(os.path.expanduser(file_path)) or os.getcwd()
            base = os.path.splitext(os.path.basename(file_path))[0]
            written = []
            try:
                for bucket, items in buckets.items():
                    if not items:
                        continue
                    out = safe_output_path(output_dir, f"{base}__{bucket}.csv")
                    cols = sorted({k for it in items for k in it})
                    with open(out, "w", newline="") as f:
                        w = _csv.DictWriter(f, fieldnames=cols)
                        w.writeheader()
                        w.writerows(items)
                    written.append(out)
            except (OSError, ValueError) as e:
                return f"Failed to write bucket CSVs: {e}"

            actions = {
                "importable": "Safe to import (IP match, no cost)",
                "already_costed": "Skip — cost already set",
                "stale_ip": "Billing team: update IP",
                "subnet_match": "Likely new member — review",
                "onboard": "Ops: add host to Zabbix",
                "cancel": "Finance: cancel or investigate",
            }
            lines = [f"**Reconciliation: {len(rows)} billing rows**\n",
                     "| Bucket | Count | $/mo | Action |",
                     "|--------|------:|-----:|--------|"]
            for b, items in buckets.items():
                if not items:
                    continue
                total = sum(i["price"] for i in items)
                lines.append(f"| {b} | {len(items)} | ${total:,.2f} | {actions[b]} |")
            lines.append(f"\nWrote {len(written)} bucket CSVs to `{output_dir}`")
            return "\n".join(lines)


    if "find_stale_billing_ips" not in skip:

        @mcp.tool()
        async def find_stale_billing_ips(
            file_path: str,
            instance: str = "",
        ) -> str:
            """Detect billing entries where name matches a Zabbix host but IP differs.

            Args:
                file_path: CSV with ip, billing_name, price_monthly columns
                instance: Zabbix instance (optional)
            """
            try:
                rows = _load_billing_csv(file_path)
            except (FileNotFoundError, OSError, ValueError) as e:
                return f"Failed to read CSV: {e}"
            if not rows:
                return "No valid billing rows."

            try:
                client = resolver.resolve(instance)
                hosts = await client.call("host.get", {
                    "output": ["hostid", "host"],
                    "selectInterfaces": ["ip"],
                    "filter": {"status": "0"},
                })
            except httpx.HTTPError as e:
                return f"Zabbix error: {e}"

            name_to_host: dict[str, dict] = {}
            ip_to_host: dict[str, dict] = {}
            for h in hosts:
                name_to_host[h["host"].lower()] = h
                hip = host_ip(h)
                if hip:
                    ip_to_host[hip] = h

            stale: list[dict] = []
            for r in rows:
                ip, name = r["ip"], r["name"].lower()
                if not name or ip in ip_to_host:
                    continue
                matched = name_to_host.get(name)
                if not matched and name:
                    first = name.split()[0]
                    matched = name_to_host.get(first)
                if matched:
                    zip_ = host_ip(matched)
                    if zip_ and zip_ != ip:
                        stale.append({
                            "billing_name": r["name"],
                            "billing_ip": ip,
                            "zabbix_host": matched["host"],
                            "zabbix_ip": zip_,
                            "price_monthly": r["price"],
                        })

            if not stale:
                return "No stale billing IPs detected."

            total = sum(s["price_monthly"] for s in stale)
            lines = [f"**{len(stale)} stale billing IPs** (${total:,.2f}/mo affected)\n",
                     "| Billing name | Billing IP | Zabbix IP | $/mo |",
                     "|-------------|-----------|-----------|-----:|"]
            for s in sorted(stale, key=lambda x: -x["price_monthly"])[:50]:
                lines.append(
                    f"| {s['billing_name']} | {s['billing_ip']} | {s['zabbix_ip']} | ${s['price_monthly']:.2f} |"
                )
            if len(stale) > 50:
                lines.append(f"\n*+{len(stale) - 50} more*")
            return "\n".join(lines)
