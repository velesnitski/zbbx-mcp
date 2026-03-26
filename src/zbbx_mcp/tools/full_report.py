"""Comprehensive server report: all dashboards, distinct servers, cost correlation."""

from __future__ import annotations

import os
from datetime import datetime
from statistics import median

import httpx
from openpyxl import Workbook

from zbbx_mcp.resolver import InstanceResolver
from zbbx_mcp.data import fetch_all_data, extract_country
from zbbx_mcp.excel import (
    BW_RED, BW_ORANGE, BW_GREEN,
    RED_FILL, BOLD_FONT, GREEN_FILL, ORANGE_FILL, LIGHT_GREEN_FILL,
    write_headers, write_data_rows, finalize_sheet, auto_width,
    bandwidth_fill, cpu_fill,
)


# Column layout for the main sheet
MAIN_HEADERS = [
    "#", "Host", "Name", "Country", "Dashboard", "Tab",
    "Product", "Tier", "Provider", "IP",
    "CPU %", "Load Avg5", "Mem Avail GB",
    "Traffic In Mbps", "Traffic Out Mbps", "Traffic Total Mbps",
    "BW Util %", "BW Tier", "Connections",
    "VPN Primary", "VPN Secondary", "VPN Tertiary", "Agent", "Templates",
    "Cost/Month ($)", "Cost/Year ($)",
    "On Dashboard", "All Tabs", "Groups",
]


def _apply_row_colors(ws, row_idx: int, row: dict) -> None:
    """Apply conditional coloring to a single data row."""
    # CPU
    fill = cpu_fill(row.get("CPU %"))
    if fill:
        ws.cell(row=row_idx, column=MAIN_HEADERS.index("CPU %") + 1).fill = fill

    # Traffic (applies to In, Total, BW Util, BW Tier)
    bw_f, bw_font = bandwidth_fill(row.get("Traffic In Mbps"))
    if bw_f:
        for col_name in ("Traffic In Mbps", "Traffic Total Mbps", "BW Util %", "BW Tier"):
            cell = ws.cell(row=row_idx, column=MAIN_HEADERS.index(col_name) + 1)
            cell.fill = bw_f
            if bw_font:
                cell.font = bw_font

    # VPN health columns
    for col_name in ("VPN Primary", "VPN Secondary", "VPN Tertiary"):
        val = row.get(col_name, "")
        if val == "DOWN":
            ws.cell(row=row_idx, column=MAIN_HEADERS.index(col_name) + 1).fill = RED_FILL
        elif val == "OK":
            ws.cell(row=row_idx, column=MAIN_HEADERS.index(col_name) + 1).fill = GREEN_FILL


def _write_dashboard_tabs_sheet(wb: Workbook, rows: list[dict], tab_data: dict) -> None:
    """Sheet 2: Dashboard tabs summary."""
    ws = wb.create_sheet("Dashboard Tabs")
    headers = ["Dashboard", "Tab", "Servers", "Median CPU %",
               "Median Traffic Mbps", f"Servers >= {BW_RED} Mbps",
               "Total Connections", "Total Cost/Month ($)"]
    write_headers(ws, headers)

    for idx, (key, tab_rows) in enumerate(sorted(tab_data.items()), 2):
        dash, tab = key.split("||", 1)
        cpu_vals = [r["CPU %"] for r in tab_rows if r["CPU %"] is not None]
        t_vals = [r["Traffic In Mbps"] for r in tab_rows if r["Traffic In Mbps"] is not None]
        high_bw = sum(1 for r in tab_rows if (r["Traffic In Mbps"] or 0) >= BW_RED)

        ws.cell(row=idx, column=1, value=dash)
        ws.cell(row=idx, column=2, value=tab)
        ws.cell(row=idx, column=3, value=len(tab_rows))
        ws.cell(row=idx, column=4, value=round(median(cpu_vals), 1) if cpu_vals else "")
        ws.cell(row=idx, column=5, value=round(median(t_vals), 1) if t_vals else "")
        cell_high = ws.cell(row=idx, column=6, value=high_bw)
        if high_bw > 0:
            cell_high.fill = RED_FILL
        ws.cell(row=idx, column=7, value=sum(r["Connections"] or 0 for r in tab_rows))
        ws.cell(row=idx, column=8, value=round(sum(r["Cost/Month ($)"] or 0 for r in tab_rows), 2) or "")

    ws.freeze_panes = "A2"
    auto_width(ws, headers)


def _write_provider_product_sheet(wb: Workbook, rows: list[dict]) -> None:
    """Sheet 3: Provider × Product matrix."""
    ws = wb.create_sheet("Provider × Product")

    matrix: dict[str, dict[str, dict]] = {}
    for r in rows:
        prov = r["Provider"] or "No IP"
        prod = f"{r['Product']}/{r['Tier']}"
        m = matrix.setdefault(prov, {}).setdefault(prod, {"count": 0, "cost": 0.0})
        m["count"] += 1
        m["cost"] += r["Cost/Month ($)"] or 0

    products = sorted(set(prod for pdata in matrix.values() for prod in pdata))
    headers = ["Provider"] + products + ["Total Servers", "Total Cost ($)"]
    write_headers(ws, headers)

    for idx, prov in enumerate(sorted(matrix, key=lambda p: -sum(
        d["count"] for d in matrix[p].values()
    )), 2):
        ws.cell(row=idx, column=1, value=prov)
        total_svrs = 0
        total_cost = 0.0
        for col, prod in enumerate(products, 2):
            d = matrix[prov].get(prod, {"count": 0, "cost": 0})
            if d["count"]:
                ws.cell(row=idx, column=col, value=d["count"])
            total_svrs += d["count"]
            total_cost += d.get("cost", 0)
        ws.cell(row=idx, column=len(products) + 2, value=total_svrs)
        ws.cell(row=idx, column=len(products) + 3, value=round(total_cost, 2) if total_cost else "")

    ws.freeze_panes = "A2"
    auto_width(ws, headers)


def _write_bandwidth_sheet(wb: Workbook, rows: list[dict]) -> None:
    """Sheet 4: Bandwidth analysis by tier."""
    ws = wb.create_sheet("Bandwidth Analysis")
    headers = ["Tier", "Range", "Servers", "% of Total", "Median CPU %", "Total Cost ($)"]
    write_headers(ws, headers)

    with_traffic = [r for r in rows if r["Traffic In Mbps"] is not None]
    no_traffic = [r for r in rows if r["Traffic In Mbps"] is None]

    tiers = [
        ("CRITICAL", f">= {BW_RED} Mbps", lambda v: v >= BW_RED, RED_FILL),
        ("HIGH", f"{BW_ORANGE}–{BW_RED} Mbps", lambda v: BW_ORANGE <= v < BW_RED, ORANGE_FILL),
        ("NORMAL", f"{BW_GREEN}–{BW_ORANGE} Mbps", lambda v: BW_GREEN <= v < BW_ORANGE, GREEN_FILL),
        ("LOW", f"< {BW_GREEN} Mbps", lambda v: v < BW_GREEN, LIGHT_GREEN_FILL),
        ("NO DATA", "No traffic data", None, None),
    ]

    for idx, (name, label, pred, fill) in enumerate(tiers, 2):
        tier_rows = no_traffic if pred is None else [r for r in with_traffic if pred(r["Traffic In Mbps"])]
        cpu_vals = [r["CPU %"] for r in tier_rows if r["CPU %"] is not None]
        cost = sum(r["Cost/Month ($)"] or 0 for r in tier_rows)
        pct = len(tier_rows) / len(rows) * 100 if rows else 0

        ws.cell(row=idx, column=1, value=name)
        ws.cell(row=idx, column=2, value=label)
        ws.cell(row=idx, column=3, value=len(tier_rows))
        ws.cell(row=idx, column=4, value=f"{pct:.1f}%")
        ws.cell(row=idx, column=5, value=round(median(cpu_vals), 1) if cpu_vals else "")
        ws.cell(row=idx, column=6, value=round(cost, 2) if cost else "")
        if fill:
            ws.cell(row=idx, column=1).fill = fill

    ws.freeze_panes = "A2"
    auto_width(ws, headers)


def _write_product_analytics_sheet(wb: Workbook, rows: list[dict]) -> None:
    """Sheet: Product analytics — aggregated metrics per product/tier."""
    ws = wb.create_sheet("Product Analytics")
    headers = [
        "Product", "Tier", "Servers", "On Dashboard",
        "Median CPU %", "Max CPU %",
        "Median Traffic Mbps", "Max Traffic Mbps", "Servers >= 650 Mbps",
        "VPN OK", "VPN DOWN",
        "Countries", "Providers",
        "Cost/Month ($)",
    ]
    write_headers(ws, headers)

    # Aggregate by product/tier
    groups: dict[str, list[dict]] = {}
    for r in rows:
        key = f"{r['Product']}||{r['Tier']}"
        groups.setdefault(key, []).append(r)

    for idx, (key, g_rows) in enumerate(sorted(groups.items()), 2):
        prod, tier = key.split("||", 1)
        cpu_vals = [r["CPU %"] for r in g_rows if r["CPU %"] is not None]
        t_vals = [r["Traffic In Mbps"] for r in g_rows if r["Traffic In Mbps"] is not None]
        on_dash = sum(1 for r in g_rows if r["On Dashboard"] == "Yes")
        high_bw = sum(1 for r in g_rows if (r["Traffic In Mbps"] or 0) >= BW_RED)
        vpn_ok = sum(1 for r in g_rows if r.get("VPN Primary") == "OK")
        vpn_down = sum(1 for r in g_rows if r.get("VPN Primary") == "DOWN")
        countries = sorted(set(r["Country"] for r in g_rows if r["Country"]))
        providers = sorted(set(r["Provider"] for r in g_rows if r["Provider"]))
        cost = sum(r["Cost/Month ($)"] or 0 for r in g_rows)

        ws.cell(row=idx, column=1, value=prod)
        ws.cell(row=idx, column=2, value=tier)
        ws.cell(row=idx, column=3, value=len(g_rows))
        ws.cell(row=idx, column=4, value=on_dash)
        ws.cell(row=idx, column=5, value=round(median(cpu_vals), 1) if cpu_vals else "")
        ws.cell(row=idx, column=6, value=round(max(cpu_vals), 1) if cpu_vals else "")
        ws.cell(row=idx, column=7, value=round(median(t_vals), 1) if t_vals else "")
        ws.cell(row=idx, column=8, value=round(max(t_vals), 1) if t_vals else "")
        cell_bw = ws.cell(row=idx, column=9, value=high_bw)
        if high_bw:
            cell_bw.fill = RED_FILL
        ws.cell(row=idx, column=10, value=vpn_ok if vpn_ok else "")
        cell_xd = ws.cell(row=idx, column=11, value=vpn_down if vpn_down else "")
        if vpn_down:
            cell_xd.fill = RED_FILL
        ws.cell(row=idx, column=12, value=", ".join(countries))
        ws.cell(row=idx, column=13, value=", ".join(providers))
        ws.cell(row=idx, column=14, value=round(cost, 2) if cost else "")

    # Totals
    total_row = len(groups) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = BOLD_FONT
    ws.cell(row=total_row, column=3, value=len(rows)).font = BOLD_FONT
    ws.cell(row=total_row, column=4, value=sum(1 for r in rows if r["On Dashboard"] == "Yes")).font = BOLD_FONT
    ws.cell(row=total_row, column=9, value=sum(1 for r in rows if (r["Traffic In Mbps"] or 0) >= BW_RED)).font = BOLD_FONT

    ws.freeze_panes = "A2"
    auto_width(ws, headers)


def _write_country_analytics_sheet(wb: Workbook, rows: list[dict]) -> None:
    """Sheet: Country analytics — servers and traffic per country."""
    ws = wb.create_sheet("Country Analytics")
    headers = [
        "Country", "Servers", "Products", "Providers",
        "Median CPU %", "Median Traffic Mbps", "Total Traffic Gbps",
        "Servers >= 650 Mbps", "VPN DOWN",
    ]
    write_headers(ws, headers)

    countries: dict[str, list[dict]] = {}
    for r in rows:
        c = r.get("Country") or "Unknown"
        countries.setdefault(c, []).append(r)

    sorted_countries = sorted(countries.items(), key=lambda x: -len(x[1]))
    for idx, (country, c_rows) in enumerate(sorted_countries, 2):
        cpu_vals = [r["CPU %"] for r in c_rows if r["CPU %"] is not None]
        t_vals = [r["Traffic In Mbps"] for r in c_rows if r["Traffic In Mbps"] is not None]
        total_gbps = sum(t_vals) / 1000 if t_vals else 0
        high_bw = sum(1 for r in c_rows if (r["Traffic In Mbps"] or 0) >= BW_RED)
        vpn_down = sum(1 for r in c_rows if r.get("VPN Primary") == "DOWN")
        products = sorted(set(r["Product"] for r in c_rows))
        providers = sorted(set(r["Provider"] for r in c_rows if r["Provider"]))

        ws.cell(row=idx, column=1, value=country)
        ws.cell(row=idx, column=2, value=len(c_rows))
        ws.cell(row=idx, column=3, value=", ".join(products))
        ws.cell(row=idx, column=4, value=", ".join(providers))
        ws.cell(row=idx, column=5, value=round(median(cpu_vals), 1) if cpu_vals else "")
        ws.cell(row=idx, column=6, value=round(median(t_vals), 1) if t_vals else "")
        ws.cell(row=idx, column=7, value=round(total_gbps, 1) if total_gbps else "")
        cell_bw = ws.cell(row=idx, column=8, value=high_bw if high_bw else "")
        if high_bw:
            cell_bw.fill = RED_FILL
        cell_xd = ws.cell(row=idx, column=9, value=vpn_down if vpn_down else "")
        if vpn_down:
            cell_xd.fill = RED_FILL

    # Totals
    total_row = len(sorted_countries) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = BOLD_FONT
    ws.cell(row=total_row, column=2, value=len(rows)).font = BOLD_FONT

    ws.freeze_panes = "A2"
    auto_width(ws, headers)


def _write_health_overview_sheet(wb: Workbook, rows: list[dict]) -> None:
    """Sheet: Health overview — problems summary."""
    ws = wb.create_sheet("Health Overview")
    headers = [
        "Metric", "Count", "% of Total", "Details",
    ]
    write_headers(ws, headers)

    total = len(rows)
    with_cpu = [r for r in rows if r["CPU %"] is not None]
    with_traffic = [r for r in rows if r["Traffic In Mbps"] is not None]

    metrics = [
        ("Total Servers", total, "", ""),
        ("On Dashboard", sum(1 for r in rows if r["On Dashboard"] == "Yes"), "", ""),
        ("Off Dashboard", sum(1 for r in rows if r["On Dashboard"] == "No"), "", "Review for dashboard coverage"),
        ("", "", "", ""),
        ("CPU >= 80%", sum(1 for r in with_cpu if r["CPU %"] >= 80), f"{sum(1 for r in with_cpu if r['CPU %'] >= 80)/total*100:.1f}%", "Overloaded"),
        ("CPU < 10%", sum(1 for r in with_cpu if r["CPU %"] < 10), f"{sum(1 for r in with_cpu if r['CPU %'] < 10)/total*100:.1f}%", "Potentially idle"),
        ("No CPU Data", total - len(with_cpu), f"{(total - len(with_cpu))/total*100:.1f}%", "Agent may be down"),
        ("", "", "", ""),
        ("Traffic >= 650 Mbps", sum(1 for r in with_traffic if r["Traffic In Mbps"] >= BW_RED), "", "Near saturation"),
        ("Traffic >= 500 Mbps", sum(1 for r in with_traffic if r["Traffic In Mbps"] >= BW_ORANGE), "", "High utilization"),
        ("No Traffic Data", total - len(with_traffic), f"{(total - len(with_traffic))/total*100:.1f}%", ""),
        ("", "", "", ""),
        ("VPN Primary OK", sum(1 for r in rows if r.get("VPN Primary") == "OK"), "", "VPN healthy"),
        ("VPN Primary DOWN", sum(1 for r in rows if r.get("VPN Primary") == "DOWN"), "", "VPN broken — investigate"),
        ("No VPN Primary Data", sum(1 for r in rows if not r.get("VPN Primary")), "", "Non-VPN server or no check"),
        ("", "", "", ""),
        ("Unique Countries", len(set(r["Country"] for r in rows if r["Country"])), "", ""),
        ("Unique Providers", len(set(r["Provider"] for r in rows if r["Provider"])), "", ""),
        ("Unique Products", len(set(r["Product"] for r in rows)), "", ""),
    ]

    for idx, (metric, count, pct, detail) in enumerate(metrics, 2):
        ws.cell(row=idx, column=1, value=metric)
        if metric:
            ws.cell(row=idx, column=1).font = BOLD_FONT if not detail else None
        ws.cell(row=idx, column=2, value=count if count != "" else "")
        ws.cell(row=idx, column=3, value=pct)
        ws.cell(row=idx, column=4, value=detail)

        # Color problem rows
        if "DOWN" in str(metric) or "Overloaded" in detail:
            ws.cell(row=idx, column=2).fill = RED_FILL
        elif "idle" in detail.lower():
            ws.cell(row=idx, column=2).fill = ORANGE_FILL
        elif "healthy" in detail.lower():
            ws.cell(row=idx, column=2).fill = GREEN_FILL

    auto_width(ws, headers)


def _write_off_dashboard_sheet(wb: Workbook, rows: list[dict]) -> None:
    """Sheet: Servers not on any dashboard — with VPN health and templates."""
    off = [r for r in rows if r["On Dashboard"] == "No"]
    if not off:
        return
    ws = wb.create_sheet(f"Off-Dashboard ({len(off)})")
    headers = ["#", "Host", "Country", "Product", "Tier", "Provider", "IP",
               "CPU %", "Traffic In Mbps", "VPN Primary", "VPN Secondary", "VPN Tertiary",
               "Agent", "Templates", "Groups"]
    write_headers(ws, headers)
    write_data_rows(ws, off, headers)

    # Color VPN status columns
    vpn1_col = headers.index("VPN Primary") + 1
    vpn2_col = headers.index("VPN Secondary") + 1
    vpn3_col = headers.index("VPN Tertiary") + 1
    for idx, r in enumerate(off, 2):
        for col, key in ((vpn1_col, "VPN Primary"), (vpn2_col, "VPN Secondary"), (vpn3_col, "VPN Tertiary")):
            val = r.get(key, "")
            if val == "DOWN":
                ws.cell(row=idx, column=col).fill = RED_FILL
            elif val == "OK":
                from zbbx_mcp.excel import GREEN_FILL
                ws.cell(row=idx, column=col).fill = GREEN_FILL

    finalize_sheet(ws, headers, len(off))


def register(mcp, resolver: InstanceResolver, skip: set[str] = frozenset()) -> None:

    if "generate_full_report" not in skip:

        @mcp.tool()
        async def generate_full_report(
            include_off_dashboard: bool = True,
            country: str = "",
            product: str = "",
            products: str = "",
            exclude_product: str = "",
            output_dir: str = "",
            instance: str = "",
        ) -> str:
            """Generate comprehensive report of ALL distinct servers across ALL dashboards.

            Sheets:
            1. All Servers — distinct, deduplicated, with all metrics
            2. Health Overview
            3. Product Analytics
            4. Country Analytics
            5. Per-dashboard tabs summary
            6. Provider × Product matrix with costs
            7. Bandwidth analysis (utilization tiers)
            8. Off-Dashboard servers

            Args:
                include_off_dashboard: Include servers not on any dashboard (default: True)
                country: Filter by country code in hostname (optional)
                product: Filter by single product name (optional)
                products: Comma-separated product names to include (optional)
                exclude_product: Comma-separated product names to exclude (optional)
                output_dir: Directory for the Excel file (default: ~/Downloads)
                instance: Zabbix instance name (optional)
            """
            try:
                client = resolver.resolve(instance)
                result = await fetch_all_data(client, include_off_dashboard)
                rows = result.rows

                # Apply filters
                include_set = {p.strip().lower() for p in products.split(",") if p.strip()} if products else set()
                exclude_set = {p.strip().lower() for p in exclude_product.split(",") if p.strip()} if exclude_product else set()

                if country or product or include_set or exclude_set:
                    filtered = []
                    for r in rows:
                        rp = r.get("Product", "").lower()
                        if country and extract_country(r.get("Host", "")).lower() != country.lower():
                            continue
                        if product and product.lower() not in rp:
                            continue
                        if include_set and not any(p in rp for p in include_set):
                            continue
                        if exclude_set and any(p in rp for p in exclude_set):
                            continue
                        filtered.append(r)
                    rows = filtered
                    # Rebuild tab_data for filtered rows
                    result.tab_data = {}
                    for r in rows:
                        if r["Dashboard"]:
                            key = f"{r['Dashboard']}||{r['Tab']}"
                            result.tab_data.setdefault(key, []).append(r)

                if not rows:
                    return "No servers found."

                wb = Workbook()

                # Sheet 1: All Servers
                ws1 = wb.active
                ws1.title = f"All Servers ({len(rows)})"
                write_headers(ws1, MAIN_HEADERS)
                write_data_rows(ws1, rows, MAIN_HEADERS)
                for idx, r in enumerate(rows, 2):
                    _apply_row_colors(ws1, idx, r)
                finalize_sheet(ws1, MAIN_HEADERS, len(rows))

                # Sheet 2–8
                _write_health_overview_sheet(wb, rows)
                _write_product_analytics_sheet(wb, rows)
                _write_country_analytics_sheet(wb, rows)
                _write_dashboard_tabs_sheet(wb, rows, result.tab_data)
                _write_provider_product_sheet(wb, rows)
                _write_bandwidth_sheet(wb, rows)
                _write_off_dashboard_sheet(wb, rows)

                # Save
                if not output_dir:
                    output_dir = os.path.expanduser("~/Downloads")
                ts = datetime.now().strftime("%Y%m%d_%H%M")
                filepath = os.path.join(output_dir, f"zabbix_full_report_{ts}.xlsx")
                wb.save(filepath)

                # Recalculate stats for filtered rows
                on = sum(1 for r in rows if r["On Dashboard"] == "Yes")
                off = sum(1 for r in rows if r["On Dashboard"] == "No")
                critical = sum(1 for r in rows if r["BW Tier"] == "CRITICAL")
                high = sum(1 for r in rows if r["BW Tier"] == "HIGH")
                cost = sum(r["Cost/Month ($)"] or 0 for r in rows)

                parts = [
                    f"**Full Server Report**",
                    f"",
                    f"**File:** `{filepath}`",
                    f"**Servers:** {len(rows)} distinct ({on} on dashboards, {off} off)",
                    f"**Bandwidth:** {critical} critical (>={BW_RED} Mbps), {high} high (>={BW_ORANGE} Mbps)",
                ]
                if cost:
                    parts.append(f"**Total cost:** ${cost:,.2f}/month")
                vpn_down = sum(1 for r in rows if r.get("VPN Primary") == "DOWN")
                countries = len(set(r["Country"] for r in rows if r["Country"]))
                products = len(set(r["Product"] for r in rows))

                if vpn_down:
                    parts.append(f"**VPN DOWN:** {vpn_down} servers")
                parts.extend([
                    f"**Countries:** {countries} | **Products:** {products}",
                    f"",
                    f"### Sheets",
                    f"1. **All Servers** — {len(rows)} × {len(MAIN_HEADERS)} columns",
                    f"2. **Health Overview** — infrastructure health summary",
                    f"3. **Product Analytics** — {products} products with metrics",
                    f"4. **Country Analytics** — {countries} countries",
                    f"5. **Dashboard Tabs** — {len(result.tab_data)} tabs",
                    f"6. **Provider × Product** — matrix with costs",
                    f"7. **Bandwidth Analysis** — utilization tiers",
                    f"8. **Off-Dashboard** — {off} unmonitored servers",
                ])

                return "\n".join(parts)
            except (httpx.HTTPError, ValueError) as e:
                return f"Error: {e}"
