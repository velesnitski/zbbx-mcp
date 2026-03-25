"""Export a single Zabbix dashboard to Excel with full server details."""

import asyncio
import os
from datetime import datetime
from statistics import median

import httpx

from zbbx_mcp.resolver import InstanceResolver
from zbbx_mcp.classify import classify_host as _classify_host, detect_provider


def register(mcp, resolver: InstanceResolver, skip: set[str] = frozenset()) -> None:

    if "export_dashboard" not in skip:

        @mcp.tool()
        async def export_dashboard(
            dashboard_id: str,
            output_dir: str = "",
            instance: str = "",
        ) -> str:
            """Export a Zabbix dashboard to Excel with full server details.

            For each server on the dashboard, includes:
            - Dashboard tab name, server hostname, visible name
            - Product, tier, hosting provider, IP
            - CPU usage, load average, memory, traffic, active connections
            - Zabbix groups

            Excel has one sheet per dashboard tab, plus a Summary sheet.

            Args:
                dashboard_id: Zabbix dashboard ID (e.g., '113')
                output_dir: Directory for the Excel file (default: ~/Downloads)
                instance: Zabbix instance name (optional, for multi-instance setups)
            """
            try:
                client = resolver.resolve(instance)

                # Phase 1: dashboard + host details (parallel)
                dash_task = client.call("dashboard.get", {
                    "dashboardids": [dashboard_id],
                    "output": ["dashboardid", "name"],
                    "selectPages": "extend",
                })
                dash_data = await dash_task

                if not dash_data:
                    return f"Dashboard '{dashboard_id}' not found."

                d = dash_data[0]
                dash_name = d["name"]

                # Extract graph IDs per tab
                tab_graph_ids: dict[str, list[str]] = {}
                all_graph_ids: set[str] = set()
                for pi, page in enumerate(d.get("pages", [])):
                    tab = page.get("name", "") or f"Page {pi + 1}"
                    gids = []
                    for w in page.get("widgets", []):
                        for f in w.get("fields", []):
                            if f.get("type") == "6":
                                gids.append(f["value"])
                                all_graph_ids.add(f["value"])
                    tab_graph_ids[tab] = gids

                if not all_graph_ids:
                    return f"Dashboard '{dash_name}' has no graph widgets."

                # Resolve graphs → hosts
                graphs = await client.call("graph.get", {
                    "graphids": list(all_graph_ids),
                    "output": ["graphid"],
                    "selectHosts": ["hostid"],
                })

                graph_to_hostid: dict[str, str] = {}
                all_hostids: set[str] = set()
                for g in graphs:
                    for h in g.get("hosts", []):
                        graph_to_hostid[g["graphid"]] = h["hostid"]
                        all_hostids.add(h["hostid"])

                if not all_hostids:
                    return f"No hosts resolved from dashboard graphs."

                # Phase 2: host details + all metrics (parallel)
                tasks = [
                    client.call("host.get", {
                        "hostids": list(all_hostids),
                        "output": ["hostid", "host", "name", "status"],
                        "selectGroups": ["name"],
                        "selectInterfaces": ["ip"],
                    }),
                    client.call("item.get", {
                        "hostids": list(all_hostids),
                        "output": ["hostid", "lastvalue"],
                        "filter": {"key_": "system.cpu.util[,idle]", "status": "0"},
                    }),
                    client.call("item.get", {
                        "hostids": list(all_hostids),
                        "output": ["hostid", "lastvalue"],
                        "filter": {"key_": "system.cpu.load[percpu,avg5]", "status": "0"},
                    }),
                    client.call("item.get", {
                        "hostids": list(all_hostids),
                        "output": ["hostid", "lastvalue"],
                        "filter": {"key_": "vm.memory.size[available]", "status": "0"},
                    }),
                    client.call("item.get", {
                        "hostids": list(all_hostids),
                        "output": ["hostid", "lastvalue"],
                        "filter": {"key_": "service_connections", "status": "0"},
                    }),
                    client.call("item.get", {
                        "hostids": list(all_hostids),
                        "output": ["hostid", "lastvalue"],
                        "search": {"name": "Incoming network traffic"},
                        "filter": {"status": "0"},
                    }),
                    client.call("usermacro.get", {
                        "hostids": list(all_hostids),
                        "output": ["hostid", "value"],
                        "filter": {"macro": "{$COST_MONTH}"},
                    }),
                ]

                results = await asyncio.gather(*tasks)
                host_list, cpu_items, load_items, mem_items, conn_items, traffic_items, cost_macros = results

                host_map = {h["hostid"]: h for h in host_list}

                # Build metrics maps
                cpu_map: dict[str, float] = {}
                for i in cpu_items:
                    try:
                        cpu_map[i["hostid"]] = round(100 - float(i["lastvalue"]), 1)
                    except (ValueError, TypeError):
                        pass

                load_map: dict[str, float] = {}
                for i in load_items:
                    try:
                        load_map[i["hostid"]] = round(float(i["lastvalue"]), 2)
                    except (ValueError, TypeError):
                        pass

                mem_map: dict[str, float] = {}
                for i in mem_items:
                    try:
                        mem_map[i["hostid"]] = round(float(i["lastvalue"]) / 1_073_741_824, 1)
                    except (ValueError, TypeError):
                        pass

                conn_map: dict[str, float] = {}
                for i in conn_items:
                    try:
                        conn_map[i["hostid"]] = float(i["lastvalue"])
                    except (ValueError, TypeError):
                        pass

                traffic_map: dict[str, float] = {}
                for i in traffic_items:
                    try:
                        val = float(i["lastvalue"])
                        if val > traffic_map.get(i["hostid"], 0):
                            traffic_map[i["hostid"]] = val
                    except (ValueError, TypeError):
                        pass

                cost_map: dict[str, float] = {}
                for m in cost_macros:
                    try:
                        cost_map[m["hostid"]] = float(m["value"])
                    except (ValueError, TypeError):
                        pass

                # Build rows per tab
                all_rows = []
                tab_rows: dict[str, list[dict]] = {}

                for tab, gids in tab_graph_ids.items():
                    seen = set()
                    rows = []
                    for gid in gids:
                        hid = graph_to_hostid.get(gid)
                        if not hid or hid in seen:
                            continue
                        seen.add(hid)
                        h = host_map.get(hid, {})
                        ip = next((i["ip"] for i in h.get("interfaces", []) if i.get("ip") != "127.0.0.1"), "")
                        prod, tier = _classify_host(h.get("groups", []))
                        provider = detect_provider(ip) if ip else ""
                        cost = cost_map.get(hid)
                        traffic = traffic_map.get(hid)
                        groups = ", ".join(g["name"] for g in h.get("groups", []))

                        row = {
                            "Tab": tab,
                            "Host": h.get("host", ""),
                            "Name": h.get("name", ""),
                            "Product": prod or "",
                            "Tier": tier or "",
                            "Provider": provider,
                            "IP": ip,
                            "CPU %": cpu_map.get(hid),
                            "Load Avg5": load_map.get(hid),
                            "Mem Avail GB": mem_map.get(hid),
                            "Traffic Mbps": round(traffic / 1e6, 1) if traffic else None,
                            "Connections": conn_map.get(hid),
                            "Cost/Month ($)": cost,
                            "Groups": groups,
                        }
                        rows.append(row)
                        all_rows.append(row)

                    tab_rows[tab] = rows

                # Generate Excel
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                wb = Workbook()
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                thin_border = Border(bottom=Side(style="thin", color="D9D9D9"))

                headers = ["#", "Host", "Name", "Product", "Tier", "Provider", "IP",
                           "CPU %", "Load Avg5", "Mem Avail GB", "Traffic Mbps",
                           "Connections", "Cost/Month ($)", "Groups"]

                def _write_sheet(ws, rows, sheet_title=None):
                    if sheet_title:
                        ws.title = sheet_title[:31]  # Excel limit
                    for col, h in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=h)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")

                    cpu_col = headers.index("CPU %") + 1
                    for idx, r in enumerate(rows, 2):
                        ws.cell(row=idx, column=1, value=idx - 1)
                        for col, key in enumerate(headers[1:], 2):
                            cell = ws.cell(row=idx, column=col, value=r.get(key, ""))
                            cell.border = thin_border
                        # Color CPU
                        cpu_cell = ws.cell(row=idx, column=cpu_col)
                        cpu_val = r.get("CPU %")
                        if cpu_val is not None:
                            if cpu_val >= 80:
                                cpu_cell.fill = red_fill
                            elif cpu_val >= 50:
                                cpu_cell.fill = yellow_fill
                            elif cpu_val < 10:
                                cpu_cell.fill = green_fill

                    # Auto-width
                    for col in range(1, len(headers) + 1):
                        max_len = len(str(ws.cell(1, col).value or ""))
                        for row in range(2, min(len(rows) + 2, 50)):
                            val = ws.cell(row, col).value
                            if val:
                                max_len = max(max_len, len(str(val)))
                        ws.column_dimensions[ws.cell(1, col).column_letter].width = min(max_len + 3, 40)

                    ws.auto_filter.ref = f"A1:{chr(64 + len(headers))}{len(rows) + 1}"
                    ws.freeze_panes = "A2"

                # Sheet 1: All servers
                ws_all = wb.active
                _write_sheet(ws_all, all_rows, f"All ({len(all_rows)})")

                # One sheet per tab
                for tab, rows in tab_rows.items():
                    if rows:
                        ws = wb.create_sheet()
                        _write_sheet(ws, rows, f"{tab} ({len(rows)})")

                # Summary sheet
                ws_sum = wb.create_sheet("Summary")
                sum_headers = ["Tab", "Servers", "Median CPU %", "Median Traffic Mbps",
                               "Total Connections", "Total Cost/Month ($)"]
                for col, h in enumerate(sum_headers, 1):
                    cell = ws_sum.cell(row=1, column=col, value=h)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

                for idx, (tab, rows) in enumerate(tab_rows.items(), 2):
                    cpu_vals = [r["CPU %"] for r in rows if r["CPU %"] is not None]
                    traffic_vals = [r["Traffic Mbps"] for r in rows if r["Traffic Mbps"] is not None]
                    total_conns = sum(r["Connections"] or 0 for r in rows)
                    total_cost = sum(r["Cost/Month ($)"] or 0 for r in rows)

                    ws_sum.cell(row=idx, column=1, value=tab)
                    ws_sum.cell(row=idx, column=2, value=len(rows))
                    ws_sum.cell(row=idx, column=3, value=round(median(cpu_vals), 1) if cpu_vals else "")
                    ws_sum.cell(row=idx, column=4, value=round(median(traffic_vals), 1) if traffic_vals else "")
                    ws_sum.cell(row=idx, column=5, value=total_conns)
                    ws_sum.cell(row=idx, column=6, value=round(total_cost, 2) if total_cost else "")

                # Totals
                total_row = len(tab_rows) + 2
                bold = Font(bold=True)
                ws_sum.cell(row=total_row, column=1, value="TOTAL").font = bold
                ws_sum.cell(row=total_row, column=2, value=len(all_rows)).font = bold

                for col in range(1, len(sum_headers) + 1):
                    ws_sum.column_dimensions[ws_sum.cell(1, col).column_letter].width = 22
                ws_sum.freeze_panes = "A2"

                # Save
                if not output_dir:
                    output_dir = os.path.expanduser("~/Downloads")
                safe_name = dash_name.replace("/", "-").replace("\\", "-")[:40]
                ts = datetime.now().strftime("%Y%m%d_%H%M")
                filename = f"{safe_name}_{ts}.xlsx"
                filepath = os.path.join(output_dir, filename)
                wb.save(filepath)

                # Text summary
                parts = [
                    f"**Dashboard Export: {dash_name}**",
                    f"",
                    f"**File:** `{filepath}`",
                    f"**Servers:** {len(all_rows)} across {len(tab_rows)} tabs",
                    f"",
                    f"### Tabs",
                ]
                for tab, rows in tab_rows.items():
                    cpu_vals = [r["CPU %"] for r in rows if r["CPU %"] is not None]
                    traffic_vals = [r["Traffic Mbps"] for r in rows if r["Traffic Mbps"] is not None]
                    med_cpu = f"{median(cpu_vals):.1f}%" if cpu_vals else "N/A"
                    med_traffic = f"{median(traffic_vals):.1f} Mbps" if traffic_vals else "N/A"
                    parts.append(f"- **{tab}**: {len(rows)} servers, median CPU {med_cpu}, median traffic {med_traffic}")

                parts.extend([
                    f"",
                    f"### Sheets",
                    f"1. **All** — {len(all_rows)} servers × {len(headers)} columns",
                ] + [
                    f"{i+2}. **{tab}** — {len(rows)} servers"
                    for i, (tab, rows) in enumerate(tab_rows.items()) if rows
                ] + [
                    f"{len(tab_rows)+2}. **Summary** — per-tab aggregates",
                ])

                return "\n".join(parts)
            except (httpx.HTTPError, ValueError) as e:
                return f"Error: {e}"
