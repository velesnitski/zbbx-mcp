"""Server cost management via {$COST_MONTH} host macros."""

import asyncio
import csv as _csv
import fnmatch
import json
import os
import re
import statistics

import httpx

from zbbx_mcp.data import host_ip
from zbbx_mcp.resolver import InstanceResolver
from zbbx_mcp.utils import resolve_group_ids

_IP_RE = re.compile(r"\b(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})\b")

# Cost source provenance tags (written into {$COST_MONTH} description).
COST_SRC_BILLING_IP = "src:billing_ip"
COST_SRC_BILLING_NAME = "src:billing_name"
COST_SRC_BILLING_TRANSLATED = "src:billing_translated"
COST_SRC_BILLING_COMPOUND = "src:billing_compound"
COST_SRC_CLUSTER_EXTRAS = "src:cluster_extras"
COST_SRC_BULK_PATTERN = "src:bulk_pattern"
COST_SRC_PRODUCT_MEDIAN = "src:product_median"
COST_SRC_PROVIDER_MEDIAN = "src:provider_median"


async def _provider_medians(client) -> dict[str, float]:
    """Compute median {$COST_MONTH} per detected provider across costed hosts."""
    from zbbx_mcp.classify import detect_provider

    hosts, macros = await asyncio.gather(
        client.call("host.get", {
            "output": ["hostid", "host"],
            "selectInterfaces": ["ip"],
            "filter": {"status": "0"},
        }),
        client.call("usermacro.get", {
            "output": ["hostid", "value"],
            "filter": {"macro": "{$COST_MONTH}"},
        }),
    )
    costs: dict[str, float] = {}
    for m in macros:
        try:
            v = float(m.get("value") or 0)
            if v > 0:
                costs[m["hostid"]] = v
        except (ValueError, TypeError):
            pass
    bucket: dict[str, list[float]] = {}
    for h in hosts:
        if h["hostid"] not in costs:
            continue
        ip = host_ip(h)
        if not ip:
            continue
        prov = detect_provider(ip)
        bucket.setdefault(prov, []).append(costs[h["hostid"]])
    return {k: statistics.median(v) for k, v in bucket.items() if v}


def _sanity_warnings(
    matches: list[tuple],
    ip_to_host: dict,
    medians: dict[str, float],
    high_factor: float = 2.0,
    low_factor: float = 0.3,
) -> list[str]:
    """Return human-readable warnings for costs far from provider median."""
    from zbbx_mcp.classify import detect_provider

    warnings = []
    # matches: (hostname, hid, ip, cost, source)
    for name, _hid, ip, cost, _src in matches:
        if not ip:
            continue
        prov = detect_provider(ip)
        med = medians.get(prov)
        if not med:
            continue
        if cost >= med * high_factor:
            warnings.append(
                f"{name}: ${cost:.2f} is {cost / med:.1f}× {prov} median ${med:.2f}"
            )
        elif cost <= med * low_factor:
            warnings.append(
                f"{name}: ${cost:.2f} is {cost / med:.1f}× {prov} median ${med:.2f} (low)"
            )
    return warnings


def _load_billing_csv(path: str) -> list[dict]:
    """Read billing CSV with ip + billing_name + price_monthly columns.

    Tolerates column aliases (ip/ipaddress, name/billing_name/hostname,
    price/price_monthly/cost). Returns list of {ip, name, price} dicts.
    Skips invalid/reserved IPs and zero/negative prices.
    """
    rows: list[dict] = []
    with open(os.path.expanduser(path)) as f:
        reader = _csv.DictReader(f)
        for raw in reader:
            # Normalize headers
            norm = {k.strip().lower(): (v or "").strip() for k, v in raw.items() if k}
            ip = norm.get("ip") or norm.get("ipaddress") or norm.get("ip_address") or ""
            name = norm.get("billing_name") or norm.get("name") or norm.get("hostname") or ""
            price_raw = (
                norm.get("price_monthly") or norm.get("price")
                or norm.get("cost") or norm.get("cost_month") or ""
            )
            try:
                price = float(price_raw)
            except (ValueError, TypeError):
                continue
            if price <= 0 or not ip:
                continue
            # Skip reserved / bogus IPs
            octets = ip.split(".")
            if len(octets) != 4:
                continue
            try:
                a, b = int(octets[0]), int(octets[1])
            except ValueError:
                continue
            if a in (0, 10, 127, 169, 172, 192, 224, 255) and (a != 172 or 16 <= b <= 31):
                # Allow most, skip only clearly-reserved
                if a in (0, 127, 224, 255):
                    continue
            rows.append({"ip": ip, "name": name, "price": price})
    return rows


def register(mcp, resolver: InstanceResolver, skip: set[str] = frozenset()) -> None:

    if "import_server_costs" not in skip:

        @mcp.tool()
        async def import_server_costs(
            costs_json: str,
            only_if_empty: bool = False,
            instance: str = "",
        ) -> str:
            """Bulk-set monthly costs on servers using {$COST_MONTH} host macros.

            Args:
                costs_json: JSON mapping hostname patterns to USD cost (e.g. {"srv-nl-*": 20})
                only_if_empty: Skip hosts that already have a non-zero {$COST_MONTH} set
                instance: Zabbix instance (optional)
            """
            try:
                cost_rules = json.loads(costs_json)
            except json.JSONDecodeError:
                return "Invalid JSON. Expected: {\"hostname-pattern\": cost, ...}"

            if not isinstance(cost_rules, dict):
                return "Expected a JSON object mapping hostname patterns to costs."

            try:
                client = resolver.resolve(instance)

                # Get all enabled hosts
                hosts = await client.call("host.get", {
                    "output": ["hostid", "host"],
                    "filter": {"status": "0"},
                })

                # Get existing {$COST_MONTH} macros
                existing = await client.call("usermacro.get", {
                    "hostids": [h["hostid"] for h in hosts],
                    "output": ["hostmacroid", "hostid", "value"],
                    "filter": {"macro": "{$COST_MONTH}"},
                })
                existing_map = {m["hostid"]: m for m in existing}

                # Match hosts to cost rules
                matched = 0
                created = 0
                updated = 0
                skipped = 0
                errors = []

                for h in hosts:
                    hostname = h["host"]
                    hid = h["hostid"]
                    cost = None

                    # Check rules in order (exact match first, then patterns)
                    for pattern, value in cost_rules.items():
                        if hostname == pattern or fnmatch.fnmatch(hostname, pattern):
                            cost = value
                            break

                    if cost is None:
                        continue

                    matched += 1
                    cost_str = str(cost)

                    try:
                        if hid in existing_map:
                            existing_val = (existing_map[hid].get("value") or "").strip()
                            if only_if_empty and existing_val not in ("", "0", "0.0", "0.00"):
                                skipped += 1
                                continue
                            # Update if value changed
                            if existing_map[hid]["value"] != cost_str:
                                await client.call("usermacro.update", {
                                    "hostmacroid": existing_map[hid]["hostmacroid"],
                                    "value": cost_str,
                                })
                                updated += 1
                            else:
                                skipped += 1
                        else:
                            # Create new macro
                            await client.call("usermacro.create", {
                                "hostid": hid,
                                "macro": "{$COST_MONTH}",
                                "value": cost_str,
                                "description": COST_SRC_BULK_PATTERN,
                            })
                            created += 1
                    except (httpx.HTTPError, ValueError) as e:
                        errors.append(f"{hostname}: {e}")
                        if len(errors) >= 10:
                            break

                parts = [
                    "**Cost import complete**",
                    "",
                    f"Matched: {matched} servers",
                    f"Created: {created} new macros",
                    f"Updated: {updated} existing macros",
                    f"Unchanged: {skipped}",
                    f"Unmatched: {len(hosts) - matched} servers (no matching pattern)",
                ]
                if errors:
                    parts.append(f"\nErrors ({len(errors)}):")
                    for e in errors:
                        parts.append(f"- {e}")

                return "\n".join(parts)
            except (httpx.HTTPError, ValueError) as e:
                return f"Error importing costs: {e}"

    if "set_bulk_cost" not in skip:

        @mcp.tool()
        async def set_bulk_cost(
            group: str,
            cost: float,
            instance: str = "",
        ) -> str:
            """Set monthly cost for all servers in a host group.

            Args:
                group: Zabbix host group name
                cost: Monthly cost in USD per server
                instance: Zabbix instance name (optional, for multi-instance setups)
            """
            try:
                client = resolver.resolve(instance)

                # Resolve group
                gids = await resolve_group_ids(client, group)
                if gids is None:
                    return f"Host group '{group}' not found."

                # Get hosts in group
                hosts = await client.call("host.get", {
                    "output": ["hostid", "host"],
                    "groupids": gids,
                    "filter": {"status": "0"},
                })
                if not hosts:
                    return f"No enabled hosts in group '{group}'."

                # Get existing macros
                hostids = [h["hostid"] for h in hosts]
                existing = await client.call("usermacro.get", {
                    "hostids": hostids,
                    "output": ["hostmacroid", "hostid", "value"],
                    "filter": {"macro": "{$COST_MONTH}"},
                })
                existing_map = {m["hostid"]: m for m in existing}

                cost_str = str(cost)
                created = 0
                updated = 0
                skipped = 0

                for h in hosts:
                    hid = h["hostid"]
                    if hid in existing_map:
                        if existing_map[hid]["value"] != cost_str:
                            await client.call("usermacro.update", {
                                "hostmacroid": existing_map[hid]["hostmacroid"],
                                "value": cost_str,
                            })
                            updated += 1
                        else:
                            skipped += 1
                    else:
                        await client.call("usermacro.create", {
                            "hostid": hid,
                            "macro": "{$COST_MONTH}",
                            "value": cost_str,
                            "description": "Monthly server cost in USD",
                        })
                        created += 1

                return (
                    f"Set ${cost}/month on {len(hosts)} servers in '{group}'. "
                    f"Created: {created}, Updated: {updated}, Unchanged: {skipped}."
                )
            except (httpx.HTTPError, ValueError) as e:
                return f"Error: {e}"

    if "import_costs_by_ip" not in skip:

        @mcp.tool()
        async def import_costs_by_ip(
            file_path: str = "",
            costs_json: str = "",
            dry_run: bool = True,
            min_cost: float = 1,
            max_cost: float = 5000,
            export_unmatched: str = "",
            only_if_empty: bool = False,
            instance: str = "",
        ) -> str:
            """Smart cost import: match by IP, then hostname (7-pass fuzzy).

            Name matching: exact → split-on-dash → IP-in-name → prefix → contains → billing name translation.

            Args:
                file_path: Path to JSON cost file (preferred over costs_json)
                costs_json: Inline JSON (alternative to file_path)
                dry_run: Preview only, don't write (default: True)
                min_cost: Skip prices below this (default: $1)
                max_cost: Skip prices above this (default: $5000)
                export_unmatched: File path to save unmatched entries as JSON (optional)
                only_if_empty: Skip hosts that already have a non-zero {$COST_MONTH} set
                instance: Zabbix instance (optional)
            """
            # Load data
            try:
                if file_path:
                    with open(os.path.expanduser(file_path)) as f:
                        raw = json.load(f)
                elif costs_json:
                    raw = json.loads(costs_json)
                else:
                    return "Provide file_path or costs_json."
            except (json.JSONDecodeError, FileNotFoundError, OSError) as e:
                return f"Failed to load: {e}"

            if isinstance(raw, dict) and ("by_ip" in raw or "by_name" in raw):
                ip_costs = raw.get("by_ip", {})
                name_costs = raw.get("by_name", {})
            elif isinstance(raw, dict):
                ip_costs = raw
                name_costs = {}
            else:
                return "Expected JSON object with by_ip/by_name or flat IP→cost map."

            def _extract_price(v):
                """Extract numeric price from value (may be dict with price/cost field or number)."""
                if isinstance(v, dict):
                    for key in ("price", "cost", "price_monthly", "monthly"):
                        if key in v:
                            return v[key]
                    return None
                return v

            def _in_range(v):
                p = _extract_price(v)
                if p is None:
                    return False
                try:
                    return min_cost <= float(p) <= max_cost
                except (ValueError, TypeError):
                    return False

            safe_ips = {k: float(_extract_price(v)) for k, v in ip_costs.items() if _in_range(v)}
            safe_names = {k: float(_extract_price(v)) for k, v in name_costs.items() if _in_range(v)}

            # If IP values are dicts with 'name' field, also populate safe_names for fuzzy matching
            for _ip, v in ip_costs.items():
                if isinstance(v, dict) and _in_range(v):
                    name = (v.get("name") or "").strip()
                    if name and name not in safe_names:
                        safe_names[name] = float(_extract_price(v))
            total_input = len(ip_costs) + len(name_costs)
            skipped_range = total_input - len(safe_ips) - len(safe_names)

            try:
                client = resolver.resolve(instance)
                hosts = await client.call("host.get", {
                    "output": ["hostid", "host"],
                    "selectInterfaces": ["ip"],
                    "filter": {"status": "0"},
                })

                # Build lookup maps
                ip_to_host: dict[str, dict] = {}
                name_to_host: dict[str, dict] = {}
                name_list: list[str] = []
                # Compound hostnames: "base-cc1 cc3 cc5" monitors multiple
                # physical servers as one entity. Expand to component names so
                # each billing entry can be summed.
                compound_components: dict[str, list[str]] = {}
                for h in hosts:
                    ip = host_ip(h)
                    if ip:
                        ip_to_host[ip] = h
                    lname = h["host"].lower()
                    name_to_host[lname] = h
                    name_list.append(lname)
                    if " " in lname:
                        parts = lname.split()
                        m = re.match(r"^(.+?)([a-z]{2}\d+)$", parts[0])
                        if m:
                            base = m.group(1)
                            derivs = [parts[0]]
                            for p in parts[1:]:
                                if re.match(r"^[a-z]{2}\d+$", p):
                                    derivs.append(base + p)
                            if len(derivs) > 1:
                                compound_components[h["hostid"]] = derivs

                # --- Pass 1: by_ip (IP is source of truth) ---
                host_costs: dict[str, float] = {}  # hostid → cost
                host_source: dict[str, str] = {}   # hostid → match method
                compound_consumed: set[str] = set()  # billing names consumed by compound sum

                for ip, cost in safe_ips.items():
                    ip = ip.strip()
                    h = ip_to_host.get(ip)
                    if h:
                        hid = h["hostid"]
                        host_costs[hid] = max(host_costs.get(hid, 0), cost)
                        host_source[hid] = "ip"

                # --- Pass 1b: partial IP /24 prefix match (3-octet IPs) ---
                # Matches "x.y.z" against hosts with IP "x.y.z.*"
                # Runtime only — no IPs stored in code
                for ip_key, cost in safe_ips.items():
                    ip_key = ip_key.strip()
                    parts = ip_key.split(".")
                    if len(parts) == 3:
                        prefix = ip_key + "."
                        for full_ip, h in ip_to_host.items():
                            if full_ip.startswith(prefix):
                                hid = h["hostid"]
                                if hid not in host_costs or cost > host_costs[hid]:
                                    host_costs[hid] = cost
                                    host_source[hid] = "ip/24"

                # --- Pass 1c: compound hostnames (twin clusters) ---
                # For hosts like "prem-cc1 cc3", sum billing entries for each
                # component name rather than taking max of one.
                for hid, comps in compound_components.items():
                    total = 0.0
                    hits = []
                    for comp in comps:
                        if comp in safe_names:
                            total += safe_names[comp]
                            hits.append(comp)
                    if hits:
                        host_costs[hid] = max(host_costs.get(hid, 0), total)
                        host_source[hid] = f"compound({len(hits)})"
                        compound_consumed.update(hits)

                # --- Pass 2-5: by_name (fill gaps, use max if IP already matched) ---
                unmatched_names: dict[str, float] = {}

                for name, cost in safe_names.items():
                    if name.lower().strip() in compound_consumed:
                        continue
                    name_lower = name.lower().strip()
                    matched_host = None

                    # Pass 2: exact lowercase match
                    if name_lower in name_to_host:
                        matched_host = name_to_host[name_lower]

                    # Pass 3: split on " - ", try last segment
                    if not matched_host and " - " in name:
                        segments = name.split(" - ")
                        for seg in reversed(segments):
                            seg_clean = seg.strip().lower()
                            if seg_clean in name_to_host:
                                matched_host = name_to_host[seg_clean]
                                break

                    # Pass 4: extract IP from name, match via interface
                    if not matched_host:
                        ip_match = _IP_RE.search(name)
                        if ip_match:
                            found_ip = ip_match.group(1)
                            if found_ip in ip_to_host:
                                matched_host = ip_to_host[found_ip]

                    # Pass 5: prefix match (name is start of a Zabbix hostname)
                    if not matched_host and len(name_lower) >= 4:
                        for zname in name_list:
                            if zname.startswith(name_lower) or name_lower.startswith(zname):
                                matched_host = name_to_host[zname]
                                break

                    # Pass 6: billing name translation
                    # Billing often uses reversed naming: cc+num+suffix → suffix-cc+num
                    if not matched_host:
                        candidates = []
                        # Reverse: ccNNN-suffix -> suffix-cc0NNN
                        m = re.match(r"^([a-z]{2})(\d+)-(.+)$", name_lower)
                        if m:
                            cc, num, suffix = m.group(1), m.group(2), m.group(3)
                            candidates.extend([
                                f"{suffix}-{cc}{num.zfill(4)}",
                                f"{suffix}-{cc}{num}",
                            ])
                        # Prefix swap: pfx-ccNNN-suffix -> pfx-suffix-cc0NNN
                        m = re.match(r"^([a-z]{2,4})-([a-z]{2})(\d+)-(.+)$", name_lower)
                        if m:
                            pfx, cc, num, suffix = m.group(1), m.group(2), m.group(3), m.group(4)
                            candidates.append(f"{pfx}-{suffix}-{cc}{num.zfill(4)}")
                        # Configurable renames via env var
                        # ZABBIX_BILLING_RENAMES="old1:new1,old2:new2"
                        renames_raw = os.environ.get("ZABBIX_BILLING_RENAMES", "")
                        if renames_raw:
                            for pair in renames_raw.split(","):
                                if ":" in pair:
                                    old, new = pair.strip().split(":", 1)
                                    if old in name_lower:
                                        candidates.append(name_lower.replace(old, new))

                        for cand in candidates:
                            if cand in name_to_host:
                                matched_host = name_to_host[cand]
                                host_source[matched_host["hostid"]] = "translated"
                                break

                    if matched_host:
                        hid = matched_host["hostid"]
                        # max(ip_cost, name_cost) when both exist
                        host_costs[hid] = max(host_costs.get(hid, 0), cost)
                        if hid not in host_source:
                            host_source[hid] = "name"
                    else:
                        unmatched_names[name] = cost

                # Filter out hosts with an existing non-zero cost if only_if_empty
                skipped_existing = 0
                if only_if_empty and host_costs:
                    existing_pre = await client.call("usermacro.get", {
                        "hostids": list(host_costs.keys()),
                        "output": ["hostid", "value"],
                        "filter": {"macro": "{$COST_MONTH}"},
                    })
                    nonempty = {
                        m["hostid"] for m in existing_pre
                        if (m.get("value") or "").strip() not in ("", "0", "0.0", "0.00")
                    }
                    skipped_existing = len(nonempty)
                    for hid in nonempty:
                        host_costs.pop(hid, None)
                        host_source.pop(hid, None)

                # Build match list
                host_map = {h["hostid"]: h for h in hosts}
                matches = []
                for hid, cost in host_costs.items():
                    h = host_map.get(hid)
                    if h:
                        ip = host_ip(h)
                        matches.append((h["host"], hid, ip or "", cost, host_source.get(hid, "?")))

                ip_matches = sum(1 for s in host_source.values() if s == "ip")
                ip24_matches = sum(1 for s in host_source.values() if s == "ip/24")
                name_matches = sum(1 for s in host_source.values() if s == "name")
                translated_matches = sum(1 for s in host_source.values() if s == "translated")
                compound_matches = sum(1 for s in host_source.values() if s.startswith("compound"))

                # Export unmatched
                if export_unmatched and unmatched_names:
                    path = os.path.expanduser(export_unmatched)
                    with open(path, "w") as f:
                        json.dump(unmatched_names, f, indent=2, ensure_ascii=False)

                if dry_run:
                    total_cost = sum(c for _, _, _, c, _ in matches)
                    medians = await _provider_medians(client)
                    warnings = _sanity_warnings(matches, ip_to_host, medians)
                    lines = [
                        f"**DRY RUN — {len(matches)} hosts matched**",
                        f"By IP: {ip_matches} | By /24: {ip24_matches} | By name: {name_matches} | Translated: {translated_matches} | Compound: {compound_matches}",
                        f"Unmatched: {len(unmatched_names)} name entries",
                        f"Skipped: {skipped_range} outside ${min_cost}-${max_cost}",
                    ]
                    if only_if_empty:
                        lines.append(f"Skipped (already costed): {skipped_existing}")
                    lines.append(f"**Total: ${total_cost:,.2f}/mo** (${total_cost*12:,.2f}/yr)\n")
                    lines.append("| Host | Match | $/mo |")
                    lines.append("|------|-------|------|")
                    for hostname, _hid, _ip, cost, src in sorted(matches, key=lambda x: -x[3])[:25]:
                        lines.append(f"| {hostname} | {src} | ${cost:.2f} |")
                    if len(matches) > 25:
                        lines.append(f"| ... | +{len(matches) - 25} more | |")
                    if warnings:
                        lines.append(f"\n**⚠ Sanity check: {len(warnings)} prices far from provider median**")
                        for w in warnings[:10]:
                            lines.append(f"- {w}")
                        if len(warnings) > 10:
                            lines.append(f"- ...+{len(warnings) - 10} more")
                    if export_unmatched and unmatched_names:
                        lines.append(f"\nUnmatched exported to `{export_unmatched}`")
                    lines.append("\nSet `dry_run=false` to apply.")
                    return "\n".join(lines)

                # Apply: set {$COST_MONTH} macros
                existing = await client.call("usermacro.get", {
                    "hostids": [m[1] for m in matches],
                    "output": ["hostmacroid", "hostid", "value"],
                    "filter": {"macro": "{$COST_MONTH}"},
                })
                existing_map = {m["hostid"]: m for m in existing}

                def _src_tag(src: str) -> str:
                    if src == "ip" or src == "ip/24":
                        return COST_SRC_BILLING_IP
                    if src == "translated":
                        return COST_SRC_BILLING_TRANSLATED
                    if src.startswith("compound"):
                        return COST_SRC_BILLING_COMPOUND
                    return COST_SRC_BILLING_NAME

                created = updated = unchanged = 0
                errors = []
                for hostname, hid, _ip, cost, src in matches:
                    cost_str = str(round(cost, 2))
                    desc = _src_tag(src)
                    try:
                        if hid in existing_map:
                            if existing_map[hid]["value"] != cost_str:
                                await client.call("usermacro.update", {
                                    "hostmacroid": existing_map[hid]["hostmacroid"],
                                    "value": cost_str,
                                    "description": desc,
                                })
                                updated += 1
                            else:
                                unchanged += 1
                        else:
                            await client.call("usermacro.create", {
                                "hostid": hid,
                                "macro": "{$COST_MONTH}",
                                "value": cost_str,
                                "description": desc,
                            })
                            created += 1
                    except (httpx.HTTPError, ValueError) as e:
                        errors.append(f"{hostname}: {e}")
                        if len(errors) >= 10:
                            break

                total_cost = sum(c for _, _, _, c, _ in matches)
                parts = [
                    f"**Cost import complete — {len(matches)} servers**",
                    f"By IP: {ip_matches} | By name: {name_matches}",
                    f"Created: {created} | Updated: {updated} | Unchanged: {unchanged}",
                    f"Total: ${total_cost:,.2f}/mo (${total_cost*12:,.2f}/yr)",
                ]
                if unmatched_names:
                    parts.append(f"Unmatched: {len(unmatched_names)} name entries")
                if export_unmatched and unmatched_names:
                    parts.append(f"Exported to `{export_unmatched}`")
                if errors:
                    parts.append(f"Errors: {len(errors)}")
                return "\n".join(parts)
            except (httpx.HTTPError, ValueError) as e:
                return f"Error: {e}"

    if "import_cluster_ip_fees" not in skip:

        @mcp.tool()
        async def import_cluster_ip_fees(
            file_path: str = "",
            fees_json: str = "",
            dry_run: bool = True,
            instance: str = "",
        ) -> str:
            """Add extra-IP billing fees to cluster primary {$COST_MONTH} macros.

            For each entry, finds the Zabbix host by name, reads the existing
            {$COST_MONTH}, adds the extra_cost_month, and writes back with an
            audit description ("base X + N extra IPs (Y)").

            Input format: list of {"cluster": str, "extra_ips": [str], "extra_cost_month": float}

            Args:
                file_path: Path to JSON file with cluster fee entries (optional)
                fees_json: Inline JSON (optional; use this OR file_path)
                dry_run: Preview changes without applying (default: True)
                instance: Zabbix instance (optional)
            """
            try:
                client = resolver.resolve(instance)

                # Parse input
                if file_path:
                    with open(os.path.expanduser(file_path)) as f:
                        data = json.load(f)
                elif fees_json:
                    data = json.loads(fees_json)
                else:
                    return "Provide either file_path or fees_json."

                if not isinstance(data, list):
                    return "Expected JSON array of {cluster, extra_ips, extra_cost_month}."

                # Batch lookup: all cluster names in one host.get call
                host_names = [e.get("cluster", "") for e in data if e.get("cluster")]
                if not host_names:
                    return "No cluster names in input."

                hosts = await client.call("host.get", {
                    "output": ["hostid", "host"],
                    "filter": {"host": host_names},
                    "selectMacros": ["hostmacroid", "macro", "value"],
                })
                host_by_name = {h["host"]: h for h in hosts}

                # Build update plan
                updates = []
                missing = []
                for entry in data:
                    name = entry.get("cluster", "")
                    extras = float(entry.get("extra_cost_month", 0))
                    ip_count = len(entry.get("extra_ips", []))
                    if not name or extras <= 0:
                        continue
                    h = host_by_name.get(name)
                    if not h:
                        missing.append(name)
                        continue

                    existing_macro = None
                    current = 0.0
                    for m in h.get("macros", []):
                        if m["macro"] == "{$COST_MONTH}":
                            existing_macro = m
                            try:
                                current = float(m["value"])
                            except (ValueError, TypeError):
                                current = 0.0
                            break

                    new_val = round(current + extras, 2)
                    desc = f"{COST_SRC_CLUSTER_EXTRAS} base {current:.2f} + {ip_count} extra IP{'s' if ip_count > 1 else ''} ({extras:.2f})"
                    updates.append({
                        "hostid": h["hostid"],
                        "host": name,
                        "macroid": existing_macro["hostmacroid"] if existing_macro else None,
                        "current": current,
                        "extras": extras,
                        "new": new_val,
                        "desc": desc,
                    })

                # Build response
                total_delta = sum(u["extras"] for u in updates)
                header = f"**Cluster IP fees import** — {len(updates)} clusters, ${total_delta:,.2f}/mo added"
                if missing:
                    header += f" ({len(missing)} missing)"

                lines = [header + ("  DRY RUN" if dry_run else ""), ""]
                lines.append("| Cluster | Current | +Extras | New | IPs |")
                lines.append("|---------|---------|---------|-----|-----|")
                for u in sorted(updates, key=lambda x: -x["extras"])[:25]:
                    ip_count = u["desc"].split()[3]
                    lines.append(f"| {u['host']} | ${u['current']:.2f} | ${u['extras']:.2f} | ${u['new']:.2f} | {ip_count} |")
                if len(updates) > 25:
                    lines.append(f"*+{len(updates) - 25} more clusters*")
                if missing:
                    lines.append(f"\n**Missing (not found in Zabbix):** {', '.join(missing[:10])}")

                if dry_run:
                    lines.append("\n*Set dry_run=false to apply.*")
                    return "\n".join(lines)

                # Apply — parallel usermacro update/create
                async def _apply(u):
                    if u["macroid"]:
                        return await client.call("usermacro.update", {
                            "hostmacroid": u["macroid"],
                            "value": str(u["new"]),
                            "description": u["desc"],
                        })
                    return await client.call("usermacro.create", {
                        "hostid": u["hostid"],
                        "macro": "{$COST_MONTH}",
                        "value": str(u["new"]),
                        "description": u["desc"],
                    })

                results = await asyncio.gather(*[_apply(u) for u in updates], return_exceptions=True)
                errors = [str(r)[:80] for r in results if isinstance(r, Exception)]
                applied = sum(1 for r in results if not isinstance(r, Exception))

                lines.append(f"\n**Applied: {applied}/{len(updates)}**")
                if errors:
                    lines.append(f"Errors: {len(errors)}")
                    for e in errors[:3]:
                        lines.append(f"  - {e}")
                return "\n".join(lines)
            except (httpx.HTTPError, ValueError, OSError) as e:
                return f"Error: {e}"

    if "get_cost_summary" not in skip:

        @mcp.tool()
        async def get_cost_summary(instance: str = "") -> str:
            """Get a summary of server costs from {$COST_MONTH} macros.

            Args:
                instance: Zabbix instance name (optional, for multi-instance setups)
            """
            try:
                from zbbx_mcp.classify import classify_host as _classify_host
                from zbbx_mcp.classify import detect_provider

                client = resolver.resolve(instance)

                hosts, macros = await asyncio.gather(
                    client.call("host.get", {
                        "output": ["hostid", "host"],
                        "selectGroups": ["name"],
                        "selectInterfaces": ["ip"],
                        "filter": {"status": "0"},
                    }),
                    client.call("usermacro.get", {
                        "output": ["hostid", "value"],
                        "filter": {"macro": "{$COST_MONTH}"},
                    }),
                )

                host_map = {h["hostid"]: h for h in hosts}
                cost_map = {}
                for m in macros:
                    try:
                        cost_map[m["hostid"]] = float(m["value"])
                    except (ValueError, TypeError):
                        pass

                # Aggregate by product and provider
                prod_costs: dict[str, dict] = {}
                prov_costs: dict[str, dict] = {}

                for hid, cost in cost_map.items():
                    h = host_map.get(hid)
                    if not h:
                        continue
                    prod, tier = _classify_host(h.get("groups", []))
                    ip = host_ip(h)
                    provider = detect_provider(ip) if ip else "No IP"

                    key = f"{prod} / {tier}"
                    p = prod_costs.setdefault(key, {"count": 0, "total": 0.0})
                    p["count"] += 1
                    p["total"] += cost

                    pv = prov_costs.setdefault(provider, {"count": 0, "total": 0.0})
                    pv["count"] += 1
                    pv["total"] += cost

                total = sum(cost_map.values())
                costed = len(cost_map)
                uncosted = len(hosts) - costed

                parts = [
                    f"**Cost Summary: ${total:,.2f}/month (${total * 12:,.2f}/year)**",
                    f"Servers with cost: {costed} | Without: {uncosted}",
                    "",
                    "## By Product",
                    "| Product / Tier | Servers | Cost/Month | Cost/Year |",
                    "|---|---|---|---|",
                ]
                for key in sorted(prod_costs, key=lambda x: -prod_costs[x]["total"]):
                    p = prod_costs[key]
                    parts.append(f"| {key} | {p['count']} | ${p['total']:,.2f} | ${p['total']*12:,.2f} |")

                parts.extend([
                    "",
                    "## By Provider",
                    "| Provider | Servers | Cost/Month | Cost/Year |",
                    "|---|---|---|---|",
                ])
                for prov in sorted(prov_costs, key=lambda x: -prov_costs[x]["total"]):
                    p = prov_costs[prov]
                    parts.append(f"| {prov} | {p['count']} | ${p['total']:,.2f} | ${p['total']*12:,.2f} |")

                return "\n".join(parts)
            except (httpx.HTTPError, ValueError) as e:
                return f"Error: {e}"

    if "export_cost_audit" not in skip:

        @mcp.tool()
        async def export_cost_audit(
            output_xlsx: str = "~/Downloads/cost_audit.xlsx",
            mode: str = "estimated",
            source_xlsx: str = "",
            instance: str = "",
        ) -> str:
            """Export hosts with {$COST_MONTH} source for accounting review.

            mode:
              - estimated: only hosts whose cost is NOT backed by an exact
                billing match (bulk patterns, product/provider medians,
                manual extrapolations, empty descriptions).
              - all: every costed host, with classification column.

            source_xlsx: optional path to the source-of-truth workbook.
              When provided, every row is cross-referenced against the
              Ip-price / Sheet11 tabs and an `in_source_of_truth` column
              records whether the host's IP is present there. This is the
              authoritative "100% sure" signal — independent of macro
              descriptions.

            Output is a single-tab XLSX with host, IP, provider, product,
            tier, country, cost, source, and (when applicable)
            in_source_of_truth.
            """
            try:
                import openpyxl
                from openpyxl.styles import Alignment, Font, PatternFill
            except ImportError:
                return "openpyxl not installed"

            from zbbx_mcp.classify import classify_host as _classify_host
            from zbbx_mcp.classify import detect_provider
            from zbbx_mcp.data import extract_country

            try:
                client = resolver.resolve(instance)
                hosts, macros = await asyncio.gather(
                    client.call("host.get", {
                        "output": ["hostid", "host"],
                        "selectGroups": ["name"],
                        "selectInterfaces": ["ip"],
                        "filter": {"status": "0"},
                    }),
                    client.call("usermacro.get", {
                        "output": ["hostid", "value", "description"],
                        "filter": {"macro": "{$COST_MONTH}"},
                    }),
                )
            except (httpx.HTTPError, ValueError) as e:
                return f"Error: {e}"

            # Optional: parse source-of-truth XLSX and record sheet+row for each IP.
            source_ip_locations: dict[str, list[tuple[str, int]]] = {}
            source_loaded = False
            source_abs_path = ""
            if source_xlsx:
                try:
                    import openpyxl
                    source_abs_path = os.path.expanduser(source_xlsx)
                    wb_src = openpyxl.load_workbook(source_abs_path, data_only=True)
                    ip_re_local = _IP_RE
                    for sheet_name in wb_src.sheetnames:
                        ws_src = wb_src[sheet_name]
                        for r_idx, row in enumerate(ws_src.iter_rows(values_only=True), start=1):
                            for cell in row:
                                if cell is None:
                                    continue
                                for ip in ip_re_local.findall(str(cell)):
                                    source_ip_locations.setdefault(ip, []).append(
                                        (sheet_name, r_idx)
                                    )
                    source_loaded = True
                except (ImportError, OSError, ValueError) as e:
                    return f"Failed to load source_xlsx: {e}"

            host_by_id = {h["hostid"]: h for h in hosts}
            # Explicit "estimated" markers — anything else is assumed billing-backed
            # (legacy macros with generic descriptions pre-date the provenance tags
            # and were almost always written from billing matches).
            estimated_prefixes = (
                "src:bulk_pattern",
                "src:product_median",
                "src:provider_median",
                "estimated from",
                "twin cluster",
            )

            rows = []
            for m in macros:
                try:
                    cost = float(m.get("value") or 0)
                except (ValueError, TypeError):
                    continue
                if cost <= 0:
                    continue
                h = host_by_id.get(m["hostid"])
                if not h:
                    continue
                desc = (m.get("description") or "").strip()
                is_estimated = any(desc.startswith(p) for p in estimated_prefixes)
                # Heuristic: our only bulk import in this cycle used $210.82 with a
                # generic description — catch it explicitly.
                if not is_estimated and abs(cost - 210.82) < 0.01 and not desc.startswith(("src:billing_", "base ")):
                    is_estimated = True
                backed = not is_estimated
                if mode == "estimated" and backed:
                    continue
                ip = host_ip(h) or ""
                prov = detect_provider(ip) if ip else "?"
                prod, tier = _classify_host(h.get("groups", []))
                cc = extract_country(h["host"]) or ""
                src_locs = source_ip_locations.get(ip, []) if source_loaded else []
                rows.append({
                    "host": h["host"],
                    "host_id": h["hostid"],
                    "ip": ip,
                    "provider": prov,
                    "product": prod,
                    "tier": tier,
                    "country": cc,
                    "cost": cost,
                    "source": desc or "(no tag)",
                    "billing_backed": "yes" if backed else "no",
                    "in_source_of_truth": (
                        ("yes" if src_locs else "no") if source_loaded else ""
                    ),
                    "source_row": "; ".join(f"{s}!{r}" for s, r in src_locs[:3]),
                })

            rows.sort(key=lambda r: (-r["cost"], r["host"]))

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "cost_audit"
            hdr_font = Font(bold=True, color="FFFFFF")
            hdr_fill = PatternFill("solid", fgColor="366092")
            center = Alignment(horizontal="center")

            headers = ["Host", "Host ID", "IP", "Provider", "Product", "Tier",
                       "Country", "Cost $/mo", "Source", "Billing-backed"]
            if source_loaded:
                headers.extend(["In source of truth", "Source row"])
            for i, h in enumerate(headers, start=1):
                c = ws.cell(row=1, column=i, value=h)
                c.font = hdr_font
                c.fill = hdr_fill
                c.alignment = center

            green_fill = PatternFill("solid", fgColor="D9EAD3")
            yes_col_idx = 11  # "In source of truth" column index when present

            for i, r in enumerate(rows, start=2):
                vals = [r["host"], r["host_id"], r["ip"], r["provider"],
                        r["product"], r["tier"], r["country"], round(r["cost"], 2),
                        r["source"], r["billing_backed"]]
                if source_loaded:
                    vals.append(r["in_source_of_truth"])
                    vals.append(r["source_row"])
                for j, v in enumerate(vals, start=1):
                    cell = ws.cell(row=i, column=j, value=v)
                    # Hyperlink the source_row cell back into the workbook
                    if (source_loaded and j == yes_col_idx + 1
                            and r["source_row"] and source_abs_path):
                        first = r["source_row"].split(";")[0].strip()
                        if "!" in first:
                            sheet_part, row_part = first.split("!", 1)
                            cell.hyperlink = (
                                f"file://{source_abs_path}#'{sheet_part}'!A{row_part}"
                            )
                            cell.font = Font(color="0563C1", underline="single")
                # Highlight row green when present in source of truth
                if source_loaded and r["in_source_of_truth"] == "yes":
                    for j in range(1, len(vals) + 1):
                        ws.cell(row=i, column=j).fill = green_fill

            total_row = len(rows) + 3
            ws.cell(row=total_row, column=1, value=f"TOTAL ({len(rows)} hosts)").font = Font(bold=True)
            ws.cell(row=total_row, column=8, value=round(sum(r["cost"] for r in rows), 2)).font = Font(bold=True)

            widths = [30, 10, 16, 14, 18, 14, 10, 12, 38, 14]
            cols = "ABCDEFGHIJ"
            if source_loaded:
                widths.extend([18, 28])
                cols = "ABCDEFGHIJKL"
            for col, w in zip(cols, widths, strict=True):
                ws.column_dimensions[col].width = w

            out = os.path.expanduser(output_xlsx)
            wb.save(out)
            total = sum(r["cost"] for r in rows)
            lines = [
                f"Exported {len(rows)} hosts (mode={mode}), ${total:,.2f}/mo",
                f"Wrote: {out}",
            ]
            if source_loaded:
                yes = sum(1 for r in rows if r["in_source_of_truth"] == "yes")
                lines.append(
                    f"In source of truth: {yes} yes (green) / {len(rows) - yes} no — review needed"
                )
            return "\n".join(lines)

    if "detect_cost_anomalies" not in skip:

        @mcp.tool()
        async def detect_cost_anomalies(
            high_factor: float = 2.5,
            low_factor: float = 0.3,
            max_results: int = 50,
            instance: str = "",
        ) -> str:
            """Flag hosts whose {$COST_MONTH} is far from its provider median.

            Useful after an import to catch mis-allocations (e.g. a bulk
            aggregate applied as a per-server rate). Compares each costed
            host's value to its provider median and reports the outliers.

            Args:
                high_factor: flag hosts at or above this multiple of median (default: 2.5)
                low_factor: flag hosts at or below this multiple of median (default: 0.3)
                max_results: cap on rows shown
                instance: Zabbix instance (optional)
            """
            from zbbx_mcp.classify import detect_provider

            try:
                client = resolver.resolve(instance)
                medians = await _provider_medians(client)
                hosts, macros = await asyncio.gather(
                    client.call("host.get", {
                        "output": ["hostid", "host"],
                        "selectInterfaces": ["ip"],
                        "filter": {"status": "0"},
                    }),
                    client.call("usermacro.get", {
                        "output": ["hostid", "value", "description"],
                        "filter": {"macro": "{$COST_MONTH}"},
                    }),
                )
                host_by_id = {h["hostid"]: h for h in hosts}
                anomalies = []
                for m in macros:
                    try:
                        v = float(m.get("value") or 0)
                    except (ValueError, TypeError):
                        continue
                    if v <= 0:
                        continue
                    h = host_by_id.get(m["hostid"])
                    if not h:
                        continue
                    ip = host_ip(h)
                    if not ip:
                        continue
                    prov = detect_provider(ip)
                    med = medians.get(prov)
                    if not med:
                        continue
                    ratio = v / med
                    if ratio >= high_factor or ratio <= low_factor:
                        anomalies.append({
                            "host": h["host"],
                            "provider": prov,
                            "cost": v,
                            "median": med,
                            "ratio": ratio,
                            "source": (m.get("description") or "").strip() or "(no tag)",
                        })
                anomalies.sort(key=lambda a: -abs(a["ratio"] - 1))
                if not anomalies:
                    return "No cost anomalies detected."
                lines = [
                    f"**{len(anomalies)} cost anomalies detected**",
                    f"Thresholds: >{high_factor}× or <{low_factor}× provider median\n",
                    "| Host | Provider | $/mo | Median | Ratio | Source |",
                    "|------|----------|------|--------|-------|--------|",
                ]
                for a in anomalies[:max_results]:
                    lines.append(
                        f"| {a['host']} | {a['provider']} | ${a['cost']:.2f} | "
                        f"${a['median']:.2f} | {a['ratio']:.1f}× | {a['source']} |"
                    )
                if len(anomalies) > max_results:
                    lines.append(f"*+{len(anomalies) - max_results} more*")
                return "\n".join(lines)
            except (httpx.HTTPError, ValueError) as e:
                return f"Error: {e}"

    if "import_from_xlsx" not in skip:

        @mcp.tool()
        async def import_from_xlsx(
            file_path: str,
            output_csv: str = "/tmp/billing_xlsx.csv",
            eur_usd: float = 1.08,
        ) -> str:
            """Parse billing XLSX (Ip-price + Sheet11 structures) to a flat CSV.

            Expects two sheet shapes used by the current accounting file:

            - A detailed sheet with a primary IP column and a price-per-server
              column (columns L–O in the current layout). Addons in the next
              column are summed into the per-server price.
            - A simple sheet with columns: Name, NS, IPv4, Price (EUR) — prices
              may use comma decimal separator.

            Output CSV columns: ip, billing_name, price_monthly (USD).
            Only the primary IP of each row gets the price; siblings in the
            same row are written with price 0 so reconcile_billing_audit can
            still bucket them without double-counting.

            Args:
                file_path: Path to the XLSX file
                output_csv: Where to write the flat CSV (default /tmp/billing_xlsx.csv)
                eur_usd: EUR→USD conversion rate for sheets priced in EUR (default 1.08)
            """
            try:
                import openpyxl
            except ImportError:
                return "openpyxl not installed. Run: uv pip install openpyxl"

            try:
                path = os.path.expanduser(file_path)
                if not os.path.isfile(path):
                    return f"File not found: {path}"
                wb = openpyxl.load_workbook(path, data_only=True)
            except (OSError, ValueError) as e:
                return f"Failed to open XLSX: {e}"

            rows: list[dict] = []
            ip_re = _IP_RE

            def _pick_price(values, max_idx=17):
                """Return first positive price in cols 13–16 (server/Price$/Addons)."""
                p1 = values[13] if len(values) > 13 else None
                p2 = values[14] if len(values) > 14 else None
                p_add = values[16] if len(values) > 16 else None
                price = None
                for p in (p1, p2):
                    if isinstance(p, (int, float)) and 0 < p < 5000:
                        price = float(p)
                        break
                if price is None:
                    return None
                if isinstance(p_add, (int, float)) and p_add > 0:
                    price += float(p_add)
                return price

            # Structured detailed sheets (Ip-price style)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                headers = [
                    str(c or "").strip().lower()
                    for c in next(ws.iter_rows(values_only=True), ())
                ]
                # Heuristic: detailed sheet has an "ip server" and "price server" header
                has_ip_col = any("ip server" in h or "ip сервера" in h for h in headers)
                has_price_col = any("price server" in h or h.startswith("price") for h in headers)
                if not (has_ip_col and has_price_col):
                    continue
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i < 1:
                        continue
                    name = row[11] if len(row) > 11 else None
                    ip_cell = row[12] if len(row) > 12 else None
                    if not ip_cell:
                        continue
                    ips = ip_re.findall(str(ip_cell))
                    if not ips:
                        continue
                    price = _pick_price(row)
                    if price is None:
                        continue
                    for j, ip in enumerate(ips):
                        rows.append({
                            "ip": ip,
                            "name": str(name or "").strip(),
                            "price": price if j == 0 else 0.0,
                        })

            # Simple sheets (Sheet11 style): Name | NS | IPv4 | Price
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                first = next(ws.iter_rows(values_only=True), ())
                headers = [str(c or "").strip().lower() for c in first]
                if len(headers) < 4:
                    continue
                if not (headers[0] in ("name", "hostname") and
                        headers[2] in ("ipv4", "ip", "ip address") and
                        headers[3].startswith("price")):
                    continue
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i < 1:
                        continue
                    name, _ns, ip, pr = row[:4]
                    if not ip:
                        continue
                    if isinstance(pr, str):
                        try:
                            price = float(pr.replace(",", "."))
                        except ValueError:
                            continue
                    elif isinstance(pr, (int, float)):
                        price = float(pr)
                    else:
                        continue
                    # Sheet11 prices are in EUR — convert.
                    rows.append({
                        "ip": str(ip).strip(),
                        "name": str(name or "").strip(),
                        "price": price * eur_usd,
                    })

            # Dedupe by ip, keep max price
            by_ip: dict[str, dict] = {}
            for r in rows:
                if r["ip"] in by_ip and r["price"] <= by_ip[r["ip"]]["price"]:
                    continue
                by_ip[r["ip"]] = r

            out_path = os.path.expanduser(output_csv)
            with open(out_path, "w", newline="") as f:
                w = _csv.writer(f)
                w.writerow(["ip", "billing_name", "price_monthly"])
                for r in by_ip.values():
                    w.writerow([r["ip"], r["name"], f"{r['price']:.2f}"])

            total = sum(r["price"] for r in by_ip.values())
            paid = sum(1 for r in by_ip.values() if r["price"] > 0)
            return (
                f"Parsed {len(by_ip)} unique IPs from {len(wb.sheetnames)} sheets\n"
                f"Priced: {paid} | Extras (price=0): {len(by_ip) - paid}\n"
                f"Total: ${total:,.2f}/mo\n"
                f"Wrote: {out_path}"
            )

    if "fill_cost_median" not in skip:

        @mcp.tool()
        async def fill_cost_median(
            group_by: str = "product",
            dry_run: bool = True,
            instance: str = "",
        ) -> str:
            """Estimate {$COST_MONTH} for empty-cost hosts using peer median.

            For each host without a non-zero {$COST_MONTH}, take the median of
            costed hosts sharing the same grouping key and assign it. Skips
            hosts without an IP (monitoring) so they stay out of cost totals.

            Args:
                group_by: "product" (product/tier from host groups) or "provider" (CIDR detection)
                dry_run: Preview only (default: True)
                instance: Zabbix instance (optional)
            """
            import statistics

            from zbbx_mcp.classify import classify_host as _classify_host
            from zbbx_mcp.classify import detect_provider

            try:
                client = resolver.resolve(instance)
                hosts, macros = await asyncio.gather(
                    client.call("host.get", {
                        "output": ["hostid", "host"],
                        "selectGroups": ["name"],
                        "selectInterfaces": ["ip"],
                        "filter": {"status": "0"},
                    }),
                    client.call("usermacro.get", {
                        "output": ["hostmacroid", "hostid", "value"],
                        "filter": {"macro": "{$COST_MONTH}"},
                    }),
                )

                cost_by_hid: dict[str, float] = {}
                macro_by_hid: dict[str, str] = {}
                for m in macros:
                    macro_by_hid[m["hostid"]] = m["hostmacroid"]
                    try:
                        v = float(m.get("value") or 0)
                        if v > 0:
                            cost_by_hid[m["hostid"]] = v
                    except (ValueError, TypeError):
                        pass

                def _key(h):
                    if group_by == "provider":
                        ip = host_ip(h)
                        return detect_provider(ip) if ip else None
                    prod, tier = _classify_host(h.get("groups", []))
                    return f"{prod}/{tier}"

                bucket: dict[str, list[float]] = {}
                for h in hosts:
                    if h["hostid"] in cost_by_hid:
                        k = _key(h)
                        if k:
                            bucket.setdefault(k, []).append(cost_by_hid[h["hostid"]])

                medians = {k: statistics.median(v) for k, v in bucket.items() if v}

                candidates = []
                skipped_no_ip = 0
                skipped_no_peer = 0
                for h in hosts:
                    if h["hostid"] in cost_by_hid:
                        continue
                    ip = host_ip(h)
                    if not ip:
                        skipped_no_ip += 1
                        continue
                    k = _key(h)
                    if not k or k not in medians:
                        skipped_no_peer += 1
                        continue
                    candidates.append((h, k, medians[k]))

                total_delta = sum(c for _, _, c in candidates)
                if dry_run:
                    lines = [
                        f"**Fill cost by {group_by} median — DRY RUN**",
                        f"Candidates: {len(candidates)} hosts, ${total_delta:,.2f}/mo",
                        f"Skipped: {skipped_no_ip} no-IP, {skipped_no_peer} no peer",
                        "",
                        f"| Host | {group_by.title()} | Median $/mo |",
                        "|------|---------|-------------|",
                    ]
                    by_group: dict[str, int] = {}
                    for _h, k, _c in candidates:
                        by_group[k] = by_group.get(k, 0) + 1
                    for _h, k, c in sorted(candidates, key=lambda x: -x[2])[:25]:
                        lines.append(f"| {_h['host']} | {k} | ${c:.2f} |")
                    if len(candidates) > 25:
                        lines.append(f"*+{len(candidates) - 25} more*")
                    lines.append("")
                    lines.append("**By group:**")
                    for k, n in sorted(by_group.items(), key=lambda x: -x[1]):
                        lines.append(f"- {k}: {n} hosts @ ${medians[k]:.2f}")
                    lines.append("\nSet dry_run=false to apply.")
                    return "\n".join(lines)

                created = 0
                errors = []
                for h, _k, cost in candidates:
                    val = str(round(cost, 2))
                    try:
                        if h["hostid"] in macro_by_hid:
                            await client.call("usermacro.update", {
                                "hostmacroid": macro_by_hid[h["hostid"]],
                                "value": val,
                                "description": COST_SRC_PRODUCT_MEDIAN if group_by == "product" else COST_SRC_PROVIDER_MEDIAN,
                            })
                        else:
                            await client.call("usermacro.create", {
                                "hostid": h["hostid"],
                                "macro": "{$COST_MONTH}",
                                "value": val,
                                "description": COST_SRC_PRODUCT_MEDIAN if group_by == "product" else COST_SRC_PROVIDER_MEDIAN,
                            })
                        created += 1
                    except (httpx.HTTPError, ValueError) as e:
                        errors.append(f"{h['host']}: {e}")
                        if len(errors) >= 10:
                            break
                parts = [
                    f"**Filled {created}/{len(candidates)} empty-cost hosts with {group_by} median**",
                    f"Added: ${total_delta:,.2f}/mo",
                ]
                if errors:
                    parts.append(f"Errors: {len(errors)}")
                return "\n".join(parts)
            except (httpx.HTTPError, ValueError) as e:
                return f"Error: {e}"

    if "get_cost_gaps" not in skip:

        @mcp.tool()
        async def get_cost_gaps(
            max_results: int = 30,
            instance: str = "",
        ) -> str:
            """Servers without cost data, grouped by provider and product.

            Args:
                max_results: Max rows (default: 30)
                instance: Zabbix instance name (optional)
            """
            try:
                from zbbx_mcp.classify import classify_host as _classify_host
                from zbbx_mcp.classify import detect_provider

                client = resolver.resolve(instance)
                hosts, macros = await asyncio.gather(
                    client.call("host.get", {
                        "output": ["hostid", "host"],
                        "selectGroups": ["name"],
                        "selectInterfaces": ["ip"],
                        "filter": {"status": "0"},
                    }),
                    client.call("usermacro.get", {
                        "output": ["hostid"],
                        "filter": {"macro": "{$COST_MONTH}"},
                    }),
                )

                has_cost = {m["hostid"] for m in macros}
                total = len(hosts)
                costed = sum(1 for h in hosts if h["hostid"] in has_cost)

                gaps: dict[str, dict] = {}
                no_ip = 0
                for h in hosts:
                    if h["hostid"] in has_cost:
                        continue
                    ip = host_ip(h)
                    if not ip:
                        no_ip += 1
                        continue
                    prod, tier = _classify_host(h.get("groups", []))
                    prov = detect_provider(ip)
                    key = f"{prov} / {prod}"
                    entry = gaps.setdefault(key, {"count": 0, "hosts": []})
                    entry["count"] += 1
                    if len(entry["hosts"]) < 3:
                        entry["hosts"].append(h["host"])

                lines = [
                    f"**Cost gaps: {total - costed} without cost** ({costed}/{total} covered)\n",
                    "| Provider / Product | Missing | Sample Hosts |",
                    "|-------------------|---------|-------------|",
                ]
                for key in sorted(gaps, key=lambda k: -gaps[k]["count"])[:max_results]:
                    g = gaps[key]
                    lines.append(f"| {key} | {g['count']} | {', '.join(g['hosts'])} |")

                if no_ip:
                    lines.append(f"\n*{no_ip} hosts without IP (monitoring) — no cost expected*")
                return "\n".join(lines)
            except (httpx.HTTPError, ValueError) as e:
                return f"Error: {e}"

    if "get_cost_efficiency" not in skip:

        @mcp.tool()
        async def get_cost_efficiency(
            max_results: int = 20,
            instance: str = "",
        ) -> str:
            """Cost per Gbps by country and provider — find overpay and waste.

            Args:
                max_results: Max rows (default: 20)
                instance: Zabbix instance name (optional)
            """
            try:
                from zbbx_mcp.classify import detect_provider
                from zbbx_mcp.data import extract_country, fetch_traffic_map

                client = resolver.resolve(instance)
                hosts, macros = await asyncio.gather(
                    client.call("host.get", {
                        "output": ["hostid", "host"],
                        "selectGroups": ["name"],
                        "selectInterfaces": ["ip"],
                        "filter": {"status": "0"},
                    }),
                    client.call("usermacro.get", {
                        "output": ["hostid", "value"],
                        "filter": {"macro": "{$COST_MONTH}"},
                    }),
                )

                cost_map = {}
                for m in macros:
                    try:
                        cost_map[m["hostid"]] = float(m["value"])
                    except (ValueError, TypeError):
                        pass

                traffic_map = await fetch_traffic_map(client, list(cost_map.keys()))

                by_country: dict[str, dict] = {}
                by_provider: dict[str, dict] = {}
                waste = []

                for hid, cost in cost_map.items():
                    h = next((x for x in hosts if x["hostid"] == hid), None)
                    if not h:
                        continue
                    traffic_mbps = traffic_map.get(hid, 0)
                    traffic_gbps = traffic_mbps / 1000
                    cc = extract_country(h["host"])
                    ip = host_ip(h)
                    prov = detect_provider(ip) if ip else "?"

                    if cc:
                        c = by_country.setdefault(cc, {"servers": 0, "cost": 0, "traffic": 0})
                        c["servers"] += 1
                        c["cost"] += cost
                        c["traffic"] += traffic_gbps

                    p = by_provider.setdefault(prov, {"servers": 0, "cost": 0, "traffic": 0})
                    p["servers"] += 1
                    p["cost"] += cost
                    p["traffic"] += traffic_gbps

                    if traffic_mbps < 1 and cost > 50:
                        waste.append((h["host"], cc, prov, cost))

                lines = [f"**Cost Efficiency** ({len(cost_map)} servers)\n"]
                lines.append("**By Country (most expensive per Gbps first):**")
                lines.append("| Country | Servers | $/mo | Gbps | $/Gbps |")
                lines.append("|---------|---------|------|------|--------|")
                sorted_cc = sorted(by_country.items(), key=lambda x: -(x[1]["cost"] / max(x[1]["traffic"], 0.001)))
                for cc, c in sorted_cc[:max_results]:
                    if c["traffic"] < 0.01:
                        lines.append(f"| {cc} | {c['servers']} | ${c['cost']:,.0f} | idle | N/A |")
                    else:
                        per = c["cost"] / c["traffic"]
                        lines.append(f"| {cc} | {c['servers']} | ${c['cost']:,.0f} | {c['traffic']:.1f} | ${per:,.0f} |")

                lines.append("\n**By Provider:**")
                lines.append("| Provider | Servers | $/mo | Gbps | $/Gbps |")
                lines.append("|----------|---------|------|------|--------|")
                for prov in sorted(by_provider, key=lambda p: -by_provider[p]["cost"])[:15]:
                    p = by_provider[prov]
                    if p["traffic"] < 0.01:
                        lines.append(f"| {prov} | {p['servers']} | ${p['cost']:,.0f} | idle | N/A |")
                    else:
                        per = p["cost"] / p["traffic"]
                        lines.append(f"| {prov} | {p['servers']} | ${p['cost']:,.0f} | {p['traffic']:.1f} | ${per:,.0f} |")

                if waste:
                    waste.sort(key=lambda w: -w[3])
                    total_waste = sum(w[3] for w in waste)
                    lines.append(f"\n**Waste: {len(waste)} servers paying ${total_waste:,.0f}/mo with 0 traffic:**")
                    lines.append("| Server | Country | Provider | $/mo |")
                    lines.append("|--------|---------|----------|------|")
                    for host, cc, prov, cost in waste[:10]:
                        lines.append(f"| {host} | {cc or '?'} | {prov} | ${cost:.0f} |")
                    if len(waste) > 10:
                        lines.append(f"*+{len(waste) - 10} more*")

                return "\n".join(lines)
            except (httpx.HTTPError, ValueError) as e:
                return f"Error: {e}"

    if "analyze_cost_import" not in skip:

        @mcp.tool()
        async def analyze_cost_import(
            file_path: str,
            output_csv: str = "",
            output_json: str = "",
            instance: str = "",
        ) -> str:
            """Tiered probability analysis of unmatched cost entries.

            Reads a JSON file of {ip: price} entries, finds which don't match
            Zabbix hosts by IP, then scores each with signals:
            - /24 subnet match (40pt)
            - Billing name fuzzy match (35pt strong / 15pt partial)
            - Known provider CIDR (15pt)

            Tiers: HIGH (>=70), MEDIUM (>=40), LOW (>=15), UNKNOWN (<15).
            Writes CSV + JSON for human review.

            Args:
                file_path: Path to JSON cost file {ip: price} or {ip: {name, price, ...}}
                output_csv: Path for CSV output (default: ~/Downloads/cost_import_analysis.csv)
                output_json: Path for JSON output (default: ~/Downloads/cost_import_analysis.json)
                instance: Zabbix instance name (optional)
            """
            from collections import defaultdict

            from zbbx_mcp.classify import detect_provider

            try:
                path = os.path.expanduser(file_path)
                if not os.path.isfile(path):
                    return f"File not found: {path}"
                with open(path) as f:
                    raw = json.load(f)

                # Normalize input: accept both {ip: price} and {ip: {name, price, ...}}
                costs: dict[str, dict] = {}
                for ip, val in raw.items():
                    if isinstance(val, dict):
                        price = val.get("price") or val.get("cost") or 0
                        name = val.get("name", "")
                    else:
                        try:
                            price = float(val)
                            name = ""
                        except (ValueError, TypeError):
                            continue
                    if price > 0:
                        costs[ip.strip()] = {"name": str(name), "price": float(price)}

                if not costs:
                    return "No valid cost entries found in file."

                client = resolver.resolve(instance)
                hosts = await client.call("host.get", {
                    "output": ["hostid", "host", "name"],
                    "selectInterfaces": ["ip"],
                })

                zabbix_ips: dict[str, str] = {}
                zabbix_by_24: dict[str, list] = defaultdict(list)
                zabbix_names: list[str] = []
                for h in hosts:
                    zabbix_names.append(h["host"].lower())
                    for iface in h.get("interfaces", []):
                        ip = iface.get("ip", "")
                        if ip and ip != "127.0.0.1":
                            zabbix_ips[ip] = h["host"]
                            parts = ip.split(".")
                            if len(parts) == 4:
                                zabbix_by_24[f"{parts[0]}.{parts[1]}.{parts[2]}"].append((ip, h["host"]))

                def _tokens(s):
                    if not s:
                        return []
                    tokens = re.findall(r"[a-z]{2,}\d*|\d{2,}", s.lower())
                    stop = {"dedicated", "server", "custom", "dual", "standard", "gold", "price", "incl", "excl", "taxes"}
                    return [t for t in tokens if t not in stop and len(t) >= 2]

                results = []
                _PRIV = ("10.", "192.168.", "172.16.", "172.17.", "172.18.", "172.19.",
                         "172.20.", "172.21.", "172.22.", "172.23.", "172.24.", "172.25.",
                         "172.26.", "172.27.", "172.28.", "172.29.", "172.30.", "172.31.")

                for ip, info in costs.items():
                    if ip.startswith(_PRIV):
                        continue
                    if ip in zabbix_ips:
                        continue  # already matched

                    score = 0
                    signals = []

                    parts = ip.split(".")
                    if len(parts) == 4:
                        subnet = zabbix_by_24.get(f"{parts[0]}.{parts[1]}.{parts[2]}", [])
                        if subnet:
                            score += 40
                            signals.append(f"/24 match: {subnet[0][1]} (+{len(subnet) - 1} more)")

                    name_tokens = _tokens(info["name"])
                    matched_hosts = []
                    if name_tokens:
                        for hn in zabbix_names:
                            if all(t in hn for t in name_tokens[:2]):
                                matched_hosts.append(hn)
                                if len(matched_hosts) >= 3:
                                    break
                        if matched_hosts:
                            score += 35
                            signals.append(f"name match: {', '.join(matched_hosts[:2])}")
                        else:
                            for hn in zabbix_names:
                                for t in name_tokens:
                                    if len(t) >= 4 and t in hn:
                                        matched_hosts.append(hn)
                                        break
                                if len(matched_hosts) >= 2:
                                    break
                            if matched_hosts:
                                score += 15
                                signals.append(f"partial name: {matched_hosts[0]}")

                    prov = detect_provider(ip)
                    if prov not in ("Unknown", "Other"):
                        score += 15
                        signals.append(f"provider: {prov}")

                    if score >= 70:
                        tier = "HIGH"
                        sug = "Likely same server (different IP or hostname) — verify then add or update"
                    elif score >= 40:
                        tier = "MEDIUM"
                        sug = "Possible match — review subnet/name before importing"
                    elif score >= 15:
                        tier = "LOW"
                        sug = "Weak signal — new server not onboarded, or decommissioned"
                    else:
                        tier = "UNKNOWN"
                        sug = "No signal — external infra, abandoned, or needs manual investigation"

                    results.append({
                        "ip": ip, "name": info["name"], "price_monthly": info["price"],
                        "tier": tier, "confidence": score, "signals": signals, "suggestion": sug,
                    })

                tier_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2, "UNKNOWN": 3}
                results.sort(key=lambda x: (tier_order[x["tier"]], -x["price_monthly"]))

                by_tier: dict[str, dict] = defaultdict(lambda: {"count": 0, "cost": 0.0})
                for r in results:
                    by_tier[r["tier"]]["count"] += 1
                    by_tier[r["tier"]]["cost"] += r["price_monthly"]

                downloads = os.path.expanduser("~/Downloads")
                out_json = os.path.expanduser(output_json) if output_json else os.path.join(downloads, "cost_import_analysis.json")
                out_csv = os.path.expanduser(output_csv) if output_csv else os.path.join(downloads, "cost_import_analysis.csv")

                with open(out_json, "w") as f:
                    json.dump(results, f, indent=2)

                import csv as _csv
                with open(out_csv, "w", newline="") as f:
                    w = _csv.writer(f)
                    w.writerow(["Tier", "Confidence", "IP", "Billing Name", "Monthly $",
                                "Annual $", "Signals", "Suggestion"])
                    for r in results:
                        w.writerow([r["tier"], r["confidence"], r["ip"], r["name"],
                                    f"{r['price_monthly']:.2f}",
                                    f"{r['price_monthly'] * 12:.2f}",
                                    " | ".join(r["signals"]), r["suggestion"]])

                total_unmatched = sum(r["price_monthly"] for r in results)
                lines = [
                    f"**Cost Import Analysis: {len(results)} unmatched IPs = ${total_unmatched:,.2f}/mo**\n",
                    "| Tier | Count | Monthly $ | Annual $ |",
                    "|------|-------|-----------|----------|",
                ]
                for t in ["HIGH", "MEDIUM", "LOW", "UNKNOWN"]:
                    d = by_tier[t]
                    lines.append(f"| {t} | {d['count']} | ${d['cost']:,.2f} | ${d['cost'] * 12:,.2f} |")
                lines.append(f"\nSaved: `{out_json}`")
                lines.append(f"Saved: `{out_csv}`")

                return "\n".join(lines)
            except (httpx.HTTPError, ValueError, OSError) as e:
                return f"Error: {e}"

    if "reconcile_billing_audit" not in skip:

        @mcp.tool()
        async def reconcile_billing_audit(
            file_path: str,
            output_dir: str = "",
            instance: str = "",
        ) -> str:
            """Categorize billing entries into finance action buckets.

            Buckets: importable, already_costed, stale_ip, subnet_match, onboard, cancel.

            Args:
                file_path: CSV with ip, billing_name, price_monthly columns
                output_dir: Directory to write per-bucket CSVs (default: same as input)
                instance: Zabbix instance (optional)
            """
            try:
                rows = _load_billing_csv(file_path)
            except (FileNotFoundError, OSError) as e:
                return f"Failed to read CSV: {e}"
            if not rows:
                return "No valid billing rows found."

            try:
                client = resolver.resolve(instance)
                hosts = await client.call("host.get", {
                    "output": ["hostid", "host"],
                    "selectInterfaces": ["ip"],
                    "filter": {"status": "0"},
                })
                macros = await client.call("usermacro.get", {
                    "output": ["hostid", "value"],
                    "filter": {"macro": "{$COST_MONTH}"},
                })
                has_cost = {m["hostid"] for m in macros if m.get("value") and m["value"] != "0"}
            except httpx.HTTPError as e:
                return f"Zabbix error: {e}"

            ip_to_host: dict[str, dict] = {}
            subnet_to_host: dict[str, list] = {}
            name_to_host: dict[str, dict] = {}
            for h in hosts:
                hip = host_ip(h)
                if hip:
                    ip_to_host[hip] = h
                    subnet_to_host.setdefault(".".join(hip.split(".")[:3]), []).append(h)
                name_to_host[h["host"].lower()] = h

            buckets: dict[str, list] = {
                "importable": [], "already_costed": [], "stale_ip": [],
                "subnet_match": [], "onboard": [], "cancel": [],
            }

            for r in rows:
                ip, name = r["ip"], r["name"]
                h = ip_to_host.get(ip)
                if h:
                    entry = {**r, "zabbix_host": h["host"]}
                    buckets["already_costed" if h["hostid"] in has_cost else "importable"].append(entry)
                    continue
                low = name.lower()
                matched = None
                if low:
                    for key, hh in name_to_host.items():
                        first = low.split()[0] if low else ""
                        if key == low or (first and key == first):
                            matched = hh
                            break
                if matched:
                    buckets["stale_ip"].append({
                        **r, "zabbix_host": matched["host"], "zabbix_ip": host_ip(matched),
                    })
                    continue
                subnet = ".".join(ip.split(".")[:3])
                if subnet in subnet_to_host:
                    sample = subnet_to_host[subnet][0]
                    buckets["subnet_match"].append({
                        **r, "zabbix_host": sample["host"], "zabbix_ip": host_ip(sample),
                    })
                    continue
                if name and any(c.isalpha() for c in name) and "." not in name[:5]:
                    buckets["onboard"].append(r)
                else:
                    buckets["cancel"].append(r)

            if not output_dir:
                output_dir = os.path.dirname(os.path.expanduser(file_path)) or os.getcwd()
            base = os.path.splitext(os.path.basename(file_path))[0]
            written = []
            for bucket, items in buckets.items():
                if not items:
                    continue
                out = os.path.join(output_dir, f"{base}__{bucket}.csv")
                cols = sorted({k for it in items for k in it})
                with open(out, "w", newline="") as f:
                    w = _csv.DictWriter(f, fieldnames=cols)
                    w.writeheader()
                    w.writerows(items)
                written.append(out)

            actions = {
                "importable": "Safe to import (IP match, no cost)",
                "already_costed": "Skip — cost already set",
                "stale_ip": "Billing team: update IP",
                "subnet_match": "Likely new member — review",
                "onboard": "Ops: add host to Zabbix",
                "cancel": "Finance: cancel or investigate",
            }
            lines = [f"**Reconciliation: {len(rows)} billing rows**\n",
                     "| Bucket | Count | $/mo | Action |",
                     "|--------|------:|-----:|--------|"]
            for b, items in buckets.items():
                if not items:
                    continue
                total = sum(i["price"] for i in items)
                lines.append(f"| {b} | {len(items)} | ${total:,.2f} | {actions[b]} |")
            lines.append(f"\nWrote {len(written)} bucket CSVs to `{output_dir}`")
            return "\n".join(lines)

    if "find_stale_billing_ips" not in skip:

        @mcp.tool()
        async def find_stale_billing_ips(
            file_path: str,
            instance: str = "",
        ) -> str:
            """Detect billing entries where name matches a Zabbix host but IP differs.

            Args:
                file_path: CSV with ip, billing_name, price_monthly columns
                instance: Zabbix instance (optional)
            """
            try:
                rows = _load_billing_csv(file_path)
            except (FileNotFoundError, OSError) as e:
                return f"Failed to read CSV: {e}"
            if not rows:
                return "No valid billing rows."

            try:
                client = resolver.resolve(instance)
                hosts = await client.call("host.get", {
                    "output": ["hostid", "host"],
                    "selectInterfaces": ["ip"],
                    "filter": {"status": "0"},
                })
            except httpx.HTTPError as e:
                return f"Zabbix error: {e}"

            name_to_host: dict[str, dict] = {}
            ip_to_host: dict[str, dict] = {}
            for h in hosts:
                name_to_host[h["host"].lower()] = h
                hip = host_ip(h)
                if hip:
                    ip_to_host[hip] = h

            stale: list[dict] = []
            for r in rows:
                ip, name = r["ip"], r["name"].lower()
                if not name or ip in ip_to_host:
                    continue
                matched = name_to_host.get(name)
                if not matched and name:
                    first = name.split()[0]
                    matched = name_to_host.get(first)
                if matched:
                    zip_ = host_ip(matched)
                    if zip_ and zip_ != ip:
                        stale.append({
                            "billing_name": r["name"],
                            "billing_ip": ip,
                            "zabbix_host": matched["host"],
                            "zabbix_ip": zip_,
                            "price_monthly": r["price"],
                        })

            if not stale:
                return "No stale billing IPs detected."

            total = sum(s["price_monthly"] for s in stale)
            lines = [f"**{len(stale)} stale billing IPs** (${total:,.2f}/mo affected)\n",
                     "| Billing name | Billing IP | Zabbix IP | $/mo |",
                     "|-------------|-----------|-----------|-----:|"]
            for s in sorted(stale, key=lambda x: -x["price_monthly"])[:50]:
                lines.append(
                    f"| {s['billing_name']} | {s['billing_ip']} | {s['zabbix_ip']} | ${s['price_monthly']:.2f} |"
                )
            if len(stale) > 50:
                lines.append(f"\n*+{len(stale) - 50} more*")
            return "\n".join(lines)
