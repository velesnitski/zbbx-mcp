"""Full infrastructure report: dashboards → hosts → metrics → Excel export."""

import asyncio
import os
import tempfile
from datetime import datetime, timezone
from statistics import median

import httpx

from zbbx_mcp.resolver import InstanceResolver
from zbbx_mcp.tools.inventory import _classify_host, detect_provider


def register(mcp, resolver: InstanceResolver, skip: set[str] = frozenset()):

    if "generate_server_report" not in skip:

        @mcp.tool()
        async def generate_server_report(
            product: str = "",
            country: str = "",
            output_dir: str = "",
            instance: str = "",
        ) -> str:
            """Generate a full server report with Excel export.

            Queries all dashboards, resolves hosts from graph widgets,
            fetches CPU/load/memory metrics, classifies by product and provider.

            The Excel file contains:
            - Sheet 1: Full server list with all columns
            - Sheet 2: Product summary (aggregated)
            - Sheet 3: Provider summary (aggregated)

            Args:
                product: Filter by product name (optional)
                country: Filter by country code in hostname (optional)
                output_dir: Directory for the Excel file (default: ~/Downloads)
                instance: Zabbix instance name (optional, for multi-instance setups)
            """
            try:
                client = resolver.resolve(instance)

                # Phase 1: Fetch dashboards + hosts in parallel (2 calls)
                dashboards_task = client.call("dashboard.get", {
                    "output": ["dashboardid", "name"],
                    "selectPages": "extend",
                })
                hosts_task = client.call("host.get", {
                    "output": ["hostid", "host", "name", "status"],
                    "selectGroups": ["name"],
                    "selectInterfaces": ["ip"],
                    "filter": {"status": "0"},
                })

                dashboards, hosts = await asyncio.gather(dashboards_task, hosts_task)

                # Extract graph IDs from dashboards
                graph_ids = set()
                for d in dashboards:
                    for page in d.get("pages", []):
                        for w in page.get("widgets", []):
                            for f in w.get("fields", []):
                                if f.get("type") == "6":
                                    graph_ids.add(f["value"])

                # Phase 2: Resolve graphs → hosts + fetch metrics (4 calls parallel)
                host_map = {h["hostid"]: h for h in hosts}
                all_ids = list(host_map.keys())

                tasks = [
                    client.call("item.get", {
                        "hostids": all_ids,
                        "output": ["hostid", "lastvalue"],
                        "filter": {"key_": "system.cpu.util[,idle]"},
                    }),
                    client.call("item.get", {
                        "hostids": all_ids,
                        "output": ["hostid", "lastvalue"],
                        "filter": {"key_": "system.cpu.load[percpu,avg5]"},
                    }),
                    client.call("item.get", {
                        "hostids": all_ids,
                        "output": ["hostid", "lastvalue"],
                        "filter": {"key_": "vm.memory.size[available]"},
                    }),
                ]
                if graph_ids:
                    tasks.append(client.call("graph.get", {
                        "graphids": list(graph_ids),
                        "output": ["graphid"],
                        "selectHosts": ["hostid"],
                    }))

                results = await asyncio.gather(*tasks)
                cpu_items, load_items, mem_items = results[0], results[1], results[2]
                graphs = results[3] if len(results) > 3 else []

                # Build graph → host mapping
                graph_to_host = {}
                dashboard_host_ids = set()
                for g in graphs:
                    for h in g.get("hosts", []):
                        graph_to_host[g["graphid"]] = h["hostid"]
                        dashboard_host_ids.add(h["hostid"])

                # Build host → dashboard+tab mapping
                host_dashboards: dict[str, set[str]] = {}
                host_tabs: dict[str, set[str]] = {}
                for d in dashboards:
                    dname = d["name"]
                    for pi, page in enumerate(d.get("pages", [])):
                        tab = page.get("name", "") or f"Page {pi + 1}"
                        for w in page.get("widgets", []):
                            for f in w.get("fields", []):
                                if f.get("type") == "6":
                                    hid = graph_to_host.get(f["value"])
                                    if hid:
                                        host_dashboards.setdefault(hid, set()).add(dname)
                                        host_tabs.setdefault(hid, set()).add(f"{dname} / {tab}")

                # Build metrics maps
                cpu_map, load_map, mem_map = {}, {}, {}
                for i in cpu_items:
                    try:
                        cpu_map[i["hostid"]] = round(100 - float(i["lastvalue"]), 1)
                    except (ValueError, TypeError):
                        pass
                for i in load_items:
                    try:
                        load_map[i["hostid"]] = round(float(i["lastvalue"]), 2)
                    except (ValueError, TypeError):
                        pass
                for i in mem_items:
                    try:
                        mem_map[i["hostid"]] = round(float(i["lastvalue"]) / 1_073_741_824, 1)
                    except (ValueError, TypeError):
                        pass

                # Build rows
                rows = []
                for hid, h in host_map.items():
                    prod, tier = _classify_host(h.get("groups", []))
                    if not prod or prod == "Unknown":
                        continue
                    if product and product.lower() not in prod.lower():
                        continue
                    hostname = h.get("host", "")
                    if country and country.lower() not in hostname.lower():
                        continue

                    ip = ""
                    for iface in h.get("interfaces", []):
                        if iface.get("ip") and iface["ip"] != "127.0.0.1":
                            ip = iface["ip"]
                            break

                    provider = detect_provider(ip) if ip else ""
                    on_dashboard = hid in dashboard_host_ids
                    dboards = ", ".join(sorted(host_dashboards.get(hid, set())))
                    tabs = ", ".join(sorted(host_tabs.get(hid, set())))
                    groups = ", ".join(g["name"] for g in h.get("groups", []))

                    rows.append({
                        "Host": hostname,
                        "Name": h.get("name", ""),
                        "Product": prod,
                        "Tier": tier,
                        "IP": ip,
                        "Provider": provider,
                        "CPU %": cpu_map.get(hid),
                        "Load Avg5": load_map.get(hid),
                        "Mem Avail GB": mem_map.get(hid),
                        "On Dashboard": "Yes" if on_dashboard else "No",
                        "Dashboards": dboards,
                        "Dashboard Tabs": tabs,
                        "Groups": groups,
                    })

                rows.sort(key=lambda r: (r["Product"], r["Tier"], r["Host"]))

                # Build product summary
                prod_summary: dict[str, dict] = {}
                for r in rows:
                    key = f"{r['Product']} / {r['Tier']}"
                    s = prod_summary.setdefault(key, {
                        "Product": r["Product"], "Tier": r["Tier"],
                        "Servers": 0, "On Dashboard": 0,
                        "CPU Values": [], "Load Values": [],
                    })
                    s["Servers"] += 1
                    if r["On Dashboard"] == "Yes":
                        s["On Dashboard"] += 1
                    if r["CPU %"] is not None:
                        s["CPU Values"].append(r["CPU %"])
                    if r["Load Avg5"] is not None:
                        s["Load Values"].append(r["Load Avg5"])

                # Build provider summary
                prov_summary: dict[str, dict] = {}
                for r in rows:
                    prov = r["Provider"] or "No IP"
                    s = prov_summary.setdefault(prov, {"Servers": 0, "CPU Values": []})
                    s["Servers"] += 1
                    if r["CPU %"] is not None:
                        s["CPU Values"].append(r["CPU %"])

                # Generate Excel
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                wb = Workbook()

                                ws1 = wb.active
                ws1.title = "Servers"
                headers = ["Host", "Name", "Product", "Tier", "IP", "Provider",
                           "CPU %", "Load Avg5", "Mem Avail GB", "On Dashboard", "Dashboards", "Dashboard Tabs", "Groups"]

                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
                thin_border = Border(
                    bottom=Side(style="thin", color="D9D9D9"),
                )

                for col, h in enumerate(headers, 1):
                    cell = ws1.cell(row=1, column=col, value=h)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

                # Color coding for CPU
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

                for row_idx, r in enumerate(rows, 2):
                    for col, key in enumerate(headers, 1):
                        cell = ws1.cell(row=row_idx, column=col, value=r.get(key, ""))
                        cell.border = thin_border
                    # Color CPU column
                    cpu_cell = ws1.cell(row=row_idx, column=7)
                    if r["CPU %"] is not None:
                        if r["CPU %"] >= 80:
                            cpu_cell.fill = red_fill
                        elif r["CPU %"] >= 50:
                            cpu_cell.fill = yellow_fill
                        else:
                            cpu_cell.fill = green_fill

                # Auto-width
                for col in range(1, len(headers) + 1):
                    max_len = len(str(ws1.cell(row=1, column=col).value))
                    for row in range(2, min(len(rows) + 2, 50)):
                        val = ws1.cell(row=row, column=col).value
                        if val:
                            max_len = max(max_len, len(str(val)))
                    ws1.column_dimensions[ws1.cell(row=1, column=col).column_letter].width = min(max_len + 3, 40)

                ws1.auto_filter.ref = f"A1:{chr(64 + len(headers))}{len(rows) + 1}"
                ws1.freeze_panes = "A2"

                                ws2 = wb.create_sheet("Products")
                prod_headers = ["Product", "Tier", "Servers", "On Dashboard", "Median CPU %", "Median Load"]
                for col, h in enumerate(prod_headers, 1):
                    cell = ws2.cell(row=1, column=col, value=h)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

                for row_idx, (key, s) in enumerate(sorted(prod_summary.items()), 2):
                    ws2.cell(row=row_idx, column=1, value=s["Product"])
                    ws2.cell(row=row_idx, column=2, value=s["Tier"])
                    ws2.cell(row=row_idx, column=3, value=s["Servers"])
                    ws2.cell(row=row_idx, column=4, value=s["On Dashboard"])
                    ws2.cell(row=row_idx, column=5, value=round(median(s["CPU Values"]), 1) if s["CPU Values"] else "")
                    ws2.cell(row=row_idx, column=6, value=round(median(s["Load Values"]), 2) if s["Load Values"] else "")

                for col in range(1, len(prod_headers) + 1):
                    ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = 18
                ws2.freeze_panes = "A2"

                                ws3 = wb.create_sheet("Providers")
                prov_headers = ["Provider", "Servers", "Median CPU %"]
                for col, h in enumerate(prov_headers, 1):
                    cell = ws3.cell(row=1, column=col, value=h)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

                for row_idx, (prov, s) in enumerate(
                    sorted(prov_summary.items(), key=lambda x: -x[1]["Servers"]), 2
                ):
                    ws3.cell(row=row_idx, column=1, value=prov)
                    ws3.cell(row=row_idx, column=2, value=s["Servers"])
                    ws3.cell(row=row_idx, column=3, value=round(median(s["CPU Values"]), 1) if s["CPU Values"] else "")

                for col in range(1, len(prov_headers) + 1):
                    ws3.column_dimensions[ws3.cell(row=1, column=col).column_letter].width = 18
                ws3.freeze_panes = "A2"

                # Save
                if not output_dir:
                    output_dir = os.path.expanduser("~/Downloads")
                ts = datetime.now().strftime("%Y%m%d_%H%M")
                filename = f"zabbix_server_report_{ts}.xlsx"
                filepath = os.path.join(output_dir, filename)
                wb.save(filepath)

                # Build text summary
                total = len(rows)
                on_dash = sum(1 for r in rows if r["On Dashboard"] == "Yes")
                cpu_vals = [r["CPU %"] for r in rows if r["CPU %"] is not None]
                load_vals = [r["Load Avg5"] for r in rows if r["Load Avg5"] is not None]

                summary = [
                    f"**Server Report Generated**",
                    f"",
                    f"**File:** `{filepath}`",
                    f"**Servers:** {total} ({on_dash} on dashboards)",
                    f"**Median CPU:** {median(cpu_vals):.1f}%" if cpu_vals else "",
                    f"**Median Load:** {median(load_vals):.2f}" if load_vals else "",
                    f"**Products:** {len(prod_summary)}",
                    f"**Providers:** {len(prov_summary)}",
                    f"",
                    f"### Sheets",
                    f"1. **Servers** — {total} rows, 13 columns (filterable, color-coded CPU)",
                    f"2. **Products** — {len(prod_summary)} product/tier groups with median metrics",
                    f"3. **Providers** — {len(prov_summary)} providers with server counts",
                ]

                return "\n".join(summary)
            except (httpx.HTTPError, ValueError) as e:
                return f"Error generating report: {e}"
