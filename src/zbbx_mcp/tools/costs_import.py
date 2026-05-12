"""Cost ingestion tools — write {$COST_MONTH} macros from billing data.

Extracted from the former monolithic costs.py.
"""

import asyncio
import csv as _csv
import fnmatch
import json
import os
import re

import httpx

from zbbx_mcp.data import host_ip
from zbbx_mcp.resolver import InstanceResolver
from zbbx_mcp.tools.costs_common import (
    _IP_RE,
    COST_SRC_BILLING_COMPOUND,
    COST_SRC_BILLING_IP,
    COST_SRC_BILLING_NAME,
    COST_SRC_BILLING_TRANSLATED,
    COST_SRC_BULK_PATTERN,
    COST_SRC_CLUSTER_EXTRAS,
    COST_SRC_PRODUCT_MEDIAN,
    COST_SRC_PROVIDER_MEDIAN,
    _cluster_new_val,
    _dedup_name_from_ip_entries,
    _prefix_name_match,
    _provider_medians,
    _sanity_warnings,
)
from zbbx_mcp.utils import resolve_group_ids


def register(mcp, resolver: InstanceResolver, skip: set[str] = frozenset()) -> None:

    if "import_server_costs" not in skip:

        @mcp.tool()
        async def import_server_costs(
            costs_json: str,
            only_if_empty: bool = False,
            instance: str = "",
        ) -> str:
            """Bulk-set monthly costs on servers using {$COST_MONTH} host macros.

            Args:
                costs_json: JSON mapping hostname patterns to USD cost (e.g. {"srv-nl-*": 20})
                only_if_empty: Skip hosts that already have a non-zero {$COST_MONTH} set
                instance: Zabbix instance (optional)
            """
            try:
                cost_rules = json.loads(costs_json)
            except json.JSONDecodeError:
                return "Invalid JSON. Expected: {\"hostname-pattern\": cost, ...}"

            if not isinstance(cost_rules, dict):
                return "Expected a JSON object mapping hostname patterns to costs."

            try:
                client = resolver.resolve(instance)

                # Get all enabled hosts
                hosts = await client.call("host.get", {
                    "output": ["hostid", "host"],
                    "filter": {"status": "0"},
                })

                # Get existing {$COST_MONTH} macros
                existing = await client.call("usermacro.get", {
                    "hostids": [h["hostid"] for h in hosts],
                    "output": ["hostmacroid", "hostid", "value"],
                    "filter": {"macro": "{$COST_MONTH}"},
                })
                existing_map = {m["hostid"]: m for m in existing}

                # Match hosts to cost rules
                matched = 0
                created = 0
                updated = 0
                skipped = 0
                errors = []

                for h in hosts:
                    hostname = h["host"]
                    hid = h["hostid"]
                    cost = None

                    # Check rules in order (exact match first, then patterns)
                    for pattern, value in cost_rules.items():
                        if hostname == pattern or fnmatch.fnmatch(hostname, pattern):
                            cost = value
                            break

                    if cost is None:
                        continue

                    matched += 1
                    cost_str = str(cost)

                    try:
                        if hid in existing_map:
                            existing_val = (existing_map[hid].get("value") or "").strip()
                            if only_if_empty and existing_val not in ("", "0", "0.0", "0.00"):
                                skipped += 1
                                continue
                            # Update if value changed
                            if existing_map[hid]["value"] != cost_str:
                                await client.call("usermacro.update", {
                                    "hostmacroid": existing_map[hid]["hostmacroid"],
                                    "value": cost_str,
                                })
                                updated += 1
                            else:
                                skipped += 1
                        else:
                            # Create new macro
                            await client.call("usermacro.create", {
                                "hostid": hid,
                                "macro": "{$COST_MONTH}",
                                "value": cost_str,
                                "description": COST_SRC_BULK_PATTERN,
                            })
                            created += 1
                    except (httpx.HTTPError, ValueError) as e:
                        errors.append(f"{hostname}: {e}")
                        if len(errors) >= 10:
                            break

                parts = [
                    "**Cost import complete**",
                    "",
                    f"Matched: {matched} servers",
                    f"Created: {created} new macros",
                    f"Updated: {updated} existing macros",
                    f"Unchanged: {skipped}",
                    f"Unmatched: {len(hosts) - matched} servers (no matching pattern)",
                ]
                if errors:
                    parts.append(f"\nErrors ({len(errors)}):")
                    for e in errors:
                        parts.append(f"- {e}")

                return "\n".join(parts)
            except (httpx.HTTPError, ValueError) as e:
                return f"Error importing costs: {e}"


    if "set_bulk_cost" not in skip:

        @mcp.tool()
        async def set_bulk_cost(
            group: str,
            cost: float,
            instance: str = "",
        ) -> str:
            """Set monthly cost for all servers in a host group.

            Args:
                group: Zabbix host group name
                cost: Monthly cost in USD per server
                instance: Zabbix instance name (optional, for multi-instance setups)
            """
            try:
                client = resolver.resolve(instance)

                # Resolve group
                gids = await resolve_group_ids(client, group)
                if gids is None:
                    return f"Host group '{group}' not found."

                # Get hosts in group
                hosts = await client.call("host.get", {
                    "output": ["hostid", "host"],
                    "groupids": gids,
                    "filter": {"status": "0"},
                })
                if not hosts:
                    return f"No enabled hosts in group '{group}'."

                # Get existing macros
                hostids = [h["hostid"] for h in hosts]
                existing = await client.call("usermacro.get", {
                    "hostids": hostids,
                    "output": ["hostmacroid", "hostid", "value"],
                    "filter": {"macro": "{$COST_MONTH}"},
                })
                existing_map = {m["hostid"]: m for m in existing}

                cost_str = str(cost)
                created = 0
                updated = 0
                skipped = 0

                for h in hosts:
                    hid = h["hostid"]
                    if hid in existing_map:
                        if existing_map[hid]["value"] != cost_str:
                            await client.call("usermacro.update", {
                                "hostmacroid": existing_map[hid]["hostmacroid"],
                                "value": cost_str,
                            })
                            updated += 1
                        else:
                            skipped += 1
                    else:
                        await client.call("usermacro.create", {
                            "hostid": hid,
                            "macro": "{$COST_MONTH}",
                            "value": cost_str,
                            "description": "Monthly server cost in USD",
                        })
                        created += 1

                return (
                    f"Set ${cost}/month on {len(hosts)} servers in '{group}'. "
                    f"Created: {created}, Updated: {updated}, Unchanged: {skipped}."
                )
            except (httpx.HTTPError, ValueError) as e:
                return f"Error: {e}"


    if "import_costs_by_ip" not in skip:

        @mcp.tool()
        async def import_costs_by_ip(
            file_path: str = "",
            costs_json: str = "",
            dry_run: bool = True,
            min_cost: float = 1,
            max_cost: float = 5000,
            export_unmatched: str = "",
            only_if_empty: bool = False,
            name_match_strict: bool = False,
            instance: str = "",
        ) -> str:
            """Smart cost import: match by IP, then hostname (7-pass fuzzy).

            Name matching: exact → split-on-dash → IP-in-name → prefix → contains → billing name translation.

            Args:
                file_path: Path to JSON cost file (preferred over costs_json)
                costs_json: Inline JSON (alternative to file_path)
                dry_run: Preview only, don't write (default: True)
                min_cost: Skip prices below this (default: $1)
                max_cost: Skip prices above this (default: $5000)
                export_unmatched: File path to save unmatched entries as JSON (optional)
                only_if_empty: Skip hosts that already have a non-zero {$COST_MONTH} set
                name_match_strict: Disable the permissive prefix pass. Only exact,
                    dash-split, IP-in-name and billing-name-translation passes run.
                    Use for reconciliation against an authoritative source where a
                    wrong bind is costlier than a missed match.
                instance: Zabbix instance (optional)
            """
            # Load data
            try:
                if file_path:
                    with open(os.path.expanduser(file_path)) as f:
                        raw = json.load(f)
                elif costs_json:
                    raw = json.loads(costs_json)
                else:
                    return "Provide file_path or costs_json."
            except (json.JSONDecodeError, FileNotFoundError, OSError) as e:
                return f"Failed to load: {e}"

            if isinstance(raw, dict) and ("by_ip" in raw or "by_name" in raw):
                ip_costs = raw.get("by_ip", {})
                name_costs = raw.get("by_name", {})
            elif isinstance(raw, dict):
                ip_costs = raw
                name_costs = {}
            else:
                return "Expected JSON object with by_ip/by_name or flat IP→cost map."

            def _extract_price(v):
                """Extract numeric price from value (may be dict with price/cost field or number)."""
                if isinstance(v, dict):
                    for key in ("price", "cost", "price_monthly", "monthly"):
                        if key in v:
                            return v[key]
                    return None
                return v

            def _in_range(v):
                p = _extract_price(v)
                if p is None:
                    return False
                try:
                    return min_cost <= float(p) <= max_cost
                except (ValueError, TypeError):
                    return False

            safe_ips = {k: float(_extract_price(v)) for k, v in ip_costs.items() if _in_range(v)}
            safe_names = {k: float(_extract_price(v)) for k, v in name_costs.items() if _in_range(v)}

            # If IP values are dicts with 'name' field, derive a consistent
            # name→price map. Names whose entries disagree on price are
            # dropped from name-matching and reported instead.
            ip_derived_names, duplicated_names = _dedup_name_from_ip_entries(
                ip_costs, _in_range, _extract_price,
            )
            for name, price in ip_derived_names.items():
                if name not in safe_names:
                    safe_names[name] = price
            # If a duplicated name also exists in name_costs, pull it — we
            # cannot trust either price in isolation.
            for name in duplicated_names:
                safe_names.pop(name, None)
            total_input = len(ip_costs) + len(name_costs)
            skipped_range = total_input - len(safe_ips) - len(safe_names)

            try:
                client = resolver.resolve(instance)
                hosts = await client.call("host.get", {
                    "output": ["hostid", "host"],
                    "selectInterfaces": ["ip"],
                    "filter": {"status": "0"},
                })

                # Build lookup maps
                ip_to_host: dict[str, dict] = {}
                name_to_host: dict[str, dict] = {}
                name_list: list[str] = []
                # Compound hostnames: "base-cc1 cc3 cc5" monitors multiple
                # physical servers as one entity. Expand to component names so
                # each billing entry can be summed.
                compound_components: dict[str, list[str]] = {}
                for h in hosts:
                    ip = host_ip(h)
                    if ip:
                        ip_to_host[ip] = h
                    lname = h["host"].lower()
                    name_to_host[lname] = h
                    name_list.append(lname)
                    if " " in lname:
                        parts = lname.split()
                        m = re.match(r"^(.+?)([a-z]{2}\d+)$", parts[0])
                        if m:
                            base = m.group(1)
                            derivs = [parts[0]]
                            for p in parts[1:]:
                                if re.match(r"^[a-z]{2}\d+$", p):
                                    derivs.append(base + p)
                            if len(derivs) > 1:
                                compound_components[h["hostid"]] = derivs

                # --- Pass 1: by_ip (IP is source of truth) ---
                host_costs: dict[str, float] = {}  # hostid → cost
                host_source: dict[str, str] = {}   # hostid → match method
                compound_consumed: set[str] = set()  # billing names consumed by compound sum

                for ip, cost in safe_ips.items():
                    ip = ip.strip()
                    h = ip_to_host.get(ip)
                    if h:
                        hid = h["hostid"]
                        host_costs[hid] = max(host_costs.get(hid, 0), cost)
                        host_source[hid] = "ip"

                # --- Pass 1b: partial IP /24 prefix match (3-octet IPs) ---
                # Matches "x.y.z" against hosts with IP "x.y.z.*"
                # Runtime only — no IPs stored in code
                for ip_key, cost in safe_ips.items():
                    ip_key = ip_key.strip()
                    parts = ip_key.split(".")
                    if len(parts) == 3:
                        prefix = ip_key + "."
                        for full_ip, h in ip_to_host.items():
                            if full_ip.startswith(prefix):
                                hid = h["hostid"]
                                if hid not in host_costs or cost > host_costs[hid]:
                                    host_costs[hid] = cost
                                    host_source[hid] = "ip/24"

                # --- Pass 1c: compound hostnames (twin clusters) ---
                # For hosts like "parent-a a-b", sum billing entries for each
                # component name rather than taking max of one.
                for hid, comps in compound_components.items():
                    total = 0.0
                    hits = []
                    for comp in comps:
                        if comp in safe_names:
                            total += safe_names[comp]
                            hits.append(comp)
                    if hits:
                        host_costs[hid] = max(host_costs.get(hid, 0), total)
                        host_source[hid] = f"compound({len(hits)})"
                        compound_consumed.update(hits)

                # --- Pass 2-5: by_name (fill gaps, use max if IP already matched) ---
                unmatched_names: dict[str, float] = {}

                for name, cost in safe_names.items():
                    if name.lower().strip() in compound_consumed:
                        continue
                    name_lower = name.lower().strip()
                    matched_host = None

                    # Pass 2: exact lowercase match
                    if name_lower in name_to_host:
                        matched_host = name_to_host[name_lower]

                    # Pass 3: split on " - ", try last segment
                    if not matched_host and " - " in name:
                        segments = name.split(" - ")
                        for seg in reversed(segments):
                            seg_clean = seg.strip().lower()
                            if seg_clean in name_to_host:
                                matched_host = name_to_host[seg_clean]
                                break

                    # Pass 4: extract IP from name, match via interface
                    if not matched_host:
                        ip_match = _IP_RE.search(name)
                        if ip_match:
                            found_ip = ip_match.group(1)
                            if found_ip in ip_to_host:
                                matched_host = ip_to_host[found_ip]

                    # Pass 5: prefix match (name is start of a Zabbix hostname)
                    if not matched_host and not name_match_strict:
                        matched_host = _prefix_name_match(name_lower, name_list, name_to_host)

                    # Pass 6: billing name translation
                    # Billing often uses reversed naming: cc+num+suffix → suffix-cc+num
                    if not matched_host:
                        candidates = []
                        # Reverse: ccNNN-suffix -> suffix-cc0NNN
                        m = re.match(r"^([a-z]{2})(\d+)-(.+)$", name_lower)
                        if m:
                            cc, num, suffix = m.group(1), m.group(2), m.group(3)
                            candidates.extend([
                                f"{suffix}-{cc}{num.zfill(4)}",
                                f"{suffix}-{cc}{num}",
                            ])
                        # Prefix swap: pfx-ccNNN-suffix -> pfx-suffix-cc0NNN
                        m = re.match(r"^([a-z]{2,4})-([a-z]{2})(\d+)-(.+)$", name_lower)
                        if m:
                            pfx, cc, num, suffix = m.group(1), m.group(2), m.group(3), m.group(4)
                            candidates.append(f"{pfx}-{suffix}-{cc}{num.zfill(4)}")
                        # Configurable renames via env var
                        # ZABBIX_BILLING_RENAMES="old1:new1,old2:new2"
                        renames_raw = os.environ.get("ZABBIX_BILLING_RENAMES", "")
                        if renames_raw:
                            for pair in renames_raw.split(","):
                                if ":" in pair:
                                    old, new = pair.strip().split(":", 1)
                                    if old in name_lower:
                                        candidates.append(name_lower.replace(old, new))

                        for cand in candidates:
                            if cand in name_to_host:
                                matched_host = name_to_host[cand]
                                host_source[matched_host["hostid"]] = "translated"
                                break

                    if matched_host:
                        hid = matched_host["hostid"]
                        # Prefer earliest pass. An IP-backed or compound-sum
                        # match (passes 1/1b/1c) is always stronger than a
                        # name-match — do not overwrite it.
                        if hid in host_costs:
                            continue
                        host_costs[hid] = cost
                        if hid not in host_source:
                            host_source[hid] = "name"
                    else:
                        unmatched_names[name] = cost

                # Filter out hosts with an existing non-zero cost if only_if_empty
                skipped_existing = 0
                if only_if_empty and host_costs:
                    existing_pre = await client.call("usermacro.get", {
                        "hostids": list(host_costs.keys()),
                        "output": ["hostid", "value"],
                        "filter": {"macro": "{$COST_MONTH}"},
                    })
                    nonempty = {
                        m["hostid"] for m in existing_pre
                        if (m.get("value") or "").strip() not in ("", "0", "0.0", "0.00")
                    }
                    skipped_existing = len(nonempty)
                    for hid in nonempty:
                        host_costs.pop(hid, None)
                        host_source.pop(hid, None)

                # Build match list
                host_map = {h["hostid"]: h for h in hosts}
                matches = []
                for hid, cost in host_costs.items():
                    h = host_map.get(hid)
                    if h:
                        ip = host_ip(h)
                        matches.append((h["host"], hid, ip or "", cost, host_source.get(hid, "?")))

                ip_matches = sum(1 for s in host_source.values() if s == "ip")
                ip24_matches = sum(1 for s in host_source.values() if s == "ip/24")
                name_matches = sum(1 for s in host_source.values() if s == "name")
                translated_matches = sum(1 for s in host_source.values() if s == "translated")
                compound_matches = sum(1 for s in host_source.values() if s.startswith("compound"))

                # Export unmatched
                if export_unmatched and unmatched_names:
                    path = os.path.expanduser(export_unmatched)
                    with open(path, "w") as f:
                        json.dump(unmatched_names, f, indent=2, ensure_ascii=False)

                if dry_run:
                    total_cost = sum(c for _, _, _, c, _ in matches)
                    medians = await _provider_medians(client)
                    warnings = _sanity_warnings(matches, ip_to_host, medians)
                    lines = [
                        f"**DRY RUN — {len(matches)} hosts matched**",
                        f"By IP: {ip_matches} | By /24: {ip24_matches} | By name: {name_matches} | Translated: {translated_matches} | Compound: {compound_matches}",
                        f"Unmatched: {len(unmatched_names)} name entries",
                        f"Skipped: {skipped_range} outside ${min_cost}-${max_cost}",
                    ]
                    if only_if_empty:
                        lines.append(f"Skipped (already costed): {skipped_existing}")
                    lines.append(f"**Total: ${total_cost:,.2f}/mo** (${total_cost*12:,.2f}/yr)\n")
                    lines.append("| Host | Match | $/mo |")
                    lines.append("|------|-------|------|")
                    for hostname, _hid, _ip, cost, src in sorted(matches, key=lambda x: -x[3])[:25]:
                        lines.append(f"| {hostname} | {src} | ${cost:.2f} |")
                    if len(matches) > 25:
                        lines.append(f"| ... | +{len(matches) - 25} more | |")
                    if warnings:
                        lines.append(f"\n**⚠ Sanity check: {len(warnings)} prices far from provider median**")
                        for w in warnings[:10]:
                            lines.append(f"- {w}")
                        if len(warnings) > 10:
                            lines.append(f"- ...+{len(warnings) - 10} more")
                    if duplicated_names:
                        lines.append(f"\n**⚠ Duplicate-name entries (dropped from name-match): {len(duplicated_names)}**")
                        for name, prices in sorted(duplicated_names.items())[:10]:
                            prices_str = ", ".join(f"${p:.2f}" for p in prices)
                            lines.append(f"- `{name}`: {prices_str}")
                        if len(duplicated_names) > 10:
                            lines.append(f"- ...+{len(duplicated_names) - 10} more")
                    if export_unmatched and unmatched_names:
                        lines.append(f"\nUnmatched exported to `{export_unmatched}`")
                    lines.append("\nSet `dry_run=false` to apply.")
                    return "\n".join(lines)

                # Apply: set {$COST_MONTH} macros
                existing = await client.call("usermacro.get", {
                    "hostids": [m[1] for m in matches],
                    "output": ["hostmacroid", "hostid", "value"],
                    "filter": {"macro": "{$COST_MONTH}"},
                })
                existing_map = {m["hostid"]: m for m in existing}

                def _src_tag(src: str) -> str:
                    if src == "ip" or src == "ip/24":
                        return COST_SRC_BILLING_IP
                    if src == "translated":
                        return COST_SRC_BILLING_TRANSLATED
                    if src.startswith("compound"):
                        return COST_SRC_BILLING_COMPOUND
                    return COST_SRC_BILLING_NAME

                created = updated = unchanged = 0
                errors = []
                for hostname, hid, _ip, cost, src in matches:
                    cost_str = str(round(cost, 2))
                    desc = _src_tag(src)
                    try:
                        if hid in existing_map:
                            if existing_map[hid]["value"] != cost_str:
                                await client.call("usermacro.update", {
                                    "hostmacroid": existing_map[hid]["hostmacroid"],
                                    "value": cost_str,
                                    "description": desc,
                                })
                                updated += 1
                            else:
                                unchanged += 1
                        else:
                            await client.call("usermacro.create", {
                                "hostid": hid,
                                "macro": "{$COST_MONTH}",
                                "value": cost_str,
                                "description": desc,
                            })
                            created += 1
                    except (httpx.HTTPError, ValueError) as e:
                        errors.append(f"{hostname}: {e}")
                        if len(errors) >= 10:
                            break

                total_cost = sum(c for _, _, _, c, _ in matches)
                parts = [
                    f"**Cost import complete — {len(matches)} servers**",
                    f"By IP: {ip_matches} | By name: {name_matches}",
                    f"Created: {created} | Updated: {updated} | Unchanged: {unchanged}",
                    f"Total: ${total_cost:,.2f}/mo (${total_cost*12:,.2f}/yr)",
                ]
                if unmatched_names:
                    parts.append(f"Unmatched: {len(unmatched_names)} name entries")
                if export_unmatched and unmatched_names:
                    parts.append(f"Exported to `{export_unmatched}`")
                if errors:
                    parts.append(f"Errors: {len(errors)}")
                return "\n".join(parts)
            except (httpx.HTTPError, ValueError) as e:
                return f"Error: {e}"


    if "import_cluster_ip_fees" not in skip:

        @mcp.tool()
        async def import_cluster_ip_fees(
            file_path: str = "",
            fees_json: str = "",
            dry_run: bool = True,
            overwrite_base: float = -1.0,
            instance: str = "",
        ) -> str:
            """Add extra-IP billing fees to cluster primary {$COST_MONTH} macros.

            Idempotent against prior cluster_extras runs. Input shape:
            list of {"cluster", "extra_ips", "extra_cost_month"}. See ADR 009.

            Args:
                file_path: Path to JSON file with cluster fee entries (optional)
                fees_json: Inline JSON (use this OR file_path)
                dry_run: Preview changes without applying (default: True)
                overwrite_base: If ≥ 0, reset the cluster's base to this value
                instance: Zabbix instance (optional)
            """
            try:
                client = resolver.resolve(instance)

                # Parse input
                if file_path:
                    with open(os.path.expanduser(file_path)) as f:
                        data = json.load(f)
                elif fees_json:
                    data = json.loads(fees_json)
                else:
                    return "Provide either file_path or fees_json."

                if not isinstance(data, list):
                    return "Expected JSON array of {cluster, extra_ips, extra_cost_month}."

                # Batch lookup: all cluster names in one host.get call
                host_names = [e.get("cluster", "") for e in data if e.get("cluster")]
                if not host_names:
                    return "No cluster names in input."

                hosts = await client.call("host.get", {
                    "output": ["hostid", "host"],
                    "filter": {"host": host_names},
                    "selectMacros": ["hostmacroid", "macro", "value", "description"],
                })
                host_by_name = {h["host"]: h for h in hosts}

                # Build update plan
                updates = []
                missing = []
                for entry in data:
                    name = entry.get("cluster", "")
                    extras = float(entry.get("extra_cost_month", 0))
                    ip_count = len(entry.get("extra_ips", []))
                    # Allow extras==0: the tool then rewrites the description
                    # at the current (or overwrite_base) base without changing
                    # the summed value. Useful for re-tagging legacy macros
                    # into the `src:cluster_extras` provenance format.
                    if not name or extras < 0:
                        continue
                    h = host_by_name.get(name)
                    if not h:
                        missing.append(name)
                        continue

                    existing_macro = None
                    current = 0.0
                    existing_desc = ""
                    for m in h.get("macros", []):
                        if m["macro"] == "{$COST_MONTH}":
                            existing_macro = m
                            existing_desc = m.get("description", "") or ""
                            try:
                                current = float(m["value"])
                            except (ValueError, TypeError):
                                current = 0.0
                            break

                    base, new_val = _cluster_new_val(current, existing_desc, extras, overwrite_base)
                    desc = f"{COST_SRC_CLUSTER_EXTRAS} base {base:.2f} + {ip_count} extra IP{'s' if ip_count > 1 else ''} ({extras:.2f})"
                    updates.append({
                        "hostid": h["hostid"],
                        "host": name,
                        "macroid": existing_macro["hostmacroid"] if existing_macro else None,
                        "current": current,
                        "base": base,
                        "extras": extras,
                        "new": new_val,
                        "desc": desc,
                    })

                # Build response
                total_delta = sum(u["extras"] for u in updates)
                header = f"**Cluster IP fees import** — {len(updates)} clusters, ${total_delta:,.2f}/mo added"
                if missing:
                    header += f" ({len(missing)} missing)"

                lines = [header + ("  DRY RUN" if dry_run else ""), ""]
                lines.append("| Cluster | Current | Base | +Extras | New | IPs |")
                lines.append("|---------|---------|------|---------|-----|-----|")
                for u in sorted(updates, key=lambda x: -x["extras"])[:25]:
                    ip_count = u["desc"].split()[3]
                    lines.append(
                        f"| {u['host']} | ${u['current']:.2f} | ${u['base']:.2f} | "
                        f"${u['extras']:.2f} | ${u['new']:.2f} | {ip_count} |"
                    )
                if len(updates) > 25:
                    lines.append(f"*+{len(updates) - 25} more clusters*")
                if missing:
                    lines.append(f"\n**Missing (not found in Zabbix):** {', '.join(missing[:10])}")

                if dry_run:
                    lines.append("\n*Set dry_run=false to apply.*")
                    return "\n".join(lines)

                # Apply — parallel usermacro update/create
                async def _apply(u):
                    if u["macroid"]:
                        return await client.call("usermacro.update", {
                            "hostmacroid": u["macroid"],
                            "value": str(u["new"]),
                            "description": u["desc"],
                        })
                    return await client.call("usermacro.create", {
                        "hostid": u["hostid"],
                        "macro": "{$COST_MONTH}",
                        "value": str(u["new"]),
                        "description": u["desc"],
                    })

                results = await asyncio.gather(*[_apply(u) for u in updates], return_exceptions=True)
                errors = [str(r)[:80] for r in results if isinstance(r, Exception)]
                applied = sum(1 for r in results if not isinstance(r, Exception))

                lines.append(f"\n**Applied: {applied}/{len(updates)}**")
                if errors:
                    lines.append(f"Errors: {len(errors)}")
                    for e in errors[:3]:
                        lines.append(f"  - {e}")
                return "\n".join(lines)
            except (httpx.HTTPError, ValueError, OSError) as e:
                return f"Error: {e}"


    if "import_from_xlsx" not in skip:

        @mcp.tool()
        async def import_from_xlsx(
            file_path: str,
            output_csv: str = "/tmp/billing_xlsx.csv",
            eur_usd: float = 1.08,
        ) -> str:
            """Parse a billing XLSX into a flat CSV (ip, billing_name, price USD).

            Handles the two sheet shapes used by the current accounting
            workbook. See ADR 002, 004 for layout details.

            Args:
                file_path: Path to the XLSX file
                output_csv: Where to write the flat CSV (default /tmp/billing_xlsx.csv)
                eur_usd: EUR→USD conversion rate for EUR-priced sheets (default 1.08)
            """
            try:
                import openpyxl
            except ImportError:
                return "openpyxl not installed. Run: uv pip install openpyxl"

            try:
                path = os.path.expanduser(file_path)
                if not os.path.isfile(path):
                    return f"File not found: {path}"
                wb = openpyxl.load_workbook(path, data_only=True)
            except (OSError, ValueError) as e:
                return f"Failed to open XLSX: {e}"

            rows: list[dict] = []
            ip_re = _IP_RE

            def _pick_price(values, max_idx=17):
                """Return first positive price in cols 13–16 (server/Price$/Addons)."""
                p1 = values[13] if len(values) > 13 else None
                p2 = values[14] if len(values) > 14 else None
                p_add = values[16] if len(values) > 16 else None
                price = None
                for p in (p1, p2):
                    if isinstance(p, (int, float)) and 0 < p < 5000:
                        price = float(p)
                        break
                if price is None:
                    return None
                if isinstance(p_add, (int, float)) and p_add > 0:
                    price += float(p_add)
                return price

            # Structured detailed sheets (Ip-price style)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                headers = [
                    str(c or "").strip().lower()
                    for c in next(ws.iter_rows(values_only=True), ())
                ]
                # Heuristic: detailed sheet has an "ip server" and "price server" header.
                # Localised workbooks can supply an additional header pattern via
                # ZABBIX_BILLING_IP_HEADER (case-insensitive substring match).
                extra_ip_header = os.environ.get("ZABBIX_BILLING_IP_HEADER", "").strip().lower()
                has_ip_col = any(
                    "ip server" in h or (extra_ip_header and extra_ip_header in h)
                    for h in headers
                )
                has_price_col = any("price server" in h or h.startswith("price") for h in headers)
                if not (has_ip_col and has_price_col):
                    continue
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i < 1:
                        continue
                    name = row[11] if len(row) > 11 else None
                    ip_cell = row[12] if len(row) > 12 else None
                    if not ip_cell:
                        continue
                    ips = ip_re.findall(str(ip_cell))
                    if not ips:
                        continue
                    price = _pick_price(row)
                    if price is None:
                        continue
                    for j, ip in enumerate(ips):
                        rows.append({
                            "ip": ip,
                            "name": str(name or "").strip(),
                            "price": price if j == 0 else 0.0,
                        })

            # Simple sheets (Sheet11 style): Name | NS | IPv4 | Price
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                first = next(ws.iter_rows(values_only=True), ())
                headers = [str(c or "").strip().lower() for c in first]
                if len(headers) < 4:
                    continue
                if not (headers[0] in ("name", "hostname") and
                        headers[2] in ("ipv4", "ip", "ip address") and
                        headers[3].startswith("price")):
                    continue
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i < 1:
                        continue
                    name, _ns, ip, pr = row[:4]
                    if not ip:
                        continue
                    if isinstance(pr, str):
                        try:
                            price = float(pr.replace(",", "."))
                        except ValueError:
                            continue
                    elif isinstance(pr, (int, float)):
                        price = float(pr)
                    else:
                        continue
                    # Sheet11 prices are in EUR — convert.
                    rows.append({
                        "ip": str(ip).strip(),
                        "name": str(name or "").strip(),
                        "price": price * eur_usd,
                    })

            # Dedupe by ip, keep max price
            by_ip: dict[str, dict] = {}
            for r in rows:
                if r["ip"] in by_ip and r["price"] <= by_ip[r["ip"]]["price"]:
                    continue
                by_ip[r["ip"]] = r

            out_path = os.path.expanduser(output_csv)
            with open(out_path, "w", newline="") as f:
                w = _csv.writer(f)
                w.writerow(["ip", "billing_name", "price_monthly"])
                for r in by_ip.values():
                    w.writerow([r["ip"], r["name"], f"{r['price']:.2f}"])

            total = sum(r["price"] for r in by_ip.values())
            paid = sum(1 for r in by_ip.values() if r["price"] > 0)
            return (
                f"Parsed {len(by_ip)} unique IPs from {len(wb.sheetnames)} sheets\n"
                f"Priced: {paid} | Extras (price=0): {len(by_ip) - paid}\n"
                f"Total: ${total:,.2f}/mo\n"
                f"Wrote: {out_path}"
            )


    if "fill_cost_median" not in skip:

        @mcp.tool()
        async def fill_cost_median(
            group_by: str = "product",
            dry_run: bool = True,
            instance: str = "",
        ) -> str:
            """Estimate {$COST_MONTH} for empty-cost hosts using peer median.

            For each host without a non-zero {$COST_MONTH}, take the median of
            costed hosts sharing the same grouping key and assign it. Skips
            hosts without an IP (monitoring) so they stay out of cost totals.

            Args:
                group_by: "product" (product/tier from host groups) or "provider" (CIDR detection)
                dry_run: Preview only (default: True)
                instance: Zabbix instance (optional)
            """
            import statistics

            from zbbx_mcp.classify import classify_host as _classify_host
            from zbbx_mcp.classify import detect_provider

            try:
                client = resolver.resolve(instance)
                hosts, macros = await asyncio.gather(
                    client.call("host.get", {
                        "output": ["hostid", "host"],
                        "selectGroups": ["name"],
                        "selectInterfaces": ["ip"],
                        "filter": {"status": "0"},
                    }),
                    client.call("usermacro.get", {
                        "output": ["hostmacroid", "hostid", "value"],
                        "filter": {"macro": "{$COST_MONTH}"},
                    }),
                )

                cost_by_hid: dict[str, float] = {}
                macro_by_hid: dict[str, str] = {}
                for m in macros:
                    macro_by_hid[m["hostid"]] = m["hostmacroid"]
                    try:
                        v = float(m.get("value") or 0)
                        if v > 0:
                            cost_by_hid[m["hostid"]] = v
                    except (ValueError, TypeError):
                        pass

                def _key(h):
                    if group_by == "provider":
                        ip = host_ip(h)
                        return detect_provider(ip) if ip else None
                    prod, tier = _classify_host(h.get("groups", []))
                    return f"{prod}/{tier}"

                bucket: dict[str, list[float]] = {}
                for h in hosts:
                    if h["hostid"] in cost_by_hid:
                        k = _key(h)
                        if k:
                            bucket.setdefault(k, []).append(cost_by_hid[h["hostid"]])

                medians = {k: statistics.median(v) for k, v in bucket.items() if v}

                candidates = []
                skipped_no_ip = 0
                skipped_no_peer = 0
                for h in hosts:
                    if h["hostid"] in cost_by_hid:
                        continue
                    ip = host_ip(h)
                    if not ip:
                        skipped_no_ip += 1
                        continue
                    k = _key(h)
                    if not k or k not in medians:
                        skipped_no_peer += 1
                        continue
                    candidates.append((h, k, medians[k]))

                total_delta = sum(c for _, _, c in candidates)
                if dry_run:
                    lines = [
                        f"**Fill cost by {group_by} median — DRY RUN**",
                        f"Candidates: {len(candidates)} hosts, ${total_delta:,.2f}/mo",
                        f"Skipped: {skipped_no_ip} no-IP, {skipped_no_peer} no peer",
                        "",
                        f"| Host | {group_by.title()} | Median $/mo |",
                        "|------|---------|-------------|",
                    ]
                    by_group: dict[str, int] = {}
                    for _h, k, _c in candidates:
                        by_group[k] = by_group.get(k, 0) + 1
                    for _h, k, c in sorted(candidates, key=lambda x: -x[2])[:25]:
                        lines.append(f"| {_h['host']} | {k} | ${c:.2f} |")
                    if len(candidates) > 25:
                        lines.append(f"*+{len(candidates) - 25} more*")
                    lines.append("")
                    lines.append("**By group:**")
                    for k, n in sorted(by_group.items(), key=lambda x: -x[1]):
                        lines.append(f"- {k}: {n} hosts @ ${medians[k]:.2f}")
                    lines.append("\nSet dry_run=false to apply.")
                    return "\n".join(lines)

                created = 0
                errors = []
                for h, _k, cost in candidates:
                    val = str(round(cost, 2))
                    try:
                        if h["hostid"] in macro_by_hid:
                            await client.call("usermacro.update", {
                                "hostmacroid": macro_by_hid[h["hostid"]],
                                "value": val,
                                "description": COST_SRC_PRODUCT_MEDIAN if group_by == "product" else COST_SRC_PROVIDER_MEDIAN,
                            })
                        else:
                            await client.call("usermacro.create", {
                                "hostid": h["hostid"],
                                "macro": "{$COST_MONTH}",
                                "value": val,
                                "description": COST_SRC_PRODUCT_MEDIAN if group_by == "product" else COST_SRC_PROVIDER_MEDIAN,
                            })
                        created += 1
                    except (httpx.HTTPError, ValueError) as e:
                        errors.append(f"{h['host']}: {e}")
                        if len(errors) >= 10:
                            break
                parts = [
                    f"**Filled {created}/{len(candidates)} empty-cost hosts with {group_by} median**",
                    f"Added: ${total_delta:,.2f}/mo",
                ]
                if errors:
                    parts.append(f"Errors: {len(errors)}")
                return "\n".join(parts)
            except (httpx.HTTPError, ValueError) as e:
                return f"Error: {e}"
