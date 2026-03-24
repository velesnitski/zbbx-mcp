"""Infrastructure cost and optimization report.

Generates a multi-sheet Excel report matching the "Infrastructure Costs" format:
- Sheet 1: Apps & Infra (servers with specs, costs, provider, product)
- Sheet 2: Unused/Underloaded (decommission candidates with savings)
- Sheet 3: Provider Summary (costs, server counts, decomm candidates)
"""

import asyncio
import os
from datetime import datetime, timezone
from statistics import median

import httpx

from zbbx_mcp.resolver import InstanceResolver
from zbbx_mcp.classify import classify_host as _classify_host, detect_provider


def register(mcp, resolver: InstanceResolver, skip: set[str] = frozenset()):

    if "generate_infra_report" not in skip:

        @mcp.tool()
        async def generate_infra_report(
            cpu_idle_threshold: float = 90.0,
            output_dir: str = "",
            instance: str = "",
        ) -> str:
            """Generate an infrastructure optimization report (Excel).

            Analyzes all servers from Zabbix: classifies by product/provider,
            detects underloaded/idle servers, and builds a cost optimization report.

            Sheets:
            1. Apps & Infra — full server inventory with specs and metrics
            2. Unused/Underloaded — decommission candidates (CPU idle > threshold)
            3. Provider Summary — aggregated by hosting provider

            Args:
                cpu_idle_threshold: CPU idle % above which a server is flagged as underloaded (default: 90%, meaning <10% CPU used)
                output_dir: Directory for the Excel file (default: ~/Downloads)
                instance: Zabbix instance name (optional, for multi-instance setups)
            """
            try:
                client = resolver.resolve(instance)

                # Phase 1: hosts + dashboards (parallel)
                hosts_task = client.call("host.get", {
                    "output": ["hostid", "host", "name", "status"],
                    "selectGroups": ["name"],
                    "selectInterfaces": ["ip", "type", "port"],
                    "filter": {"status": "0"},
                })
                dashboards_task = client.call("dashboard.get", {
                    "output": ["dashboardid", "name"],
                    "selectPages": "extend",
                })
                hosts, dashboards = await asyncio.gather(hosts_task, dashboards_task)

                # Extract graph IDs
                graph_ids = set()
                graph_context = {}
                for d in dashboards:
                    dname = d["name"]
                    for pi, page in enumerate(d.get("pages", [])):
                        tab = page.get("name", "") or f"Page {pi + 1}"
                        for w in page.get("widgets", []):
                            for f in w.get("fields", []):
                                if f.get("type") == "6":
                                    gid = f["value"]
                                    graph_ids.add(gid)
                                    graph_context[gid] = {"dashboard": dname, "tab": tab}

                host_map = {h["hostid"]: h for h in hosts}
                all_ids = list(host_map.keys())

                # Phase 2: metrics + graphs + cost macros (parallel)
                tasks = [
                    client.call("item.get", {
                        "hostids": all_ids,
                        "output": ["hostid", "lastvalue"],
                        "filter": {"key_": "system.cpu.util[,idle]"},
                    }),
                    client.call("item.get", {
                        "hostids": all_ids,
                        "output": ["hostid", "lastvalue"],
                        "filter": {"key_": "system.cpu.load[percpu,avg5]"},
                    }),
                    client.call("item.get", {
                        "hostids": all_ids,
                        "output": ["hostid", "lastvalue"],
                        "filter": {"key_": "vm.memory.size[available]"},
                    }),
                    client.call("item.get", {
                        "hostids": all_ids,
                        "output": ["hostid", "lastvalue"],
                        "filter": {"key_": "vm.memory.size[total]"},
                    }),
                    # Fetch {$COST_MONTH} macro from all hosts
                    client.call("usermacro.get", {
                        "hostids": all_ids,
                        "output": ["hostid", "hostmacroid", "macro", "value"],
                        "filter": {"macro": "{$COST_MONTH}"},
                    }),
                ]
                if graph_ids:
                    tasks.append(client.call("graph.get", {
                        "graphids": list(graph_ids),
                        "output": ["graphid"],
                        "selectHosts": ["hostid"],
                    }))

                results = await asyncio.gather(*tasks)
                cpu_items, load_items, mem_avail_items, mem_total_items = results[:4]
                cost_macros = results[4]
                graphs = results[5] if len(results) > 5 else []

                # Build cost map from macros
                cost_map: dict[str, float] = {}
                for m in cost_macros:
                    try:
                        cost_map[m["hostid"]] = float(m.get("value", "0"))
                    except (ValueError, TypeError):
                        pass

                # Build dashboard mapping
                graph_to_host = {}
                for g in graphs:
                    for h in g.get("hosts", []):
                        graph_to_host[g["graphid"]] = h["hostid"]

                host_dashboard: dict[str, str] = {}
                host_tab: dict[str, str] = {}
                for gid, ctx in graph_context.items():
                    hid = graph_to_host.get(gid)
                    if hid:
                        host_dashboard[hid] = ctx["dashboard"]
                        host_tab[hid] = f"{ctx['dashboard']} / {ctx['tab']}"

                # Build metrics
                cpu_map, load_map, mem_avail_map, mem_total_map = {}, {}, {}, {}
                for i in cpu_items:
                    try:
                        cpu_map[i["hostid"]] = round(float(i["lastvalue"]), 1)
                    except (ValueError, TypeError):
                        pass
                for i in load_items:
                    try:
                        load_map[i["hostid"]] = round(float(i["lastvalue"]), 2)
                    except (ValueError, TypeError):
                        pass
                for i in mem_avail_items:
                    try:
                        mem_avail_map[i["hostid"]] = round(float(i["lastvalue"]) / 1_073_741_824, 1)
                    except (ValueError, TypeError):
                        pass
                for i in mem_total_items:
                    try:
                        mem_total_map[i["hostid"]] = round(float(i["lastvalue"]) / 1_073_741_824, 1)
                    except (ValueError, TypeError):
                        pass

                # Build rows
                rows = []
                unused_rows = []
                for hid, h in host_map.items():
                    prod, tier = _classify_host(h.get("groups", []))
                    if not prod or prod == "Unknown":
                        continue

                    ip = ""
                    for iface in h.get("interfaces", []):
                        if iface.get("ip") and iface["ip"] != "127.0.0.1":
                            ip = iface["ip"]
                            break

                    provider = detect_provider(ip) if ip else ""
                    cpu_idle = cpu_map.get(hid)
                    cpu_used = round(100 - cpu_idle, 1) if cpu_idle is not None else None
                    load_avg = load_map.get(hid)
                    mem_total = mem_total_map.get(hid)
                    mem_avail = mem_avail_map.get(hid)
                    groups = ", ".join(g["name"] for g in h.get("groups", []))
                    dashboard = host_dashboard.get(hid, "")
                    tab = host_tab.get(hid, "")

                    cost_month = cost_map.get(hid)
                    cost_year = round(cost_month * 12, 2) if cost_month else None

                    row = {
                        "Host": h.get("host", ""),
                        "Name": h.get("name", ""),
                        "Product": prod,
                        "Tier": tier,
                        "Provider": provider,
                        "IP": ip,
                        "RAM Total GB": mem_total,
                        "RAM Avail GB": mem_avail,
                        "CPU Used %": cpu_used,
                        "Load Avg5": load_avg,
                        "Cost/Month ($)": cost_month,
                        "Cost/Year ($)": cost_year,
                        "Dashboard": dashboard,
                        "Dashboard Tab": tab,
                        "Groups": groups,
                        "Status": "Active",
                    }
                    rows.append(row)

                    # Flag underloaded
                    if cpu_idle is not None and cpu_idle >= cpu_idle_threshold:
                        unused_rows.append({
                            "Resource": h.get("host", ""),
                            "Type": "Server",
                            "Host/Domain": ip,
                            "Provider": provider,
                            "Product": prod,
                            "Tier": tier,
                            "CPU Used %": cpu_used,
                            "Load Avg5": load_avg,
                            "Cost/Month ($)": cost_month,
                            "Reason": f"CPU idle {cpu_idle:.0f}% (used only {cpu_used:.1f}%)",
                            "Recommendation": "Review for decommission",
                            "Priority": "High" if cpu_used < 3 else "Medium",
                            "Dashboard": dashboard,
                        })

                rows.sort(key=lambda r: (r["Product"], r["Tier"], r["Host"]))
                unused_rows.sort(key=lambda r: (r.get("CPU Used %") or 100))

                # Provider summary
                prov_data: dict[str, dict] = {}
                for r in rows:
                    prov = r["Provider"] or "No IP"
                    p = prov_data.setdefault(prov, {
                        "servers": 0, "cpu_vals": [], "decomm": 0,
                        "total_cost": 0.0, "savings": 0.0,
                    })
                    p["servers"] += 1
                    if r["CPU Used %"] is not None:
                        p["cpu_vals"].append(r["CPU Used %"])
                    if r["Cost/Month ($)"] is not None:
                        p["total_cost"] += r["Cost/Month ($)"]

                for r in unused_rows:
                    prov = r["Provider"] or "No IP"
                    if prov in prov_data:
                        prov_data[prov]["decomm"] += 1
                        if r.get("Cost/Month ($)"):
                            prov_data[prov]["savings"] += r["Cost/Month ($)"]

                # Generate Excel
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                wb = Workbook()
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                thin_border = Border(bottom=Side(style="thin", color="D9D9D9"))

                def _write_headers(ws, headers):
                    for col, h in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=h)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")

                def _auto_width(ws, headers, max_rows=50):
                    for col in range(1, len(headers) + 1):
                        max_len = len(str(ws.cell(1, col).value or ""))
                        for row in range(2, min(ws.max_row + 1, max_rows)):
                            val = ws.cell(row, col).value
                            if val:
                                max_len = max(max_len, len(str(val)))
                        ws.column_dimensions[ws.cell(1, col).column_letter].width = min(max_len + 3, 40)

                                ws1 = wb.active
                ws1.title = "Apps & Infra"
                h1 = ["#", "Host", "Name", "Product", "Tier", "Provider", "IP",
                       "RAM Total GB", "RAM Avail GB", "CPU Used %", "Load Avg5",
                       "Cost/Month ($)", "Cost/Year ($)",
                       "Dashboard", "Dashboard Tab", "Groups", "Status"]
                _write_headers(ws1, h1)

                cpu_col = h1.index("CPU Used %") + 1
                for idx, r in enumerate(rows, 2):
                    ws1.cell(row=idx, column=1, value=idx - 1)
                    for col, key in enumerate(h1[1:], 2):
                        cell = ws1.cell(row=idx, column=col, value=r.get(key, ""))
                        cell.border = thin_border
                    # Color CPU
                    cpu_cell = ws1.cell(row=idx, column=cpu_col)
                    if r["CPU Used %"] is not None:
                        if r["CPU Used %"] >= 80:
                            cpu_cell.fill = red_fill
                        elif r["CPU Used %"] >= 50:
                            cpu_cell.fill = yellow_fill
                        elif r["CPU Used %"] < 10:
                            cpu_cell.fill = green_fill

                ws1.auto_filter.ref = f"A1:{chr(64 + len(h1))}{len(rows) + 1}"
                ws1.freeze_panes = "A2"
                _auto_width(ws1, h1)

                                ws2 = wb.create_sheet("Unused & Underloaded")
                h2 = ["#", "Resource", "Type", "Host/Domain", "Provider", "Product",
                       "Tier", "CPU Used %", "Load Avg5", "Cost/Month ($)",
                       "Reason", "Recommendation", "Priority", "Dashboard"]
                _write_headers(ws2, h2)

                for idx, r in enumerate(unused_rows, 2):
                    ws2.cell(row=idx, column=1, value=idx - 1)
                    for col, key in enumerate(h2[1:], 2):
                        cell = ws2.cell(row=idx, column=col, value=r.get(key, ""))
                        cell.border = thin_border
                    # Highlight priority
                    prio_cell = ws2.cell(row=idx, column=h2.index("Priority") + 1)
                    if r.get("Priority") == "High":
                        prio_cell.fill = red_fill
                    elif r.get("Priority") == "Medium":
                        prio_cell.fill = yellow_fill

                ws2.auto_filter.ref = f"A1:{chr(64 + len(h2))}{len(unused_rows) + 1}"
                ws2.freeze_panes = "A2"
                _auto_width(ws2, h2)

                                ws3 = wb.create_sheet("Provider Summary")
                h3 = ["Provider", "Servers", "Cost/Month ($)", "Cost/Year ($)",
                       "Decomm Candidates", "Potential Savings/Month ($)",
                       "Median CPU %", "Servers < 10% CPU"]
                _write_headers(ws3, h3)

                sorted_provs = sorted(prov_data.items(), key=lambda x: -x[1]["total_cost"])
                for idx, (prov, p) in enumerate(sorted_provs, 2):
                    ws3.cell(row=idx, column=1, value=prov)
                    ws3.cell(row=idx, column=2, value=p["servers"])
                    ws3.cell(row=idx, column=3, value=round(p["total_cost"], 2) if p["total_cost"] else "")
                    ws3.cell(row=idx, column=4, value=round(p["total_cost"] * 12, 2) if p["total_cost"] else "")
                    ws3.cell(row=idx, column=5, value=p["decomm"])
                    ws3.cell(row=idx, column=6, value=round(p["savings"], 2) if p["savings"] else "")
                    ws3.cell(row=idx, column=7, value=round(median(p["cpu_vals"]), 1) if p["cpu_vals"] else "")
                    low_cpu = sum(1 for v in p["cpu_vals"] if v < 10)
                    ws3.cell(row=idx, column=8, value=low_cpu)

                # Totals row
                total_row = len(sorted_provs) + 2
                bold = Font(bold=True)
                ws3.cell(row=total_row, column=1, value="TOTAL").font = bold
                ws3.cell(row=total_row, column=2, value=sum(p["servers"] for p in prov_data.values())).font = bold
                total_cost = sum(p["total_cost"] for p in prov_data.values())
                ws3.cell(row=total_row, column=3, value=round(total_cost, 2)).font = bold
                ws3.cell(row=total_row, column=4, value=round(total_cost * 12, 2)).font = bold
                ws3.cell(row=total_row, column=5, value=sum(p["decomm"] for p in prov_data.values())).font = bold
                total_savings = sum(p["savings"] for p in prov_data.values())
                ws3.cell(row=total_row, column=6, value=round(total_savings, 2)).font = bold

                ws3.freeze_panes = "A2"
                _auto_width(ws3, h3)

                # Save
                if not output_dir:
                    output_dir = os.path.expanduser("~/Downloads")
                ts = datetime.now().strftime("%Y%m%d_%H%M")
                filename = f"zabbix_infra_report_{ts}.xlsx"
                filepath = os.path.join(output_dir, filename)
                wb.save(filepath)

                # Summary
                total_cost = sum(p["total_cost"] for p in prov_data.values())
                total_savings = sum(p["savings"] for p in prov_data.values())
                costed_servers = sum(1 for r in rows if r["Cost/Month ($)"])

                parts = [
                    f"**Infrastructure Report Generated**",
                    f"",
                    f"**File:** `{filepath}`",
                    f"**Total servers:** {len(rows)} ({costed_servers} with cost data)",
                    f"**Underloaded (CPU <{100-cpu_idle_threshold:.0f}%):** {len(unused_rows)}",
                    f"**Providers:** {len(prov_data)}",
                ]
                if total_cost:
                    parts.append(f"**Total cost:** ${total_cost:,.2f}/month (${total_cost * 12:,.2f}/year)")
                if total_savings:
                    parts.append(f"**Potential savings:** ${total_savings:,.2f}/month from underloaded servers")
                parts.extend([
                    f"",
                    f"### Sheets",
                    f"1. **Apps & Infra** — {len(rows)} servers × {len(h1)} columns",
                    f"2. **Unused & Underloaded** — {len(unused_rows)} decommission candidates",
                    f"3. **Provider Summary** — {len(prov_data)} providers with costs and savings",
                ])

                if costed_servers == 0:
                    parts.extend([
                        f"",
                        f"*No cost data found. Set `{{$COST_MONTH}}` macro on hosts "
                        f"or use `import_server_costs` to bulk-import from a spreadsheet.*",
                    ])

                if unused_rows:
                    parts.append(f"\n### Top Decommission Candidates")
                    for r in unused_rows[:5]:
                        cost_str = f" — ${r['Cost/Month ($)']}/mo" if r.get("Cost/Month ($)") else ""
                        parts.append(
                            f"- **{r['Resource']}** ({r['Provider']}/{r['Product']}) "
                            f"— CPU {r.get('CPU Used %', '?')}%{cost_str}"
                        )

                return "\n".join(parts)
            except (httpx.HTTPError, ValueError) as e:
                return f"Error generating report: {e}"
