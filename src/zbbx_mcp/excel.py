"""Shared Excel formatting utilities for all report modules.

openpyxl is imported lazily — only when Excel functions are actually called.
This saves ~15MB RAM for tools that never generate Excel reports.
"""

from __future__ import annotations

__all__ = [
    "BW_MAX", "BW_RED", "BW_ORANGE", "BW_GREEN",
    "classify_bandwidth", "write_headers", "write_data_rows",
    "finalize_sheet", "auto_width",
]


def _init_openpyxl():
    """Lazy-initialize openpyxl styles. Called once on first Excel operation."""
    global HEADER_FONT, HEADER_FILL, BOLD_FONT
    global RED_FILL, DARK_RED_FILL, DARK_RED_FONT, ORANGE_FILL
    global GREEN_FILL, LIGHT_GREEN_FILL, THIN_BORDER
    global _openpyxl_loaded

    if _openpyxl_loaded:
        return

    from openpyxl.styles import Border, Font, PatternFill, Side

    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    BOLD_FONT = Font(bold=True)
    RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    DARK_RED_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    DARK_RED_FONT = Font(color="FFFFFF", bold=True)
    ORANGE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    LIGHT_GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))
    _openpyxl_loaded = True


_openpyxl_loaded = False
# Placeholders — replaced by _init_openpyxl() on first use
HEADER_FONT = HEADER_FILL = BOLD_FONT = None
RED_FILL = DARK_RED_FILL = DARK_RED_FONT = ORANGE_FILL = None
GREEN_FILL = LIGHT_GREEN_FILL = THIN_BORDER = None



BW_MAX = 800    # Mbps – practical NIC limit
BW_RED = 650    # Mbps – near saturation
BW_ORANGE = 500  # Mbps – high utilization
BW_GREEN = 200   # Mbps – normal


def classify_bandwidth(mbps: float | None) -> str:
    """Classify traffic into a tier string."""
    if mbps is None:
        return ""
    if mbps >= BW_RED:
        return "CRITICAL"
    if mbps >= BW_ORANGE:
        return "HIGH"
    if mbps >= BW_GREEN:
        return "NORMAL"
    return "LOW"


def bandwidth_fill(mbps: float | None) -> tuple:
    """Return (fill, font) for a traffic value."""
    _init_openpyxl()
    if mbps is None:
        return None, None
    if mbps >= BW_MAX:
        return DARK_RED_FILL, DARK_RED_FONT
    if mbps >= BW_RED:
        return RED_FILL, None
    if mbps >= BW_ORANGE:
        return ORANGE_FILL, None
    if mbps >= BW_GREEN:
        return GREEN_FILL, None
    return LIGHT_GREEN_FILL, None


def cpu_fill(pct: float | None):
    """Return fill for a CPU usage value."""
    _init_openpyxl()
    if pct is None:
        return None
    if pct >= 80:
        return RED_FILL
    if pct >= 50:
        return ORANGE_FILL
    if pct < 10:
        return GREEN_FILL
    return None



def write_headers(ws, headers: list[str]) -> None:
    """Write styled header row."""
    _init_openpyxl()
    from openpyxl.styles import Alignment
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def auto_width(ws, headers: list[str], sample_rows: int = 50) -> None:
    """Auto-size columns based on header and first N data rows."""
    from openpyxl.utils import get_column_letter
    for col in range(1, len(headers) + 1):
        max_len = len(str(ws.cell(1, col).value or ""))
        for row in range(2, min(ws.max_row + 1, sample_rows + 2)):
            val = ws.cell(row, col).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 45)


def finalize_sheet(ws, headers: list[str], row_count: int) -> None:
    """Apply filters, freeze panes, and auto-width."""
    from openpyxl.utils import get_column_letter
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{row_count + 1}"
    ws.freeze_panes = "A2"
    auto_width(ws, headers)


def write_data_rows(
    ws,
    rows: list[dict],
    headers: list[str],
) -> None:
    """Write data rows with borders and row numbers."""
    for idx, r in enumerate(rows, 2):
        ws.cell(row=idx, column=1, value=idx - 1)
        for col, key in enumerate(headers[1:], 2):
            cell = ws.cell(row=idx, column=col, value=r.get(key, ""))
            cell.border = THIN_BORDER
